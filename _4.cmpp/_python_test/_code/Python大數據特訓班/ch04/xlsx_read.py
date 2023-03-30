import openpyxl
 
#  讀取檔案
workbook = openpyxl.load_workbook('test.xlsx')
# 取得第 1 個工作表
sheet = workbook.worksheets[0]

# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)

# 取得總行、列數
print(sheet.max_row, sheet.max_column)

# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()

sheet['A3'] = 'David' 
workbook.save('test.xlsx')      