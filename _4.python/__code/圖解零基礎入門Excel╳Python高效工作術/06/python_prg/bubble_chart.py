import openpyxl
from openpyxl.chart import Series, Reference, BubbleChart

wb = openpyxl.load_workbook(r"..\data\bubble_chart.xlsx")
sh = wb.active

chart = BubbleChart()
chart.style = 18
for row in range(2,sh.max_row+1):
    xvalues = Reference(sh, min_col=3, min_row=row)
    yvalues = Reference(sh, min_col=2, min_row=row)
    size = Reference(sh, min_col=4, min_row=row)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size,  title=sh.cell(row,1).value)
    chart.series.append(series)

sh.add_chart(chart, "F2")
wb.save(r"..\data\bubble_chart.xlsx")