import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("..\data\column_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=3, max_col=3, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #橫條圖
# 轉換成橫條圖之後，有部分的客戶名稱會消失，所以要指定chart整體的高度
#chart.height = 10

chart.style = 28    #1灰色、11藍色、28橙色、30黃色、37圖表的背景設定為淺灰色、45整體的背景設定為黑色
chart.title = "各客戶業績"
chart.y_axis.title = "營業額"
chart.x_axis.title = "客戶名稱"

chart.add_data(data,titles_from_data=True)  #以當月業績作為圖例
chart.set_categories(labels)
sh.add_chart(chart, "E3")

wb.save("..\data\column_chart.xlsx")