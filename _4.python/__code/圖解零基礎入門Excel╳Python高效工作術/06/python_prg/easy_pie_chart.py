import openpyxl
from openpyxl.chart import PieChart, Reference

wb = openpyxl.load_workbook("..\data\pie_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "D3")
wb.save("..\data\pie_chart.xlsx")