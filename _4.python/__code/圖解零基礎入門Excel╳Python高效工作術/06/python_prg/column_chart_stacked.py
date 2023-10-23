import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("..\data\column_chart_stacked.xlsx")
sh = wb.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.overlap = 100
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
wb.save("..\data\column_chart_stacked.xlsx")