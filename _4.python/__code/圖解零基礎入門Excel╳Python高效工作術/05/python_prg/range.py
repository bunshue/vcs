import openpyxl


wb = openpyxl.load_workbook(r"..\data\range.xlsx")
sheet = wb.active

getted_list = []
for row in sheet:
    for cell in row:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for row in range(2, sheet.max_row+1):
    for col in range(2,sheet.max_column+1):
        getted_list.append(sheet.cell(row,col).value)

print(getted_list)

getted_list = []
for rows in sheet["B2":"C3"]:
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for rows in sheet.iter_rows(min_row=2, min_col=2, max_row=3, max_col=3):
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)