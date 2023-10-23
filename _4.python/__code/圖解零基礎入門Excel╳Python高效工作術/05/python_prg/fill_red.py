import openpyxl
import random
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


wb = openpyxl.Workbook()
sh = wb.active
values = random.sample(range(50,150), 10)
for i, value in enumerate(values):
    sh.cell(i + 1, 1 ).value = value


less_than_rule = CellIsRule( 
    operator="lessThan",
    formula=[100],
    stopIfTrue=True,
    fill=PatternFill("solid", start_color="FF0000", end_color="FF0000")
)
sh.conditional_formatting.add("A1:A10", less_than_rule)

wb.save(r"..\data\fill_red.xlsx")