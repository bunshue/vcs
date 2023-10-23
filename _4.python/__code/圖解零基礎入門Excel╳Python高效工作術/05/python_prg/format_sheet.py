import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

#常數
TITLE_CELL_COLOR = "AA8866"

wb = openpyxl.load_workbook("..\data\orders_aggregate.xlsx")
sh = wb.active

#固定為第1列、第2欄
sh.freeze_panes = "C2"
#設定欄寬
col_widths = {"A":8, "B":15, "C":10, "D":10, "E":10, \
    "F":10, "G":10, "H":10}
for col_name in col_widths:
    sh.column_dimensions[col_name].width = col_widths[col_name]

#隱藏A欄
#sh.column_dimensions["A"].hidden=False

for i in range(2, sh.max_row+1):
    sh.row_dimensions[i].height = 18
    for j in range(3, sh.max_column+1):
        #千分位樣式
        sh.cell(row=i,column=j).number_format = "#,##0" 
        if j == 8:
            sh.cell(row=i,column=j).font = Font(bold=True)


#建立字體
font_header = Font(name="MS PGothic",size=12,bold=True,color="FFFFFF")

for rows in sh["A1":"H1"]:
    for cell in rows:
        #儲存格背景色
        cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
        #水平置中
        cell.alignment = Alignment(horizontal="distributed")
        #設定字型
        cell.font = font_header

side = Side(style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
for row in sh:
    for cell in row:
        cell.border = border
        #sh[cell.coordinate].border = border




wb.save("..\data\orders_aggregate_ed.xlsx")