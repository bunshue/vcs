import pathlib  # 標準函式庫
import openpyxl # 外部函式庫　pip install openpyxl
import csv      # 標準函式庫

lwb = openpyxl.Workbook()           #業績一覽表活頁簿
lsh = lwb.active                    #業績一覽工作表
list_row = 1
path = pathlib.Path("data\sales")    #指定相對路徑
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(9,19):
                if sh.cell(dt_row, 2).value != None:
                    #更便於說明的代碼
                    #lsh.cell(row=list_row, column=1).value = \
                    #    sh.cell(row=2, column=7).value   #傳票NO
                    lsh.cell(list_row, 1).value = sh.cell(2, 7).value   #傳票NO
                    lsh.cell(list_row, 2).value = sh.cell(3, 7).value   #日期
                    lsh.cell(list_row, 3).value = sh.cell(4, 3).value   #客戶代碼
                    lsh.cell(list_row, 4).value = sh.cell(3, 2).value.strip("敬啟")   #客戶名稱
                    lsh.cell(list_row, 5).value = sh.cell(7, 8).value   #負責人代碼
                    lsh.cell(list_row, 6).value = sh.cell(7, 7).value   #負責人姓名
                    lsh.cell(list_row, 7).value = sh.cell(dt_row, 1).value #No                    
                    lsh.cell(list_row, 8).value = sh.cell(dt_row, 2).value #商品代碼 
                    lsh.cell(list_row, 9).value = sh.cell(dt_row, 3).value #商品名稱
                    lsh.cell(list_row, 10).value = sh.cell(dt_row, 4).value #數量
                    lsh.cell(list_row, 11).value = sh.cell(dt_row, 5).value #單價
                    lsh.cell(list_row, 12).value = sh.cell(dt_row, 4).value * \
                                                sh.cell(dt_row, 5).value #金額
                    lsh.cell(list_row, 13).value = sh.cell(dt_row, 7).value #備註                                      
                    list_row += 1

lwb.save("tmp_salesList.xlsx")


