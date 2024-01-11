import openpyxl

def print_header():
    osh["A1"].value = "負責人"
    osh["B1"].value = "數量"
    osh["C1"].value = "金額"
    osh["D1"].value = "客戶"
    osh["E1"].value = "數量"
    osh["F1"].value = "金額"

wb = openpyxl.load_workbook("data\salesList.xlsx")
sh = wb.active
sales_data = {}
for row in range(1, sh.max_row + 1):
    person = sh["E" + str(row)].value
    customer = sh["C" + str(row)].value
    quantity = sh["J" + str(row)].value
    amount = sh["L" + str(row)].value
    sales_data.setdefault(person, {"name": sh["F" + str(row)].value , "quantity": 0, "amount":0})
    sales_data[person].setdefault(customer, {"name": sh["D" + str(row)].value , "quantity": 0, "amount":0})
    sales_data[person][customer]["quantity"] += int(quantity)
    sales_data[person][customer]["amount"] += int(amount)
    sales_data[person]["quantity"] += int(quantity)
    sales_data[person]["amount"] += int(amount)
    #print(sales_data)

owb = openpyxl.Workbook()
osh = owb.active
print_header()

row = 2
for person_data in sales_data.values():
    osh["A" + str(row)].value = person_data["name"]
    osh["B" + str(row)].value = person_data["quantity"]
    osh["C" + str(row)].value = person_data["amount"]
    for customer_data in person_data.values():
        #print(customer_data)
        if isinstance(customer_data,dict):
            for item in customer_data.values():
                osh["D" + str(row)].value = customer_data["name"]
                osh["E" + str(row)].value = customer_data["quantity"]
                osh["F" + str(row)].value = customer_data["amount"]
            row +=1 

osh["F" + str(row)].value =  "=SUM(F2:F" + str(row-1) + ")"
osh["E" + str(row)].value =  "合計"

owb.save("tmp_sales_aggregate.xlsx")
