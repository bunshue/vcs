import openpyxl

categorys = ((0,""),(10,"POLO衫"), (11,"禮服襯衫"), (12,"休閒襯衫"), \
            (13,"T恤"), (15,"開襟羊毛衫"),(16,"毛衣"),(17,"吸汗上衣"), \
            (18,"連帽T"))
sizes = ("代碼","分類名稱","S","M","L","LL","XL")
#製作二維列表
# 下面的程式不行，因為元素的列表會是相同的物件
#order_amount= [[0]*len(sizes)] * len(categorys)
order_amount= [[0]*len(sizes) for i in range(len(categorys))]

for j in range(len(sizes)):
    order_amount[0][j] = sizes[j]
for i in range(1,len(categorys)):
    order_amount[i][0] = categorys[i][0]
    order_amount[i][1] = categorys[i][1]
#print(order_amount)   

wb = openpyxl.load_workbook("data\ordersList.xlsx")
sh = wb.active
for row in range(2, sh.max_row + 1):
    category = sh["I" + str(row)].value
    size = sh["L" + str(row)].value
    amount = sh["M" + str(row)].value
    for i in range(1,len(categorys)):
        if category == order_amount[i][0]:
            for j in range(2,len(sizes)):
                if size == order_amount[0][j]:
                    order_amount[i][j] += amount


owb = openpyxl.Workbook()
osh = owb.active
row = 1
for order_row in order_amount:
    col = 1
    size_sum = 0 
    for order_col in order_row:
        osh.cell(row, col).value = order_col
        if  row > 1 and col > 2: 
            #print(order_col)
            size_sum += order_col
        col += 1
    if row == 1:
        osh.cell(row, col).value =  "合計"
    else:
        osh.cell(row, col).value =  size_sum
    row += 1

owb.save("tmp_orders_aggregate.xlsx")
