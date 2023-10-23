import pathlib 
import openpyxl 
import csv      

try:
    lwb = openpyxl.Workbook()
    lsh = lwb.active
    list_row = 1
    path = pathlib.Path("d:/038采實文化/9593/日經BP/excel_python/03/data/sales") 
    for pass_obj in path.iterdir():
        if pass_obj.match("*.xlsx"):
            wb = openpyxl.load_workbook(pass_obj)
            for sh in wb:
                for dt_row in range(9,19):
                    if sh.cell(dt_row, 2).value != None:
                        lsh.cell(list_row, 1).value = sh.cell(2, 7).value
                        lsh.cell(list_row, 2).value = sh.cell(3, 7).value
                        lsh.cell(list_row, 3).value = sh.cell(4, 3).value
                        lsh.cell(list_row, 4).value = sh.cell(7, 8).value
                        lsh.cell(list_row, 5).value = sh.cell(dt_row, 1).value                    
                        lsh.cell(list_row, 6).value = sh.cell(dt_row, 2).value 
                        lsh.cell(list_row, 7).value = sh.cell(dt_row, 3).value
                        lsh.cell(list_row, 8).value = sh.cell(dt_row, 4).value
                        lsh.cell(list_row, 9).value = sh.cell(dt_row, 5).value
                        lsh.cell(list_row, 10).value = sh.cell(dt_row, 4).value * \
                                                    sh.cell(dt_row, 5).value
                        lsh.cell(list_row, 11).value = sh.cell(dt_row, 7).value                                     
                        list_row += 1

    with open("..\data\sales\salesList.csv","w",encoding="utf_8_sig") as fp:
        writer = csv.writer(fp, lineterminator="\n")
        for row in lsh.rows:
            writer.writerow([col.value for col in row])
            
except PermissionError as ex: 
    print(ex.filename,"是Permission錯誤") 
except:
    print("出現例外了")