
print("------------------------------------------------------------")  # 60個

# 共同
import os
import sys
import math
import random
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

font_filename = "C:/_git/vcs/_1.data/______test_files1/_font/msch.ttf"
# 設定中文字型及負號正確顯示
# 設定中文字型檔
plt.rcParams["font.sans-serif"] = "Microsoft JhengHei"  # 將字體換成 Microsoft JhengHei
# 設定負號
plt.rcParams["axes.unicode_minus"] = False  # 讓負號可正常顯示
plt.rcParams["font.size"] = 12  # 設定字型大小

print("------------------------------------------------------------")  # 60個

# 索引
df = pd.DataFrame({"a":[1,3], "b":[2,4]}, index=['line1', 'line2'])
print(df.index) # 顯示行索引
print(df.columns) # 顯示列索引

print('------------------------------------------------------------')	#60個

# 多重索引
df = pd.read_excel('data/test.xlsx', header=[0,1]) # 指定前兩行爲列索引
print(df)
print(df.columns.values) # 查看列索引內容

df.columns = ['_'.join(col).strip() for col in df.columns.values] # 重置字段名
print(df)

print('------------------------------------------------------------')	#60個

# Python日期時間處理
# 時間點
from datetime import datetime

d1 = datetime.now() # 獲取當前時間
print(d1)
print(d1.year, d1.month, d1.day, d1.hour, d1.minute, d1.second)
d2 = datetime(2019, 3, 27) # 通過指定日期構造datetime
print(d2)

# 時間段
from datetime import timedelta
delta = d2-d1 # 通過時間日期相減獲取
print(type(delta))
print(delta)
delta = timedelta(days=3) # 通過指定時定差獲取
print(d1+delta)# 利用時間段計算新日期時間

# 時間戳

import time
print(time.time())

d = datetime.now()
t = time.mktime(d.timetuple()) # 從datetime格式轉換
print(t)
print(time.mktime(time.strptime("2019-03-27", "%Y-%m-%d"))) # 從字符串轉換
print(datetime.fromtimestamp(t)) 
print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(t)))


# 時間類型轉換
d = datetime.strptime('2019-03-27', '%Y-%m-%d')
print(d)

from dateutil.parser import parse
d = parse('2019/03/27')
print(d)
print(str(d))

print(d.strftime("%Y/%m/%d %H:%M:%S"))

print('------------------------------------------------------------')	#60個

# Pandas日期時間處理
# 時間點TimeStamp

t = pd.to_datetime('2019-03-01 00:00:00') # 從字符串轉換
print(type(t), t)
t = pd.to_datetime(datetime.now()) # 從datetime格式轉換
print(type(t), t)

# 時間間隔
t1 = pd.to_datetime('2019-03-01 00:00:00')
t2 = pd.to_datetime(datetime.now())
delta = t2-t1 # 通過TimeStamp相減獲取
print(type(delta), delta, delta.days, delta.seconds)

delta = pd.Timedelta(days=27) # 構造時間間隔爲27天
print(t2 + delta)

# 時間段Period
t = pd.to_datetime(datetime.now())
p = pd.Period(t, freq='H')
print(p, p.start_time, p.end_time) # 顯示時間段起止時間

# 批量轉換
arr = ['2019-03-01','2019-03-02','2019-03-03']
df = pd.DataFrame({'d':arr})
df['d'] = pd.to_datetime(df['d'])
print(df)

print('------------------------------------------------------------')	#60個

# 時間序列操作
# 時間日期類型索引
df.index = pd.to_datetime(df['d']) # 本例中使用了上例中構造的df[‘d’]
print(df.index)

df = pd.DataFrame()
df['date'] = pd.date_range(start='2017-12-30',end='2019-01-05',freq='d') # 創建時間數據
df['val'] = df['date'].apply(lambda x: x.weekday()) # 計算該日是星期幾
df.set_index('date', inplace = True) # 設置時間索引
print(df.head(3)) # 顯示前三條

# 時間段類型索引
df_period = df.to_period(freq='M') # 按月創建時間段
print(type(df_period.index)) # 查看類型
print(len(df_period)) # 查看記錄個數，與原記錄個數一致
print(df_period.head(3))

print(df_period.index[0].start_time, df_period.index[0].end_time)
print(df_period.index[1].start_time, df_period.index[1].end_time)
print(df.index.is_unique, df_period.index.is_unique)

df_dt = df_period.to_timestamp()
print(df_dt.head(3))
print(type(df_dt.index))

print('------------------------------------------------------------')	#60個

# 篩選和切分
print(df['2019'])  # 篩選2019全年數據
print(df['2019-01'])  #  篩選2019年一月全月數據
print(df['2018':'2019'].head()) # 篩選2018年初到2019年底的所有數據
print(df['2018-12-31':].head()) # 篩選2018-12-31及之後的數據

# 重採樣
tmp = df.resample('w').sum() # 使用疊加方式按周重採樣
print(tmp.head(3))

tmp = df.resample('M').ohlc() # 使用用ohlc方式按月降採樣
print(tmp.head(3))

tmp = df.resample('M').sum().to_period('M') # 按月降採樣，同時將時間變爲時間段
print(tmp.head(3))

df1 = pd.DataFrame({'val':[8,7,6]})
df1.index = pd.to_datetime(['2019-03-01','2019-03-15','2019-03-31']) # 僅含三條數據
df2 = df1.resample('D').interpolate() # 用插值方式升採樣
print(len(df2))
print(df2.head(3))

df3 = df1.asfreq('D')
print(df3.head(3))

# 偏移
df['prev'] = df['val'].shift() # 取前一條數據的val值作爲當前記錄中prev字段的值
print(df.head(3))

# 計算滑動窗口
df['sw'] = df['val'].rolling(window=3).mean() # 計算窗口中數據的均值
print(df.head(3))

df['emw_3'] = df['val'].ewm(span=3).mean()
df['emw_7'] = df['val'].ewm(span=7).mean()
df['rolling'] = df['val'].rolling(7).mean()

print('------------------------------------------------------------')	#60個

# 時區轉換

import pytz
print('時區個數 :', len(pytz.common_timezones))
print('前3個 :', pytz.common_timezones[:3])

import datetime
t = datetime.datetime.now()
print(t)

utc_dt = pytz.utc.localize(t)
print(utc_dt)

from pytz import timezone
tz = timezone('Asia/Shanghai') # 將時區設爲上海
print(utc_dt.astimezone(tz)) # 轉換時區

df = pd.DataFrame()
df['date'] = pd.date_range(start='2018-12-31',end='2019-01-01',freq='d')
df.set_index('date', inplace=True) # 設置時間索引
print(df.index)

df.index = df.index.tz_localize('UTC')
print(df.index.values, df.index)

df.index = df.index.tz_convert('Asia/Shanghai')
print(df.index.values)
print(df.index)

print('------------------------------------------------------------')	#60個

# 數據重排
# 數據錶轉置
df = pd.DataFrame({"a":[1,2],"b":[3,4]}, index=['l1','l2'])
print(df)
print(df.T)

# 行轉列和列轉行
df1 = df.stack() # 列轉行
print(df1)

print(df1.unstack()) # 將內層行索引轉爲列索引
print(df1.unstack(level=0)) # 將外層行索引轉爲列索引

# 透視轉換
df = pd.DataFrame({"時間":['期中','期末','期中','期末'],
                   "學科":['語文','語文','數學','數學'],
                   "分數":[89,75,90,95]})
df1 = df.pivot(index='時間', columns='學科', values='分數')
print(df, df1)


print("------------------------------------------------------------")  # 60個

import seaborn as sns
import statsmodels.api as sm # 示例使用了statsmodels庫中的自帶的數據
import matplotlib as mpl

sns.set(style='darkgrid',color_codes=True) # 帶灰色網格的背景風格
tips=sns.load_dataset('tips')  # 示例中的基本數據

# 4.2.2 連續變量相關圖
# Relplot關係類型圖表
sns.relplot(x="total_bill", y="tip", hue="day",col="time", row="sex", data=tips)

plt.show()

print('------------------------------------------------------------')	#60個



# 點圖
sns.scatterplot(x="total_bill", y="tip", hue="size", size="size", data=tips)

plt.show()

print('------------------------------------------------------------')	#60個

# 線圖
sns.lineplot(x="tip", y="total_bill", hue="sex", style="sex", data=tips)

plt.show()

print('------------------------------------------------------------')	#60個

# 4.2.3 分類變量圖
# stripplot散點圖
sns.stripplot(x='day', y='total_bill', data=tips, jitter=True)

plt.show()

print('------------------------------------------------------------')	#60個

# swarmplot散點圖
sns.swarmplot(x='day',y='total_bill',data=tips)

plt.show()

print('------------------------------------------------------------')	#60個

# violinplot小提琴圖
sns.violinplot(x="day", y="total_bill", hue="sex", split=True, data=tips)

plt.show()

print('------------------------------------------------------------')	#60個

# boxplot箱式圖
sns.boxplot(x="day", y="total_bill", hue="sex", data=tips);

plt.show()

print('------------------------------------------------------------')	#60個

# boxenplot變種箱式圖
sns.boxenplot(x="day", y="total_bill", hue="sex", data=tips)

plt.show()

print('------------------------------------------------------------')	#60個

# pointplot分類統計圖
sns.pointplot(x="sex", y="total_bill", hue="smoker", data=tips,
palette={"Yes": "g", "No": "m"},
markers=["^", "o"], linestyles=["-", "--"]);

plt.show()

print('------------------------------------------------------------')	#60個

# barplot柱對比圖
sns.barplot(x='smoker',y='total_bill',hue='sex',data=tips)

plt.show()

print('------------------------------------------------------------')	#60個


# 4.2.4 迴歸圖
# 連續變量回歸圖
sns.lmplot(x="total_bill", y="tip", data=tips)

plt.show()

print('------------------------------------------------------------')	#60個


# 分類變量回歸圖
sns.lmplot(x="size", y="total_bill", data=tips, x_estimator=np.mean)

plt.show()

print('------------------------------------------------------------')	#60個


""" fail
# 4.2.5 多圖組合
# jointplot兩變量圖

import statsmodels.api as sm
import seaborn as sns

sns.set(style="darkgrid")
data = sm.datasets.ccard.load_pandas().data
g = sns.jointplot('AVGEXP', 'AGE', data=data, kind="reg",
                 xlim=(0, 1000), ylim=(0, 50), color="m")
plt.show()

print('------------------------------------------------------------')	#60個

"""

# pairplot多變量圖
data = sm.datasets.ccard.load_pandas().data
sns.pairplot(data, vars=['AGE','INCOME', 'INCOMESQ','OWNRENT'])

plt.show()


print('------------------------------------------------------------')	#60個

""" fail
# factorplot兩變量關係圖
data = sm.datasets.fair.load_pandas().data
sns.factorplot(x='occupation', y='affairs', hue='religious', data=data)

plt.show()
"""

print('------------------------------------------------------------')	#60個

""" fail
# FacetGrid結構化繪圖網格
g = sns.FacetGrid(tips, col = 'time', row = 'smoker') # 按行和列的分類做N個圖
g.map(plt.hist, 'total_bill', bins = 10) # 指定做圖方式

plt.show()
"""

print('------------------------------------------------------------')	#60個


# 4.2.6 熱力圖
data = sns.load_dataset('planets')
corr=data[['number','orbital_period','mass','distance']].corr(method='pearson')
sns.heatmap(corr, cmap="YlGnBu") 

plt.show()

print('------------------------------------------------------------')	#60個


""" fail
# 印刷品作圖

sns.set_style("whitegrid")

with sns.cubehelix_palette(start=2.7, rot=0, dark=.5, light=.8, 
          reverse=True, n_colors=5):
    # 此處放置具體繪圖函數
    sns.stripplot(x='day', y='total_bill', data=tips, jitter=True)

plt.show()
"""

print('------------------------------------------------------------')	#60個

""" fail

# 準備工作

import pyecharts

attr = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
v1 = [2.0, 4.9, 7.0, 23.2, 25.6, 76.7, 135.6, 162.2, 32.6, 20.0, 6.4, 3.3]
v2 = [2.6, 5.9, 9.0, 26.4, 28.7, 70.7, 175.6, 182.2, 48.7, 18.8, 6.0, 2.3]

# 4.3.3 繪製交互圖
# 柱圖
bar = pyecharts.Bar("Title1", "Title2")
bar.add("v1", attr, v1, mark_line=["average"], mark_point=["max", "min"])
bar.add("v2", attr, v2, mark_line=["average"], mark_point=["max", "min"])
bar.render('test.html')
bar

"""

print('------------------------------------------------------------')	#60個
'''

import logging

# 獲取logger對象,取名mylog
logger = logging.getLogger("mylog")
# 輸出DEBUG及以上級別的信息，針對所有輸出的第一層過濾
logger.setLevel(level=logging.DEBUG)

# 獲取文件日誌句柄並設置日誌級別，第二層過濾
handler = logging.FileHandler("log.txt")
handler.setLevel(logging.INFO)	

# 生成並設置文件日誌格式，其中name爲上面設置的mylog
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# 獲取流句柄並設置日誌級別，第二層過濾
console = logging.StreamHandler()
console.setLevel(logging.WARNING)

# 爲logger對象添加句柄
logger.addHandler(handler)
logger.addHandler(console)

# 記錄日誌
logger.info("show info")
logger.debug("show debug")
logger.warning("show warning")


print('------------------------------------------------------------')	#60個

print('讀寫XML文件')

from xml.dom import minidom

dom=minidom.Document()
root_node=dom.createElement('root') # 創建根節點
dom.appendChild(root_node) # 添加根節點

book_node=dom.createElement('blog') # 創建第一個子節點
book_node.setAttribute('level','3') # 添加屬性
root_node.appendChild(book_node) # 爲root添加子節點

name_node=dom.createElement('addr') # 創建第二個子節點
name_text=dom.createTextNode('https://blog.csdn.net/xieyan0811') # 添加文字
name_node.appendChild(name_text)
root_node.appendChild(name_node)

# toxml() 轉換成字符串, toprettyxml()轉換成樹形縮進版式
print(dom.toprettyxml())
with open('test_dom.xml','w') as fh:
    dom.writexml(fh, indent='',addindent='\t', newl='\n', encoding='UTF-8')

print('------------------------------------------------------------')	#60個


from xml.dom import minidom
with open('test_dom.xml','r') as fh:
    dom = minidom.parse(fh) # 獲取dom對象
    root = dom.documentElement # 獲取根節點
    print("node name", root.nodeName) # 顯示節點名: root
    print("node type", root.nodeType) # 顯示節點類型
    print("child nodes", root.childNodes) # 列出所有子節點
    blog = root.getElementsByTagName('blog')[0] # 根據標籤名獲取元素列表
    print(blog.getAttribute('level')) # 獲取屬性值
    addr = root.getElementsByTagName('addr')[0]
    print("addr's child nodes", addr.childNodes)
    text_node = addr.childNodes[0] # 獲取文本節點內容
    print("text data", text_node.data)
    print("parent", addr.parentNode.nodeName) # 顯示name的父節點名稱


print('------------------------------------------------------------')	#60個

print('讀寫Json文件')

import json

data = [{"group":0,"param":["one","two","three"]},
        {"group":1,"param":["1","2","3"]}] 

jsonstr = json.dumps(data)
print(jsonstr)
jsonstr = json.dumps(data, sort_keys=True, 
                 indent=4, separators=(',', ': '))
print(jsonstr)
data1 = json.loads(jsonstr)
print(data1, type(data1))

with open('json.txt','w') as json_file:
    json.dump(data, json_file)
    json_file.close()

with open('json.txt','r') as json_file:
    data = json.load(json_file)
    json_file.close()
print(data1, type(data1))

print('------------------------------------------------------------')	#60個

print('讀寫CSV文件')

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
print(df.info()) # 顯示dataframe相關信息
df.to_csv("tmp.csv", index=False, header=True, columns=['Name','Sex','Age'])

df1 = pd.read_csv("tmp.csv")
print(df1.info())
print(df1)

print('------------------------------------------------------------')	#60個

print('讀寫PKL文件')

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
print(df.info())
df.to_pickle("tmp.pkl")

df1 = pd.read_pickle("tmp.pkl")
print(df1.info())


print('------------------------------------------------------------')	#60個

import pickle
data1 = {'a': [1, 2.0, 4+6j],
         'b': ('string1', u'Unicode string'),
         'c': None}
output = open('tmp2.pkl', 'wb')
pickle.dump(data1, output)
output.close()

pkl_file = open('tmp2.pkl', 'rb')
data2 = pickle.load(pkl_file)
print(data2)
pkl_file.close()




print('------------------------------------------------------------')	#60個

"""
from sklearn import svm
from sklearn import datasets
from sklearn.externals import joblib

clf = svm.SVC()
iris = datasets.load_iris()
clf.fit(iris.data, iris.target)
joblib.dump(clf, "tmp3.pkl")

clf1 = joblib.load("tmp3.pkl")
print(clf1.predict(iris.data[:2]))
"""

print('------------------------------------------------------------')	#60個

print('讀寫HDF5文件')

import h5py

f = h5py.File('tmp.h5','w')
f['data'] = np.zeros((3,3))
f['labels'] = np.array([1,2,3,4,5])
f.close()

f = h5py.File('tmp.h5','r')
for key in f.keys():
    print(f[key].name)
    print(f[key].shape)
    #print(f[key].value)
f.close()

print('------------------------------------------------------------')	#60個

print('讀寫Excel文件')

import openpyxl

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
df.to_excel("tmp.xlsx")

df1 = pd.read_excel("tmp.xlsx")
print(df1)

wb = openpyxl.load_workbook('tmp.xlsx')
sheets = wb.sheetnames
print(sheets)
for i in range(len(sheets)):
    sheet = wb[sheets[i]]
    print('title', sheet.title)
    for col in sheet.iter_cols(min_row=0, min_col=0, max_row=3, max_col=3):
        for cell in col:
            print(cell.value)


print('------------------------------------------------------------')	#60個


print('Python存取Sqlite數據庫')

import sqlite3

conn = sqlite3.connect('test.db')
c = conn.cursor()
c.execute("""CREATE TABLE IF NOT EXISTS TIPS 
       (NAME           TEXT    NOT NULL,
       ADDRESS        CHAR(50),
       BILL         REAL);""") # 創建數據表
c.execute("INSERT INTO TIPS (NAME,ADDRESS,BILL) \
      VALUES ('Zhang', 'Beijing', 1004.00 )"); # 向表中輸入數據
cursor = c.execute("SELECT * from TIPS")
for row in cursor:
    print(row)
conn.commit()
conn.close()

print('------------------------------------------------------------')	#60個

print('抓取網站數據')

import urllib.request
from bs4 import BeautifulSoup

response = urllib.request.urlopen("https://blog.csdn.net/xieyan0811")
html = response.read().decode("utf-8","ignore") # 返回網頁爲utf-8編碼，解碼時忽略錯誤
reg = r'http://'  
soup = BeautifulSoup(html, 'html.parser')
for link in soup.find_all('a'):
    addr = link.get('href')
    # 顯示包含關鍵字的所有地址
    if addr != None and addr.find('xieyan0811/article') != -1: 
        print(addr)
        

print('------------------------------------------------------------')	#60個

print('使用POST方式抓取數據')

import requests
params = {'key1': 'value1', 'key2': 'value2'}
r = requests.post("http://httpbin.org/post", data=params)
print(r.text)


print('------------------------------------------------------------')	#60個

print('基本類型轉換')

dic = {
     'string': ['dog', 'snake', 'cat', 'dog', 'monkey', 'elephant'],
     'integer': [2000, 2000, 2001, 2002, 2003, np.nan],
     'float': [1.5, 1.5, 1.7, np.nan, np.nan, 8.3],
     'dtime': ['2018-01-01', '2018/01/02', '2018-01-03', '2018-01-04', '2018-01-05', np.nan],
     'mix': [1, 1, 0, '+', 0, 1],
     'classify': ['A', 'B', 'A', 'B', 'A', 'A']
                }
data = pd.DataFrame(dic)
print(data.dtypes)


data['dtime'] = pd.to_datetime(data['dtime'], infer_datetime_format=True)
data['mix']=pd.to_numeric(data['mix'],errors='coerce')
data['classify']=pd.Categorical(data['classify'])
data['float']=data['float'].astype(np.float32)
print(data.dtypes)



print('------------------------------------------------------------')	#60個

print('缺失值處理')

dic = {   
     'state': ['Ohio', 'Ohio', 'Ohio', 'Ohio', 'Nevada', 'Nevada'],
     'year': [2000, 2000, 2001, 2002, 2003, 3456],
     'score': [1.5, 1.5, 1.7, np.nan, np.nan, 8.3],
     'desc': [np.nan, np.nan, np.nan, np.nan, np.nan, 3],
     'val1': [1, 1, 0, '+', 0, 1],
}
data = pd.DataFrame(dic)

print(data['desc'].nunique()) # 不同取值個數
print(data['desc'].unique()) # 不同取值列表
print(data['year'].value_counts()) # 不同取值出現次數



print(data['desc'].isnull()) # 是否缺失
print(data['desc'].isnull().any()) # 是否含有任意缺失
print(data['desc'].isnull().all()) # 是否全部缺失
print(data['desc'].isnull().sum(), len(data)) # 空值個數與記錄個數
print(data.dropna(axis=1, how='all'))
print(data['score'].fillna(data['score'].mean()))
print(data['score'].fillna(method='ffill', limit=1))

print(data.interpolate(mdthod='polynomial', order=2)) # 二次多項式插值
print(data.interpolate(mdthod='spline', order=3)) # 三次樣條插值


print('------------------------------------------------------------')	#60個


print('異常值處理')

print(data.query('year<2050'))
print(data[data['year']<2050])

data['val1'] = data['val1'].apply(lambda x: 1 if x == '+' else x)


print('去重處理')

print(data.drop_duplicates(keep='last'))
print(data.drop_duplicates(keep='last', subset='year'))



print('------------------------------------------------------------')	#60個

print('模型篩選特徵')

from sklearn.datasets import load_iris 
from sklearn import tree 

iris = load_iris()
clf = tree.DecisionTreeClassifier()
clf = clf.fit(iris.data, iris.target)
print(clf.feature_importances_)




print('------------------------------------------------------------')	#60個

print('數學方法降維')

from sklearn.decomposition import PCA
from sklearn import datasets
import seaborn as sns

iris = datasets.load_iris()
data = pd.DataFrame(iris.data, columns=['SpealLen', 'SpealWid', 
                             'PetalLen', 'PetalWid'])
mat = data.corr()
sns.heatmap(mat, annot=True, vmax=1, vmin=-1, xticklabels= True, 
            yticklabels= True, square=True, cmap="gray")

plt.show()

print('------------------------------------------------------------')	#60個

pca = PCA(n_components=2)
data1 = pca.fit_transform(data)
print(data1.shape)
print(pca.explained_variance_ratio_, 
      pca.explained_variance_ratio_.sum())
plt.scatter(data1[:,0], data1[:,1], c = np.array(iris.target), 
            cmap=plt.cm.copper)


plt.show()


print('------------------------------------------------------------')	#60個

print('簡單隨機抽樣')

import statsmodels.api as sm

data = sm.datasets.anes96.load_pandas().data
df = data.sample(50)
print(df.head())

print('系統抽樣')

index_list = [i for i in range(len(data)) if i % 10 == 0]
df = data.iloc[index_list]
print(df.head())

print('分層抽樣')

def typicalSampling(grp, typicalFracDict):
    name = grp.name
    frac = typicalFracDict[name]
    return grp.sample(frac=frac)

typicalFracDict = {
    0.0: 0.35,  
    1.0: 0.5,  
}
df = data.groupby('vote').apply(typicalSampling, typicalFracDict)
print(df.head())

print('整羣抽樣')
unique = np.unique(data['income'])
sample = random.sample(list(unique),2)
df = pd.DataFrame()
for label in sample:
    tmp = data[data['income']==label]  
    df = pd.concat([df, tmp])
print(df.head())

print('------------------------------------------------------------')	#60個

print('merge方法')

df1 = pd.DataFrame({'id': [1, 2, 3], 'val1': [2, 4, 6]})
df2 = pd.DataFrame({'id': [3, 2, 2], 'val2': [9, 6, 5]})
print(pd.merge(df1, df2, how='left'))

print('------------------------------------------------------------')	#60個

print('concat方法')

df1 = pd.DataFrame({'id': [1, 2, 3], 'val1': [2, 4, 6]})
df2 = pd.DataFrame({'id': [3, 2, 2], 'val2': [9, 6, 5]})
print(pd.concat([df1, df2]))
print(pd.concat([df1, df2], axis=1))

print('------------------------------------------------------------')	#60個

print('數值型特徵')

dic = {'height': [1.6, 1.7, 1.8],
      'weight': [60, 70, 90]}
data = pd.DataFrame(dic)
data['bmi'] = data['weight'] / (data['height'] **2)
print(data)
data['overweight'] = data['bmi'] > 25
print(data)
data['overweight'] = data['overweight'].map({True:'Yes', False:'No'})
print(data)

print('------------------------------------------------------------')	#60個

print('類型特徵')

dic = {'string': ['第一組', '第二組', '第二組']}
data = pd.DataFrame(dic)
print(pd.factorize(data.string)) # 轉換成數值型編碼

data['num'] = pd.factorize(data['string'])[0]
df = pd.get_dummies(data['string'], prefix='組別')  # 轉換成onehot類型編碼
new_data = pd.concat([data, df], axis=1)
print(new_data)

print('------------------------------------------------------------')	#60個

print('關鍵字特徵')

from scipy import stats
import jieba
import re

def do_split(test_text):
    pattern = r',|\.|/|;|\'|`|\[|\]|<|>|\?|:|"|\{|\}|\~|!|？|@|#|\$|%|\^|&|\(|\)|-|=|\_|\+|，|。|、|；|‘|’|【|】03   |·|！| |…|（|）' 
    return re.split(pattern, test_text) 



def get_keywords(data, feat):
    ret = []
    data[feat] = data[feat].apply(lambda x: x.strip())
    for i in data[feat].unique():
        # 將短句作爲關鍵字
        if len(i) <= 50 and i not in ret:
            ret.append(i)
        # 將子句作爲關鍵字
        for sentence in do_split(i):
            if len(sentence) <= 50 and sentence not in ret:
                ret.append(sentence)
        # 將詞作爲關鍵字
        for word in jieba.cut(i, cut_all=True):
            if len(word) > 1 and word not in ret:
                ret.append(word)
    return ret

def check_freq(data, feat, keywords, limit):
    ret = []
    for key in keywords:
        try:
            if len(data[data[feat].str.contains(key)]) > limit:
                ret.append(key)
        except:
            pass
    return ret

def do_test(data, feat, key, y, debug=False):
    arr1 = data[data[feat].str.contains(key) == True][y]
    arr2 = data[data[feat].str.contains(key) == False][y]
    ret1 = stats.ttest_ind(arr1, arr2, equal_var = False)
    ret2 = stats.levene(arr1, arr2)
    if ret1.pvalue < 0.05 or ret2.pvalue < 0.05:
        return True    
    return False

def check(data, feat, y):
    ret = []
    keywords = get_keywords(data, feat)
    arr = check_freq(data, feat, keywords, 5)
    for word in arr:
        if do_test(data, feat, word, y):
            ret.append(word)
    return ret

# 讀取數據文件的前500條數據，其中第6個字段是微博內容，第5個字段爲點贊次數。
data = pd.read_csv('data/weibo_train_data.txt', sep='\t', 
                   header=None, nrows=500)
print(check(data, 6, 5))


print('------------------------------------------------------------')	#60個

from scipy import stats

arr1 = [96,95,95,95,95,95,95,95,95,95,95,95,95,95,95,
        95,95,95,95,95,95,95,95,95,95,95,95,95,95,95]
arr2 = [90,91,92,93,94,90,91,92,93,94,90,91,92,93,94,
        90,91,92,93,94,90,91,92,93,94,90,91,92,93,94]
print(stats.ttest_1samp(arr1, 92))
print(stats.ttest_1samp(arr2, 92))
print((np.mean(arr1)-92)/(np.std(arr1)/np.sqrt(len(arr1)-1)))
print((np.mean(arr2)-92)/(np.std(arr2)/np.sqrt(len(arr2)-1)))

print('------------------------------------------------------------')	#60個

from scipy import stats


print('正態性檢驗')

# 檢驗樣本是否服從某一分佈
np.random.seed(12345678)
x = stats.norm.rvs(loc=0, scale=1, size=300) # loc爲均值，scale爲方差
print(stats.kstest(x,'norm'))

# 數據的正態性檢驗
np.random.seed(12345678)
x = stats.norm.rvs(loc=10, scale=2, size=70) 
print(stats.shapiro(x))

# 作圖法檢驗正態分佈
np.random.seed(12345678)
x = stats.norm.rvs(loc=10, scale=2, size=300) 
plt.hist(x)
plt.show()

print('------------------------------------------------------------')	#60個

print('方差齊性檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500)  
rvs2 = stats.norm.rvs(loc=25,scale=9,size=500)
print(stats.levene(rvs1, rvs2))


print('------------------------------------------------------------')	#60個

from scipy import stats

print('兩獨立樣本T檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500)  
rvs2 = stats.norm.rvs(loc=6,scale=10,size=500)
print(stats.ttest_ind(rvs1,rvs2))


print('配對樣本T檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500) 
rvs2 = (stats.norm.rvs(loc=5,scale=10,size=500) + stats.norm.rvs(scale=0.2,size=500)) 
print(stats.ttest_rel(rvs1,rvs2))



print('------------------------------------------------------------')	#60個


from scipy import stats

a = [47,56,46,56,48,48,57,56,45,57]  # 分組1
b = [87,85,99,85,79,81,82,78,85,91]  # 分組2
c = [29,31,36,27,29,30,29,36,36,33]  # 分組3
print(stats.f_oneway(a,b,c))


print('------------------------------------------------------------')	#60個

from scipy import stats

A = [6, 15, 22, 36, 40, 48, 53]
B = [3, 4, 5, 12, 17, 18, 21]
print(stats.ranksums(A, B))

C = [1, 2, 3, 4, 5, 6, 7]
print(stats.kruskal(A, B, C))


print('------------------------------------------------------------')	#60個

import statsmodels.api as sm 
import scipy.stats as stats

data = sm.datasets.anes96.load_pandas().data
contingency = pd.crosstab(data['vote'], [data['educ']])
print(stats.chi2_contingency(contingency)) # 卡方檢驗

print('------------------------------------------------------------')	#60個

from scipy import stats

print('圖形描述相關性')
import statsmodels.api as sm
data = sm.datasets.ccard.load_pandas().data
plt.scatter(data['INCOMESQ'], data['INCOME'])

plt.show()

print('正態資料的相關分析')
np.random.seed(12345678)
a = np.random.normal(0,1,100)
b = np.random.normal(2,2,100)
print(stats.pearsonr(a, b))


print('非正態資料的相關分析')
print(stats.spearmanr([1,2,3,4,5], [1,6,7,8,20]))


print('------------------------------------------------------------')	#60個

print('多元線性迴歸')
import statsmodels.api as sm 

data = sm.datasets.ccard.load_pandas().data
model = sm.OLS(endog = data['AVGEXP'],
     exog = data[['AGE','INCOME','INCOMESQ','OWNRENT']]).fit()
print(model.summary())

print('------------------------------------------------------------')	#60個

# 邏輯迴歸

import statsmodels.api as sm
data = sm.datasets.ccard.load_pandas().data
data['OWNRENT'] = data['OWNRENT'].astype(int)
model = sm.Logit(endog = data['OWNRENT'], 
     exog = data[['AVGEXP','AGE','INCOME','INCOMESQ']]).fit()
print(model.summary())


print('------------------------------------------------------------')	#60個

import statsmodels as sm
import tableone 

data = sm.datasets.anes96.load_pandas().data
categorical = ['TVnews', 'selfLR', 'ClinLR', 'educ', 'income'] 
groupby = 'vote'
mytable = tableone.TableOne(data, categorical=categorical, 
                            groupby=groupby, pval=True)
print(mytable)
mytable.to_excel("a.xlsx")

print('------------------------------------------------------------')	#60個

print('方差、協方差、協方差矩陣')

# 數據準備
df = pd.DataFrame({'身高':[1.7, 1.8, 1.65, 1.75, 1.8], 
                   '體重':[140, 170, 135,  150,  200]})
print(df)

print('均值')
print(df['身高'].mean())

print('方差')
print(df['身高'].var())
print((sum((df['身高']-df['身高'].mean())**2))/(len(df)-1))

print('標準差')
print(df['身高'].std())

print('協方差')
print((sum((df['體重']-df['體重'].mean())*(df['身高']-df['身高'].mean()))/(len(df)-1)))

print('協方差矩陣')
print(df.cov())

print('相關係數和相關係數矩陣')
print(df.corr())

print('------------------------------------------------------------')	#60個

print('距離與範數')

from scipy.spatial.distance import pdist  # 導入科學計算庫中的距離計算工具

df = pd.DataFrame({'身高':[1.7, 1.8, 1.65, 1.75, 1.8], 
                   '體重':[140, 170, 135,  150,  200]})
x = df.loc[0,:]  # 取第一條實例x
print(x)
y = df.loc[1,:]  # 取第二條實例y
print(y)

print('歐氏距離')
d1 = np.sqrt(np.sum(np.square(x-y))) # 公式計算
d2 = pdist([x,y])  # 調用距離函數
print(d1, d2)

print('曼哈頓距離')
d1 = np.sum(np.abs(x-y))
d2 = pdist([x,y],'cityblock')
print(d1, d2)

print('海明距離')
d1 = pdist([x,y], 'hamming') 
d2 = pdist([[0,0,0,1],[0,0,0,8]], 'hamming') # 對比兩數組的海明距離
print(d1, d2)

print('閔氏距離')
d1=np.sqrt(np.sum(np.square(x-y)))
d2=pdist([x,y],'minkowski',p=2) # 求取p=2時的閔氏距離
print(d1, d2)

print('切比雪夫距離')
d1 = np.max(np.abs(x-y))
d2 = pdist([x,y],'chebyshev')
print(d1, d2)

print('馬氏距離')
delta = x-y
S=df.cov()   #協方差矩陣
SI = np.linalg.inv(S) #協方差矩陣的逆矩陣
d1=np.sqrt(np.dot(np.dot(delta,SI),delta.T))
d2=pdist([x,y], 'mahalanobis', VI=SI)
print(d1, d2)

print('------------------------------------------------------------')	#60個

print('迴歸效果評估')


print('MSE均方誤差')

from sklearn.metrics import mean_squared_error
y_true = [1, 1.25, 2.37]
y_pred = [1, 1, 2]
print(mean_squared_error(y_true, y_pred))

print('MAE平均絕對誤差')
from sklearn.metrics import mean_absolute_error
y_true = [1, 1.25, 2.37]
y_pred = [1, 1, 2]
print(mean_absolute_error(y_true, y_pred))

print('R-Squared擬合度')
from sklearn.metrics import r2_score
y_true = [1, 1.25, 2.37]
y_pred = [1, 1, 2]
print(r2_score(y_true,y_pred))

print('------------------------------------------------------------')	#60個

print('分類效果評估')
print('FP/FN/TP/TN')

y_pred = [0, 0, 0, 1, 1, 1, 0, 1, 0, 0]  # 預測值
y_real = [0, 1, 1, 1, 1, 1, 0, 0, 0, 0]  # 實際值

from sklearn.metrics import confusion_matrix
cm = confusion_matrix(y_real, y_pred)
tn, fp, fn, tp = cm.ravel()
print("tn", tn, "fp", fp, "fn", fn, "tp", tp)

print('準確率')
from sklearn.metrics import accuracy_score
print(accuracy_score(y_real, y_pred))

print('召回率')
from sklearn.metrics import recall_score
print(recall_score(y_real, y_pred))

print('精度')
from sklearn.metrics import precision_score
print(precision_score(y_real, y_pred))

print('F值')

from sklearn.metrics import f1_score
from sklearn.metrics import fbeta_score

print(f1_score(y_real, y_pred))  # 計算f1
print(fbeta_score(y_real, y_pred, beta=2)) # 計算fn

print('Logloss')
from sklearn.metrics import log_loss
y_real = [0, 1, 1, 1, 1, 1, 0, 0, 0, 0]
y_score=[0.9, 0.75, 0.86, 0.47, 0.55, 0.56, 0.74, 0.22, 0.5, 0.26]
print(log_loss(y_real,y_score))

print('ROC曲線和AUC')
from sklearn.metrics import roc_auc_score, roc_curve

print(roc_auc_score(y_real, y_score)) # AUC值

fpr, tpr, thresholds = roc_curve(y_real, y_score) 
plt.plot(fpr, tpr) # 繪圖
plt.show()

# P-R曲線
from sklearn.metrics import precision_recall_curve
precision, recall, _ = precision_recall_curve(y_real, y_score)
plt.plot(recall,precision)

plt.show()

print('------------------------------------------------------------')	#60個

print('多指標評分')

from sklearn.metrics import classification_report

y_real = [0, 1, 1, 1, 1, 1, 0, 0, 0, 0]
y_score=[0.9, 0.75, 0.86, 0.47, 0.55, 0.56, 0.74, 0.22, 0.5, 0.26]
y_pred = [round(i) for i in y_score]
print(classification_report(y_real, y_pred))

print('------------------------------------------------------------')	#60個

print('K近鄰算法')

from sklearn import neighbors, datasets
from sklearn.model_selection import train_test_split

data = datasets.load_breast_cancer()
X = data.data # 自變量
y = data.target # 因變量
x_train,x_test,y_train,y_test = train_test_split(X,y,test_size=0.1,random_state=0)
clf = neighbors.KNeighborsClassifier(5) # 設鄰居數爲5個
clf.fit(x_train, y_train) # 訓練模型
print(clf.score(x_test, y_test)) # 給模型打分
print(clf.predict([x_test[0]]), y_test[0], clf.predict_proba([x_test[0]]))

print('------------------------------------------------------------')	#60個

from sklearn.metrics import accuracy_score
from scipy.spatial import distance
import operator

def classify(inX, dataSet, labels, k):
    #S=np.cov(dataSet.T)   #協方差矩陣，爲計算馬氏距離
    #SI = np.linalg.inv(S)  #協方差矩陣的逆矩陣
    #distances = np.array(distance.cdist(dataSet, [inX], 'mahalanobis', VI=SI)).reshape(-1)
    distances = np.array(distance.cdist(dataSet, [inX], 'euclidean').reshape(-1))
    sortedDistIndicies = distances.argsort() # 取排序的索引，用於label排序
    classCount={}
    for i in range(k): # 訪問距離最近的五個實例
        voteILabel = labels[sortedDistIndicies[i]]
        classCount[voteILabel]=classCount.get(voteILabel,0)+1
    sortedClassCount = sorted(classCount.items(), 
             key=operator.itemgetter(1), reverse=True)
    return sortedClassCount[0][0] # 取最多的分類

ret = [classify(x_test[i], x_train, y_train, 5) for i in range(len(x_test))]
print(accuracy_score(y_test, ret))

print('------------------------------------------------------------')	#60個

print('聚類算法')

from sklearn.datasets import make_blobs  # 數據支持
from sklearn.cluster import KMeans  # 聚類方法

X,y = make_blobs(n_samples=100, random_state=150) 
y_pred = KMeans(n_clusters=3, random_state=5).fit_predict(X)  # 訓練
plt.scatter(X[:,0],X[:,1],c=y_pred)
plt.show()

print('------------------------------------------------------------')	#60個

print('------------------------------------------------------------')	#60個

print('線性迴歸')

def train(xArr,yArr):  # 訓練模型
    m,n = np.shape(xArr)
    xMat = np.mat(np.ones((m, n+1))) # 加第一列設爲1，用於計算截距
    x = np.mat(xArr)
    xMat[:,1:n+1] = x[:,0:n]
    yMat = np.mat(yArr).T  
    xTx = xMat.T*xMat
    if np.linalg.det(xTx) == 0.0:  #行列式的值爲0，無逆矩陣
        print("This matrix is sigular, cannot do inverse") 
        return None
    ws = xTx.I*(xMat.T*yMat)  
    return ws  

def predict(xArr, ws): # 預測
    m,n = np.shape(xArr)
    xMat = np.mat(np.ones((m, n+1))) # 加第一列設爲1, 爲計算截距
    x = np.mat(xArr)
    xMat[:,1:n+1] = x[:,0:n];
    return xMat*ws

if __name__ == '__main__':
    x = [[1], [2], [3], [4]]
    y = [4.1, 5.9, 8.1, 10.1]
    ws = train(x,y)
    if isinstance(ws, np.ndarray):
        print(ws)  # 返回結果：[[2.  ] [2.02]]
        print(predict([[5]], ws)) # 返回結果：[[12.1]]
        plt.scatter(x, y, s=20) # 繪圖
        yHat = predict(x, ws)
        plt.plot(x, yHat, linewidth=2.0) 
        plt.show()

print('------------------------------------------------------------')	#60個

print('邏輯迴歸')

def sigmoid(x): # S函數實現
    return 1.0 / (1.0 + np.exp(-x))
x = np.arange(-10,10,0.2) # 生成從-10, 10， 間隔爲0.2的數組
y = [sigmoid(i) for i in x]
plt.grid(True) # 顯示網格
plt.plot(x,y)

plt.show()


print('------------------------------------------------------------')	#60個

from sklearn import svm
from sklearn.model_selection import train_test_split
from sklearn.datasets import load_iris

iris=load_iris()
X = iris.data  # 獲取自變量
y = iris.target  # 獲取因變量
X_train, X_test, y_train ,y_test = train_test_split(X,y,test_size=0.2,random_state=0)
clf = svm.SVC(C=0.8, kernel='rbf', gamma=1) # 高斯核，鬆弛度0.8
#clf = svm.SVC(C=0.5, kernel='linear') # 線性核，鬆弛度0.5
clf.fit(X_train, y_train.ravel())

print('trian pred:%.3f' %(clf.score(X_train, y_train))) # 對訓練集打分
print('test pred:%.3f' %(clf.score(X_test, y_test))) # 對測試集打分
print(clf.support_vectors_) #支持向量列表，從中看到切分邊界
print(clf.n_support_) # 每類別持向量個數

plt.plot(X_train[:,0], X_train[:,1],'o', color = '#bbbbbb')
plt.plot(clf.support_vectors_[:,0], clf.support_vectors_[:,1],'o')

plt.show()

print('------------------------------------------------------------')	#60個

print('信息量和熵')

import math
def entropy(*c):
    if(len(c)<=0):
        return -1
    result = 0
    for x in c:
        result+=(-x)*math.log(x,2)
    return result;
print(entropy(0.99,0.01))
print(entropy(0.5,0.5))
print(entropy(0.333,0.333,0.333))


print('------------------------------------------------------------')	#60個

print('決策樹')

from sklearn.datasets import load_iris # 鳶尾花數據集
from sklearn.model_selection import train_test_split # 切分數據集工具
from sklearn import tree # 決策樹工具
import pydotplus # 做圖工具
import io

iris=load_iris()
X = iris.data  # 獲取自變量
y = iris.target  # 獲取因變量
X_train, X_test, y_train ,y_test = train_test_split(X,y,test_size=0.2,random_state=0)
clf = tree.DecisionTreeClassifier(max_depth=5)
clf.fit(X_train,y_train) # 訓練模型
print("score:", clf.score(X_test,y_test)) # 模型打分
# 生成決策樹圖片
dot_data = io.StringIO()
tree.export_graphviz(clf,out_file=dot_data, 
                     feature_names=iris.feature_names,
                     filled=True,rounded=True,
                     impurity=False)
graph = pydotplus.graph_from_dot_data(dot_data.getvalue())

#fail
#open('a.jpg','wb').write(graph.create_jpg()) # 保存圖片

print('------------------------------------------------------------')	#60個

print('Apriori關聯規則')


from numpy import *

def load1(): 
    return [['香蕉','蘋果','梨','葡萄','櫻桃','西瓜','芒果','枇杷'],
            ['蘋果','菠蘿' ,'梨','香蕉','荔枝','芒果','橙子'],
            ['菠蘿','香蕉','桔子','橙子'],
            ['菠蘿','梨','枇杷'],
            ['蘋果','香蕉' ,'梨','荔枝','枇杷','芒果','香瓜']]

# 建立所有物品集合
def create_collection_1(data):
    c = []
    for item in data:
        for g in item:
            if not [g] in c:
                c.append([g])                
    c.sort()
    return list(map(frozenset, c))

def check_support(d_list, c_list, min_support):
    # d_list是購物數據，c_list是物品集合，support是支持度
    c_dic = {} # 組合計數
    for d in d_list: # 每次購物
        for c in c_list: # 每個組
            if c.issubset(d):
                if c in c_dic: 
                    c_dic[c]+=1 # 組合計數加1
                else: 
                    c_dic[c]=1 # 將組合加入字典
    d_count = float(len(d_list)) # 購物次數
    ret = []
    support_dic = {}
    for key in c_dic:
        support = c_dic[key]/d_count
        if support >= min_support: # 判斷支持度
            ret.append(key)
        support_dic[key] = support # 記錄支持度
    return ret, support_dic # 返回滿足支持率的組和支持度字典

def create_collection_n(lk, k):
    ret = []
    for i in range(len(lk)):
        for j in range(i+1, len(lk)): 
            l1 = list(lk[i])[:k-2];
            l1.sort()
            l2 = list(lk[j])[:k-2]
            l2.sort()
            if l1==l2:
                ret.append(lk[i] | lk[j])
    return ret

def apriori(data, min_support = 0.5):
    c1 = create_collection_1(data)
    d_list = list(map(set, data)) # 將購物列表轉換成集合列表
    l1, support_dic = check_support(d_list, c1, min_support)
    l = [l1]
    k = 2
    while (len(l[k-2]) > 0):
        ck = create_collection_n(l[k-2], k) # 建立新組合
        lk, support = check_support(d_list, ck, min_support) # 判斷新組是否適合支持率
        support_dic.update(support)
        l.append(lk) # 將本次結果加入整體
        k += 1
    return l, support_dic

data = load1()
l,support_dic = apriori(data)
print(l)
print()
print(support_dic)

print('------------------------------------------------------------')	#60個

print('樸素貝葉斯')

import jieba

def load2():
    arr = ['不知道該說什麼, 這麼爛的抄襲片也能上映, 我感到很尷尬',
       '天吶。一個大寫的滑稽。',
       '劇情太狗血，演技太浮誇，結局太無語。總體太渣了。這一個半小時廢了。',
       '畫面很美，音樂很好聽，主角演的很到位，很值得一看的電影，男主角很帥很帥，贊贊贊',
       '超級喜歡的一部愛情影片',
       '故事情節吸引人，演員演的也很好，電影裏的歌也好聽，總之值得一看，看了之後也會很感動的。']
    ret = []
    for i in arr:
        words = jieba.cut(i) # 將句子切分成詞
        ret.append(words)
    return ret,[0,0,0,1,1,1]

def create_vocab(data):
    vocab_set = set([])# 使用set集合操作去掉重複出現的詞彙
    for document in data:
        vocab_set = vocab_set | set(document) 
    return list(vocab_set)

def words_to_vec(vocab_list, vocab_set):  # 將句轉換成詞表格式
    ret = np.zeros(len(vocab_list)) # 創建數據表中的一行，並置初值爲0（不存在）
    for word in vocab_set:
        if word in vocab_list:
            ret[vocab_list.index(word)] = 1  # 若該詞在本句中出現，則設置爲1
    return ret

def train(X, y):
    rows = X.shape[0]
    cols = X.shape[1]
    percent = sum(y)/float(rows) # 正例佔比
    p0_arr = np.ones(cols) # 設置初值爲1，後作爲分子
    p1_arr = np.ones(cols)
    p0_count = 2.0 # 設初值爲2，後作爲分母
    p1_count = 2.0
    for i in range(rows): # 按每句遍歷
        if y[i] == 1:
            p1_arr += X[i] # 數組按每個值相加
            p1_count += sum(X[i]) # 句子所有詞個數相加(只計詞彙表中詞)
        else:
            p0_arr += X[i]
            p0_count += sum(X[i])
    p1_vec = np.log(p1_arr/p1_count) # 正例時，每個詞出現概率
    p0_vec = np.log(p0_arr/p0_count)
    return p0_vec, p1_vec, percent

def predict(X, p0_vec, p1_vec, percent):
    p1 = sum(X * p1_vec) + np.log(percent) # 爲1的概率
    p0 = sum(X * p0_vec) + np.log(1.0 - percent) #爲0的概率
    if p1 > p0:
        return 1
    else:
        return 0

sentences,y = load2()
vocab_list = create_vocab(sentences)
X=[]
for sentence in sentences:
    X.append(words_to_vec(vocab_list, sentence))
p0_vec, p1_vec, percent = train(np.array(X), np.array(y))
test = jieba.cut('抄襲得那麼明顯也是醉了！')
print(test)
test_X = np.array(words_to_vec(vocab_list, test))
print(test_X)
print(test,'分類',predict(test_X, p0_vec, p1_vec, percent))

print('------------------------------------------------------------')	#60個

""" some fail
from hmmlearn import hmm

states = ["A", "B", "C"] # 定義隱藏狀態
n_states = len(states)

observations = ["down","up"] # 定義觀測狀態
n_observations = len(observations)

p = np.array([0.7, 0.2, 0.1]) # 設置初始值概率pi
a = np.array([  # 設置狀態轉移矩陣A
   [0.5, 0.2, 0.3],
   [0.3, 0.5, 0.2],
   [0.2, 0.3, 0.5]
])
b = np.array([  # 設置狀態對觀測的生成矩陣B
  [0.6, 0.2],
  [0.3, 0.3],
  [0.1, 0.5]
])
o = np.array([[1, 0, 1, 1, 1]]).T # 設置觀測狀態

model = hmm.MultinomialHMM(n_components=n_states)
model.startprob_= p
model.transmat_= a
model.emissionprob_= b

logprob, h = model.decode(o, algorithm="viterbi")
print("The hidden h", ", ".join(map(lambda x: states[x], h))) # 顯示隱藏狀態

"""


print('------------------------------------------------------------')	#60個

"""
# load_boston 已被移除 但可以試一下 從 warning 訊息

from sklearn import ensemble
from sklearn import datasets
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split

boston = datasets.load_boston() # 讀取Sklearn自帶的數據集
X_train,X_test,y_train,y_test = train_test_split(boston.data, boston.target,
                                                 test_size=0.2,random_state=13)
params = {'n_estimators': 200, 'max_depth': 5, 
          'min_samples_split': 5,'learning_rate': 0.01,
          'loss': 'ls', 'random_state': 0}
clf = ensemble.GradientBoostingRegressor(**params)
clf.fit(X_train, y_train) # 訓練模型
print("MSE: %.2f" % mean_squared_error(y_test, clf.predict(X_test)))

test_score = []
for i, y_pred in enumerate(clf.staged_predict(X_test)):
    test_score.append(clf.loss_(y_test, y_pred)) # 計算測試集誤差
plt.plot(clf.train_score_, 'y-') # 黃色(淺色)
plt.plot(test_score, 'b-') # 藍色(深色)

plt.show()
"""

print('------------------------------------------------------------')	#60個

""" 安裝 auto-sklearn fail
print('Auto-Sklearn')

#pip install auto-sklearn
import autosklearn.classification
import statsmodels.api as sm
import warnings
warnings.filterwarnings('ignore')
  
data = sm.datasets.anes96.load_pandas().data
label = 'vote'
features = [i for i in data.columns if i != label]
X_train = data[features]
y_train = data[label]
automl = autosklearn.classification.AutoSklearnClassifier(
    time_left_for_this_task=120, per_run_time_limit=120, # 兩分鐘
    include_estimators=["random_forest"])
automl.fit(X_train, y_train)
print(automl.score(X_train, y_train))
"""

print('------------------------------------------------------------')	#60個

""" 安裝 auto-ml fail
print('Auto-ML')

#pip install auto-ml

from auto_ml import Predictor
import statsmodels.api as sm

data = sm.datasets.anes96.load_pandas().data
column_descriptions = {
    'vote': 'output',
    'TVnews': 'categorical',
    'educ': 'categorical',
    'income': 'categorical',
}

ml_predictor = Predictor(type_of_estimator='classifier', 
                         column_descriptions=column_descriptions)
model = ml_predictor.train(data)
model.score(data, data.vote)
"""

print('------------------------------------------------------------')	#60個

""" 一些 fail
print('Auto-Keras')

from keras.datasets import mnist
from autokeras import ImageClassifier
from autokeras.constant import Constant
import autokeras
from keras.utils import plot_model
    
(x_train, y_train), (x_test, y_test) = mnist.load_data()
x_train = x_train.reshape(x_train.shape + (1,))
x_test = x_test.reshape(x_test.shape + (1,))
clf = ImageClassifier(verbose=True, augment=False)
clf.fit(x_train, y_train, time_limit=500 * 60)
clf.final_fit(x_train, y_train, x_test, y_test, retrain=True)
y = clf.evaluate(x_test, y_test)
print(y * 100)
clf.export_keras_model('model.h5')
plot_model(clf, to_file='model.png')
"""

print('------------------------------------------------------------')	#60個

print('分詞工具')

import jieba
print(' '.join(jieba.cut('今天我去參觀展覽館', cut_all=True))) # 全模式
print(' '.join(jieba.cut('今天我去參觀展覽館', cut_all=False))) # 精確模式

jieba.load_userdict('data/a.txt')
print(jieba.cut('今天我去參觀展覽館'))

import jieba.posseg as pseg
words = pseg.cut("今天我去參觀展覽館")
for w in words:
    print("%s %s" %(w.word, w.flag))

print('------------------------------------------------------------')	#60個

print('TF-IDF逆文本頻率指數')

import jieba
from sklearn.feature_extraction.text import CountVectorizer 
from sklearn.feature_extraction.text import TfidfTransformer  

arr = ['第一天我參觀了美術館',
       '第二天我參觀了博物館',
       '第三天我參觀了動物園',]

arr = [' '.join(jieba.cut(i)) for i in arr] # 分詞
print(arr)


vectorizer = CountVectorizer() 
X = vectorizer.fit_transform(arr) 
word = vectorizer.get_feature_names_out() 
df = pd.DataFrame(X.toarray(), columns=word)
print(df)

transformer = TfidfTransformer()
tfidf = transformer.fit_transform(X)
weight = tfidf.toarray()
for i in range(len(weight)): # 訪問每一句
    print("第{}句：".format(i))
    for j in range(len(word)):  # 訪問每個詞
        if weight[i][j] > 0.05:  # 只顯示重要關鍵字
            print(word[j],round(weight[i][j],2))  # 保留兩位小數

print('------------------------------------------------------------')	#60個

# 寫程序實現TF-IDF方法

from collections import Counter

countlist = []
for i in range(len(arr)):
    count = Counter(arr[i].split(' ')) # 用空格將字串切分成字符串列表，統計每個詞出現次數
    countlist.append(count)
print(countlist)

def tf(word, count): 
    return count[word] / sum(count.values())
def contain(word, count_list): # 統計包含關鍵詞word的句子數量
    return sum(1 for count in count_list if word in count)
def idf(word, count_list):
    return np.log(len(count_list) / (contain(word, count_list)) + 1)  #爲避免分母爲0，分母加1
def tfidf(word, count, count_list):
    return tf(word, count) * idf(word, count_list)
for i, count in enumerate(countlist):
    print("第{}句：".format(i))
    scores = {word: tfidf(word, count, countlist) for word in count}
    for word, score in scores.items():
        print(word, round(score, 2))

'''

print('------------------------------------------------------------')	#60個

print('切分數據集與交叉驗證')

from sklearn.model_selection import train_test_split
from sklearn.datasets import load_iris

iris = load_iris()
X = iris.data
y = iris.target
X_train,X_test,y_train,y_test=train_test_split(X,y,test_size=0.3,random_state=10)
X_train,X_test=train_test_split(X,test_size=0.3,random_state=10)

print('------------------------------------------------------------')	#60個

""" some fail
from sklearn.cross_validation import KFold
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.datasets import load_iris
from sklearn.svm import SVC

iris = load_iris()
X_train,X_test,y_train,y_test=train_test_split(iris.data,iris.target,
                                               test_size=0.3,random_state=10)
num = 5 # 5折交叉驗證
train_preds = np.zeros(X_train.shape[0]) # 用於保存預測結果
test_preds = np.zeros((X_test.shape[0], num))
kf = KFold(len(X_train), n_folds = num, shuffle=True, random_state=0)
for i, (train_index, eval_index) in enumerate(kf):
    clf = SVC(C=1, gamma=0.125, kernel='rbf')
    clf.fit(X_train[train_index], y_train[train_index])
    train_preds[eval_index] += clf.predict(X_train[eval_index])
    test_preds[:,i] = clf.predict(X_test)
print(accuracy_score(y_train, train_preds)) # 返回結果: 0.971428571429
print(test_preds.mean(axis=1))

from sklearn.model_selection import cross_val_score # python 3使用
# from sklearn.cross_validation import cross_val_score # python 2 使用
print(cross_val_score(clf, iris.data, iris.target).mean())
"""
print('------------------------------------------------------------')	#60個

print('模型調參')

# 網格搜索
from sklearn.model_selection import GridSearchCV
from sklearn.svm import SVC
from sklearn.datasets import load_iris

iris = load_iris()
model = SVC(random_state=1)
param_grid = {'kernel':('linear', 'rbf'), 'C':[1, 2, 4], # 制定參數範圍
              'gamma':[0.125, 0.25, 0.5 ,1, 2, 4]}
gs = GridSearchCV(estimator=model, param_grid=param_grid, scoring='accuracy', 
                  cv=10, n_jobs=-1)
gs = gs.fit(iris.data, iris.target)
y_pred = gs.predict(iris.data)  # 預測
print(gs.best_score_)
print(gs.best_params_)

print('------------------------------------------------------------')	#60個

print('Hyperopt')

# pip install hyperopt
from hyperopt import fmin, tpe, hp, Trials
trials = Trials()
best = fmin(
    fn=lambda x: (x-1)**2, # 最小化目標，如誤差函數
    space=hp.uniform('x', -10, 10), # 定義搜索空間, 名稱爲x，範圍-10~10
    algo=tpe.suggest, # 指定搜索算法
    trials=trials, # 保存每次迭代的具體信息
    max_evals=50) # 評估次數
print(best) # 返回結果：{'x': 0.980859461591201}
for t in trials.trials:
    print(t['result'])

print('------------------------------------------------------------')	#60個

""" some fail
from sklearn.datasets import load_iris
from sklearn.cross_validation import cross_val_score
from hyperopt import hp,STATUS_OK,Trials,fmin,tpe
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC

def f(params): # 定義評價函數
    t = params['type']
    del params['type']
    if t == 'svm':
        clf = SVC(**params)
    elif t == 'randomforest':
        clf = RandomForestClassifier(**params)
    else:
        return 0
    acc = cross_val_score(clf, iris.data, iris.target).mean() 
    return {'loss': -acc, 'status': STATUS_OK} # 求最小值:準確率加負號

iris=load_iris()
space = hp.choice('classifier_type', [ # 定義可選參數
    {
        'type': 'svm',
        'C': hp.uniform('C', 0, 10.0),
        'kernel': hp.choice('kernel', ['linear', 'rbf']),
        'gamma': hp.uniform('gamma', 0, 20.0)
    },
    {
        'type': 'randomforest',
        'max_depth': hp.choice('max_depth', range(1,20)),
        'max_features': hp.choice('max_features', range(1,5)),
        'n_estimators': hp.choice('n_estimators', range(1,20)),
        'criterion': hp.choice('criterion', ["gini", "entropy"])
    }
])
best = fmin(f, space, algo=tpe.suggest, max_evals=100)
print('best:',best) 
"""
print('------------------------------------------------------------')	#60個

print('學習曲線和驗證曲線')

from sklearn import datasets
from sklearn.ensemble import RandomForestClassifier

def draw_curve(params, train_score, test_score):
    train_mean =  np.mean(train_score,axis=1) # 均值
    train_std = np.std(train_score,axis=1) # 標準差
    test_mean = np.mean(test_score,axis=1)
    test_std=np.std(test_score,axis=1)
    plt.plot(params,train_mean,'--',color = 'g',label = 'training')
    plt.fill_between(params,train_mean+train_std,train_mean-train_std, 
                     alpha=0.2,color='g') # 以半透明方式繪圖區域
    plt.plot(params,test_mean,'o-',color = 'b',label = 'testing')
    plt.fill_between(params,test_mean+test_std,test_mean-test_std, 
                     alpha=0.2,color='b')
    plt.grid() # 顯示網格
    plt.legend() # 顯示圖例文字
    plt.ylim(0.5,1.05) # 設定y軸顯示範圍
    plt.show()


from sklearn.model_selection import learning_curve
breast_cancer = datasets.load_breast_cancer()
X = breast_cancer.data
y = breast_cancer.target

clf = RandomForestClassifier()
params = np.linspace(0.1,1.0,10) # 從0.1到1，切分成10份
train_sizes,train_score,test_score = learning_curve(clf,X,y,
                            train_sizes=params,
                            cv=10,scoring='accuracy') # 10折交叉驗證
draw_curve(params, train_score, test_score)



from sklearn.model_selection import validation_curve
params = [10,20,40,80,160,240]
train_score,test_score = validation_curve(RandomForestClassifier(),
                     X,y,param_name='n_estimators',cv=10,scoring='accuracy',
                     param_range=params)
draw_curve(params, train_score, test_score)

print('------------------------------------------------------------')	#60個

'''
# 數據集和數據處理

from pandas import Series,DataFrame

# 繪圖分析
import seaborn as sns
sns.set_style('whitegrid')

# 機器學習
from sklearn.linear_model import LogisticRegression # 邏輯迴歸
from sklearn.svm import SVC, LinearSVC # 支持向量機
from sklearn.ensemble import RandomForestClassifier # 隨機森林
#from sklearn.neighbors import KneighborsClassifier # K近鄰
from sklearn.naive_bayes import GaussianNB# 數據集和數據處理


print('------------------------------------------------------------')	#60個

titanic_df = pd.read_csv('data/train.csv')
test_df = pd.read_csv('data/test.csv')
print(titanic_df.head())
print(titanic_df.info())
print(titanic_df.describe())



print('------------------------------------------------------------')	#60個


facet = sns.FacetGrid(titanic_df, hue="Survived",aspect=4)
facet.map(sns.kdeplot,'Age',shade= True)
facet.set(xlim=(0, titanic_df['Age'].max()))
facet.add_legend()
plt.show()

fig, axis1 = plt.subplots(1,1,figsize=(18,4))
average_age = titanic_df[["Age", "Survived"]].groupby(['Age'],as_index=False).mean()
sns.barplot(x='Age', y='Survived', data=average_age)
plt.show()


print('------------------------------------------------------------')	#60個

def get_person(passenger): # 小於16歲的分類爲兒童
    age,sex = passenger
    return 'child' if age < 16 else sex

def conv(df):
    df['Person'] = df[['Age','Sex']].apply(get_person,axis=1) # 組合特徵
    df['Fare'] = df['Fare'].fillna(df['Fare'].mean()) # 缺失值填充爲均值
    df["Embarked"] = df["Embarked"].fillna("S") # 缺失值填充爲S
    df['Fare'] = df['Fare'].astype(int) # 類型轉換

    person_dummies  = pd.get_dummies(df['Person']) # onehot編碼
    person_dummies.columns = ['Child','Female','Male']
    df = df.join(person_dummies) # 連接原數據與onehot數據
    df = df.drop(['PassengerId','Name','Ticket','Person','Sex','Embarked','Cabin','Age'], 
                         axis=1) # 刪除非數據型特徵
    return df

titanic_df = conv(titanic_df)
test_df = conv(test_df)

print('------------------------------------------------------------')	#60個

# 生成模型所需的訓練集和測試集
X_train = titanic_df.drop("Survived",axis=1)
Y_train = titanic_df["Survived"]
X_test  = test_df.copy()

from sklearn.linear_model import LogisticRegression

logreg = LogisticRegression() # 初始化模型
logreg.fit(X_train, Y_train) # 訓練模型
print(logreg.score(X_train, Y_train)) # 模型評分
Y_pred = logreg.predict(X_test) # 預測

'''
print('------------------------------------------------------------')	#60個

"""
實際數據請從天池競賽平臺下載
https://tianchi.aliyun.com/competition/gameList/activeList
https://tianchi.aliyun.com/competition/activeList
"""

import datetime
from pandas.api.types import is_numeric_dtype # 用於判斷特徵類型
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier #分類模型
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor # 迴歸模型
from sklearn.model_selection import cross_val_score, train_test_split # 切分數據集
from sklearn.metrics import mean_squared_error # 評價函數

"""
無csv資料
data = pd.read_csv('data/happiness_train_min.csv', encoding='gb2312')

test = pd.read_csv('data/happiness_test_min.csv', encoding='gb2312')

print(data.columns.tolist()) # 查看所有特徵
print(data.dtypes) # 查看各特徵類型

print('------------------------------------------------------------')	#60個

# 特徵工程

features = []
label = 'happiness' # 目標變量

for col in data.columns:
    if not is_numeric_dtype(data[col]): # 非數值型特徵
        print(col, data[col].dtype)
        print(data[col].unique()[:5])
    elif col != label and col != 'id': # 加入可直接代入模型的特徵
        features.append(col)
        
x = data[features] # 自變量
y = data[label] # 目標變量
x_train, x_val, y_train, y_val = train_test_split(x, y, test_size=0.25, random_state=0)
x_train = x_train.fillna(x.mean()) # 空值填充訓練集
x_val = x_val.fillna(x.mean()) # 空值填充驗證集
x_test = test.fillna(x.mean()) # 空值填充測試集
x = x.fillna(x.mean()) # 空值填充全集


print('------------------------------------------------------------')	#60個

# 訓練模型生成提交數據

#clf = RandomForestRegressor(criterion='mse', random_state=0) # 隨機森林迴歸
#clf = GradientBoostingClassifier(criterion='mse',random_state=0) # GBDT分類
clf = GradientBoostingRegressor(criterion='mse', random_state=0) # GBDT迴歸

if True: # 用於本地測試
    clf.fit(x_train, y_train)
    mse = mean_squared_error(y_val, [round(i) for i in clf.predict(x_val)])
    print("MSE: %.4f" % mse)
else: # 用於遠程提交
    clf.fit(x, y) # 全量數據訓練
    df = pd.DataFrame()
    df['id'] = test.id
    df['happiness'] = clf.predict(x_test[features])
    df.to_csv('out/submit_{}.csv'.format(datetime.datetime.now().strftime('%Y%m%d_%H%M%S')),index=False)

print('------------------------------------------------------------')	#60個

import datetime
from pandas.api.types import is_numeric_dtype # 用於判斷特徵類型
from sklearn.model_selection import cross_val_score, train_test_split # 切分數據集
from sklearn.metrics import mean_squared_error # 評價函數

data = pd.read_csv('data/happiness_train_min.csv', encoding='gb2312')
test = pd.read_csv('data/happiness_test_min.csv', encoding='gb2312')

print('------------------------------------------------------------')	#60個

# 特徵工程

def get_mean(fea, data, test): # 同時變換訓練集和測試集
    arr1 = data[fea].unique()
    arr2 = test[fea].unique()
    arr3 = list(arr1)
    arr3.extend(arr2) # 有的數據只出現在訓練集或測試集中
    arr4 = list(set(arr3))
    dic = {}
    for x in arr4:
        dic[x] = data[data[fea] == x][label].mean() # 取其因變量均值
    data[fea] = data[fea].apply(lambda x: dic[x]) # 數據替換
    test[fea] = test[fea].apply(lambda x: dic[x])
    return data,test

label = 'happiness' # 目標變量
features = []

data, test = get_mean('city', data, test)
data, test = get_mean('invest_other', data, test)
data, test = get_mean('province', data, test)

for col in data.columns:
    if not is_numeric_dtype(data[col]): # 非數值型特徵
        continue
    elif col != label and col != 'id' and col not in ['public_service_7']: # 去掉干擾特徵
        features.append(col)
        data[col] = data[col].apply(lambda x: np.nan if x < 0 else x) # 優化點一
        test[col] = test[col].apply(lambda x: np.nan if x < 0 else x)

data_all = pd.concat([data,test]) # 優化點二
data = data[data['happiness'] > 0] # 去掉因變量缺失的數據
x = data[features] # 自變量
y = data[label] # 目標變量
x_train, x_val, y_train, y_val = train_test_split(x, y, test_size=0.25, random_state=0)
x_train = x_train.fillna(data_all[features].mean()) # 空值填充訓練集
x_val = x_val.fillna(data_all[features].mean()) # 空值填充驗證集
x_test = test.fillna(data_all[features].mean()) # 空值填充測試集
x = x.fillna(data_all[features].mean()) # 空值填充全集

print('------------------------------------------------------------')	#60個

# 訓練模型

import xgboost as xgb
from sklearn.cross_validation import KFold

def my_eval(preds, train): # 自定義評價函數
    score = mean_squared_error(train.get_label(), preds)
    return 'myeval', score

my_params = {"booster":'gbtree','eta': 0.005, 'max_depth': 6, 'subsample': 0.7, 
              'colsample_bytree': 0.8, 'objective': 'reg:linear', 'eval_metric': 'rmse', 
              'silent': True, 'nthread': 4} # 模型參數

train_preds = np.zeros(len(data)) # 用於保存預測結果
test_preds = np.zeros(len(test))
kf = KFold(len(data), n_folds = 5, shuffle=True, random_state=0) # 5折交叉驗證
for fold, (trn_idx, val_idx) in enumerate(kf):
    print("fold {}".format(fold+1))
    train_data = xgb.DMatrix(data[features].iloc[trn_idx], data[label].iloc[trn_idx]) # 訓練集
    val_data = xgb.DMatrix(data[features].iloc[val_idx], data[label].iloc[val_idx]) # 驗證集
    watchlist = [(train_data, 'train'), (val_data, 'valid_data')]
    clf = xgb.train(dtrain=train_data, num_boost_round=5000, evals=watchlist, 
               early_stopping_rounds=200, verbose_eval=100, 
               params=my_params,feval = my_eval)
    train_preds[val_idx] = clf.predict(xgb.DMatrix(data[features].iloc[val_idx]),
               ntree_limit=clf.best_ntree_limit)
    test_preds += clf.predict(xgb.DMatrix(test[features]), 
               ntree_limit=clf.best_ntree_limit) / kf.n_folds
print("CV score: {:<8.8f}".format(mean_squared_error(train_preds, data[label])))

df = pd.DataFrame() # 生成提交結果
df['id'] = test.id
df['happiness'] = test_preds
df.to_csv('out/submit_{}.csv'.format(datetime.datetime.now().strftime('%Y%m%d_%H%M%S')),index=False)


print('------------------------------------------------------------')	#60個

fig,ax = plt.subplots()
fig.set_size_inches(40,6)
xgb.plot_tree(clf, ax=ax, num_trees=0) # 顯示模型中的第一棵樹
plt.savefig('tmp.png',dpi=300)


print('------------------------------------------------------------')	#60個

# 檢測干擾變量

from sklearn.ensemble import GradientBoostingRegressor

baseline = 0.4887 # 誤差baseline
for i in features:
    features_new = [x for x in features if x != i]
    clf = GradientBoostingRegressor(criterion='mse', random_state=0)
    clf.fit(x_train[features_new], y_train)
    mse = mean_squared_error(y_val, [round(i) for i in clf.predict(x_val[features_new])])
    if mse < baseline:
        print("remove", i, "MSE: %.4f" % mse)

"""
print('------------------------------------------------------------')	#60個






print('------------------------------------------------------------')	#60個




print('------------------------------------------------------------')	#60個



print('------------------------------------------------------------')	#60個







print('------------------------------------------------------------')	#60個











print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
