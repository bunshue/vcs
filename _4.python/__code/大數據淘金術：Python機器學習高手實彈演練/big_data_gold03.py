import sys

import pandas as pd

print('------------------------------------------------------------')	#60個
'''

import logging

# 獲取logger對象,取名mylog
logger = logging.getLogger("mylog")
# 輸出DEBUG及以上級別的信息，針對所有輸出的第一層過濾
logger.setLevel(level=logging.DEBUG)

# 獲取文件日誌句柄並設置日誌級別，第二層過濾
handler = logging.FileHandler("log.txt")
handler.setLevel(logging.INFO)	

# 生成並設置文件日誌格式，其中name爲上面設置的mylog
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)

# 獲取流句柄並設置日誌級別，第二層過濾
console = logging.StreamHandler()
console.setLevel(logging.WARNING)

# 爲logger對象添加句柄
logger.addHandler(handler)
logger.addHandler(console)

# 記錄日誌
logger.info("show info")
logger.debug("show debug")
logger.warning("show warning")


print('------------------------------------------------------------')	#60個

print('讀寫XML文件')

from xml.dom import minidom

dom=minidom.Document()
root_node=dom.createElement('root') # 創建根節點
dom.appendChild(root_node) # 添加根節點

book_node=dom.createElement('blog') # 創建第一個子節點
book_node.setAttribute('level','3') # 添加屬性
root_node.appendChild(book_node) # 爲root添加子節點

name_node=dom.createElement('addr') # 創建第二個子節點
name_text=dom.createTextNode('https://blog.csdn.net/xieyan0811') # 添加文字
name_node.appendChild(name_text)
root_node.appendChild(name_node)

# toxml() 轉換成字符串, toprettyxml()轉換成樹形縮進版式
print(dom.toprettyxml())
with open('test_dom.xml','w') as fh:
    dom.writexml(fh, indent='',addindent='\t', newl='\n', encoding='UTF-8')

print('------------------------------------------------------------')	#60個


from xml.dom import minidom
with open('test_dom.xml','r') as fh:
    dom = minidom.parse(fh) # 獲取dom對象
    root = dom.documentElement # 獲取根節點
    print("node name", root.nodeName) # 顯示節點名: root
    print("node type", root.nodeType) # 顯示節點類型
    print("child nodes", root.childNodes) # 列出所有子節點
    blog = root.getElementsByTagName('blog')[0] # 根據標籤名獲取元素列表
    print(blog.getAttribute('level')) # 獲取屬性值
    addr = root.getElementsByTagName('addr')[0]
    print("addr's child nodes", addr.childNodes)
    text_node = addr.childNodes[0] # 獲取文本節點內容
    print("text data", text_node.data)
    print("parent", addr.parentNode.nodeName) # 顯示name的父節點名稱


print('------------------------------------------------------------')	#60個

print('讀寫Json文件')

import json

data = [{"group":0,"param":["one","two","three"]},
        {"group":1,"param":["1","2","3"]}] 

jsonstr = json.dumps(data)
print(jsonstr)
jsonstr = json.dumps(data, sort_keys=True, 
                 indent=4, separators=(',', ': '))
print(jsonstr)
data1 = json.loads(jsonstr)
print(data1, type(data1))

with open('json.txt','w') as json_file:
    json.dump(data, json_file)
    json_file.close()

with open('json.txt','r') as json_file:
    data = json.load(json_file)
    json_file.close()
print(data1, type(data1))

print('------------------------------------------------------------')	#60個

print('讀寫CSV文件')

import pandas as pd

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
print(df.info()) # 顯示dataframe相關信息
df.to_csv("tmp.csv", index=False, header=True, columns=['Name','Sex','Age'])

df1 = pd.read_csv("tmp.csv")
print(df1.info())
print(df1)

print('------------------------------------------------------------')	#60個

print('讀寫PKL文件')

import pandas as pd

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
print(df.info())
df.to_pickle("tmp.pkl")

df1 = pd.read_pickle("tmp.pkl")
print(df1.info())


print('------------------------------------------------------------')	#60個

import pickle
data1 = {'a': [1, 2.0, 4+6j],
         'b': ('string1', u'Unicode string'),
         'c': None}
output = open('tmp2.pkl', 'wb')
pickle.dump(data1, output)
output.close()

pkl_file = open('tmp2.pkl', 'rb')
data2 = pickle.load(pkl_file)
print(data2)
pkl_file.close()




print('------------------------------------------------------------')	#60個

"""
from sklearn import svm
from sklearn import datasets
from sklearn.externals import joblib

clf = svm.SVC()
iris = datasets.load_iris()
clf.fit(iris.data, iris.target)
joblib.dump(clf, "tmp3.pkl")

clf1 = joblib.load("tmp3.pkl")
print(clf1.predict(iris.data[:2]))
"""

print('------------------------------------------------------------')	#60個

print('讀寫HDF5文件')

import h5py
import numpy as np

f = h5py.File('tmp.h5','w')
f['data'] = np.zeros((3,3))
f['labels'] = np.array([1,2,3,4,5])
f.close()

f = h5py.File('tmp.h5','r')
for key in f.keys():
    print(f[key].name)
    print(f[key].shape)
    #print(f[key].value)
f.close()



print('------------------------------------------------------------')	#60個

print('讀寫Excel文件')

import pandas as pd
import openpyxl

df = pd.DataFrame({'Name': ['Smith', 'Lucy'], 'Age': ['25', '20'], 'Sex': ['男','女']})
df.to_excel("tmp.xlsx")

df1 = pd.read_excel("tmp.xlsx")
print(df1)

wb = openpyxl.load_workbook('tmp.xlsx')
sheets = wb.sheetnames
print(sheets)
for i in range(len(sheets)):
    sheet = wb[sheets[i]]
    print('title', sheet.title)
    for col in sheet.iter_cols(min_row=0, min_col=0, max_row=3, max_col=3):
        for cell in col:
            print(cell.value)


print('------------------------------------------------------------')	#60個


print('Python存取Sqlite數據庫')

import sqlite3

conn = sqlite3.connect('test.db')
c = conn.cursor()
c.execute("""CREATE TABLE IF NOT EXISTS TIPS 
       (NAME           TEXT    NOT NULL,
       ADDRESS        CHAR(50),
       BILL         REAL);""") # 創建數據表
c.execute("INSERT INTO TIPS (NAME,ADDRESS,BILL) \
      VALUES ('Zhang', 'Beijing', 1004.00 )"); # 向表中輸入數據
cursor = c.execute("SELECT * from TIPS")
for row in cursor:
    print(row)
conn.commit()
conn.close()

print('------------------------------------------------------------')	#60個

print('抓取網站數據')

import urllib.request
from bs4 import BeautifulSoup

response = urllib.request.urlopen("https://blog.csdn.net/xieyan0811")
html = response.read().decode("utf-8","ignore") # 返回網頁爲utf-8編碼，解碼時忽略錯誤
reg = r'http://'  
soup = BeautifulSoup(html, 'html.parser')
for link in soup.find_all('a'):
    addr = link.get('href')
    # 顯示包含關鍵字的所有地址
    if addr != None and addr.find('xieyan0811/article') != -1: 
        print(addr)
        

print('------------------------------------------------------------')	#60個

print('使用POST方式抓取數據')

import requests
params = {'key1': 'value1', 'key2': 'value2'}
r = requests.post("http://httpbin.org/post", data=params)
print(r.text)


print('------------------------------------------------------------')	#60個

print('基本類型轉換')


import pandas as pd
import numpy as np
dic = {
     'string': ['dog', 'snake', 'cat', 'dog', 'monkey', 'elephant'],
     'integer': [2000, 2000, 2001, 2002, 2003, np.nan],
     'float': [1.5, 1.5, 1.7, np.nan, np.nan, 8.3],
     'dtime': ['2018-01-01', '2018/01/02', '2018-01-03', '2018-01-04', '2018-01-05', np.nan],
     'mix': [1, 1, 0, '+', 0, 1],
     'classify': ['A', 'B', 'A', 'B', 'A', 'A']
                }
data = pd.DataFrame(dic)
print(data.dtypes)


data['dtime'] = pd.to_datetime(data['dtime'], infer_datetime_format=True)
data['mix']=pd.to_numeric(data['mix'],errors='coerce')
data['classify']=pd.Categorical(data['classify'])
data['float']=data['float'].astype(np.float32)
print(data.dtypes)



print('------------------------------------------------------------')	#60個


print('缺失值處理')

import pandas as pd
import numpy as np

dic = {   
     'state': ['Ohio', 'Ohio', 'Ohio', 'Ohio', 'Nevada', 'Nevada'],
     'year': [2000, 2000, 2001, 2002, 2003, 3456],
     'score': [1.5, 1.5, 1.7, np.nan, np.nan, 8.3],
     'desc': [np.nan, np.nan, np.nan, np.nan, np.nan, 3],
     'val1': [1, 1, 0, '+', 0, 1],
}
data = pd.DataFrame(dic)

print(data['desc'].nunique()) # 不同取值個數
print(data['desc'].unique()) # 不同取值列表
print(data['year'].value_counts()) # 不同取值出現次數



print(data['desc'].isnull()) # 是否缺失
print(data['desc'].isnull().any()) # 是否含有任意缺失
print(data['desc'].isnull().all()) # 是否全部缺失
print(data['desc'].isnull().sum(), len(data)) # 空值個數與記錄個數
print(data.dropna(axis=1, how='all'))
print(data['score'].fillna(data['score'].mean()))
print(data['score'].fillna(method='ffill', limit=1))

print(data.interpolate(mdthod='polynomial', order=2)) # 二次多項式插值
print(data.interpolate(mdthod='spline', order=3)) # 三次樣條插值


print('------------------------------------------------------------')	#60個


print('異常值處理')

print(data.query('year<2050'))
print(data[data['year']<2050])

data['val1'] = data['val1'].apply(lambda x: 1 if x == '+' else x)


print('去重處理')

print(data.drop_duplicates(keep='last'))
print(data.drop_duplicates(keep='last', subset='year'))



print('------------------------------------------------------------')	#60個

print('模型篩選特徵')

from sklearn.datasets import load_iris 
from sklearn import tree 

iris = load_iris()
clf = tree.DecisionTreeClassifier()
clf = clf.fit(iris.data, iris.target)
print(clf.feature_importances_)




print('------------------------------------------------------------')	#60個

print('數學方法降維')

from sklearn.decomposition import PCA
from sklearn import datasets
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

iris = datasets.load_iris()
data = pd.DataFrame(iris.data, columns=['SpealLen', 'SpealWid', 
                             'PetalLen', 'PetalWid'])
mat = data.corr()
sns.heatmap(mat, annot=True, vmax=1, vmin=-1, xticklabels= True, 
            yticklabels= True, square=True, cmap="gray")

plt.show()

print('------------------------------------------------------------')	#60個

pca = PCA(n_components=2)
data1 = pca.fit_transform(data)
print(data1.shape)
print(pca.explained_variance_ratio_, 
      pca.explained_variance_ratio_.sum())
plt.scatter(data1[:,0], data1[:,1], c = np.array(iris.target), 
            cmap=plt.cm.copper)


plt.show()


print('------------------------------------------------------------')	#60個

print('簡單隨機抽樣')

import statsmodels.api as sm
import pandas as pd
import numpy as np
import random

data = sm.datasets.anes96.load_pandas().data
df = data.sample(50)
print(df.head())

print('系統抽樣')

index_list = [i for i in range(len(data)) if i % 10 == 0]
df = data.iloc[index_list]
print(df.head())

print('分層抽樣')

def typicalSampling(grp, typicalFracDict):
    name = grp.name
    frac = typicalFracDict[name]
    return grp.sample(frac=frac)

typicalFracDict = {
    0.0: 0.35,  
    1.0: 0.5,  
}
df = data.groupby('vote').apply(typicalSampling, typicalFracDict)
print(df.head())

print('整羣抽樣')
unique = np.unique(data['income'])
sample = random.sample(list(unique),2)
df = pd.DataFrame()
for label in sample:
    tmp = data[data['income']==label]  
    df = pd.concat([df, tmp])
print(df.head())

print('------------------------------------------------------------')	#60個

print('merge方法')

df1 = pd.DataFrame({'id': [1, 2, 3], 'val1': [2, 4, 6]})
df2 = pd.DataFrame({'id': [3, 2, 2], 'val2': [9, 6, 5]})
print(pd.merge(df1, df2, how='left'))

print('------------------------------------------------------------')	#60個

print('concat方法')

df1 = pd.DataFrame({'id': [1, 2, 3], 'val1': [2, 4, 6]})
df2 = pd.DataFrame({'id': [3, 2, 2], 'val2': [9, 6, 5]})
print(pd.concat([df1, df2]))
print(pd.concat([df1, df2], axis=1))

print('------------------------------------------------------------')	#60個

print('數值型特徵')

dic = {'height': [1.6, 1.7, 1.8],
      'weight': [60, 70, 90]}
data = pd.DataFrame(dic)
data['bmi'] = data['weight'] / (data['height'] **2)
print(data)
data['overweight'] = data['bmi'] > 25
print(data)
data['overweight'] = data['overweight'].map({True:'Yes', False:'No'})
print(data)


print('------------------------------------------------------------')	#60個

print('類型特徵')

import pandas as pd
dic = {'string': ['第一組', '第二組', '第二組']}
data = pd.DataFrame(dic)
print(pd.factorize(data.string)) # 轉換成數值型編碼

data['num'] = pd.factorize(data['string'])[0]
df = pd.get_dummies(data['string'], prefix='組別')  # 轉換成onehot類型編碼
new_data = pd.concat([data, df], axis=1)
print(new_data)

print('------------------------------------------------------------')	#60個

print('關鍵字特徵')

import pandas as pd
import numpy as np
from scipy import stats
import jieba
import re

def do_split(test_text):
    pattern = r',|\.|/|;|\'|`|\[|\]|<|>|\?|:|"|\{|\}|\~|!|？|@|#|\$|%|\^|&|\(|\)|-|=|\_|\+|，|。|、|；|‘|’|【|】03   |·|！| |…|（|）' 
    return re.split(pattern, test_text) 



def get_keywords(data, feat):
    ret = []
    data[feat] = data[feat].apply(lambda x: x.strip())
    for i in data[feat].unique():
        # 將短句作爲關鍵字
        if len(i) <= 50 and i not in ret:
            ret.append(i)
        # 將子句作爲關鍵字
        for sentence in do_split(i):
            if len(sentence) <= 50 and sentence not in ret:
                ret.append(sentence)
        # 將詞作爲關鍵字
        for word in jieba.cut(i, cut_all=True):
            if len(word) > 1 and word not in ret:
                ret.append(word)
    return ret

def check_freq(data, feat, keywords, limit):
    ret = []
    for key in keywords:
        try:
            if len(data[data[feat].str.contains(key)]) > limit:
                ret.append(key)
        except:
            pass
    return ret

def do_test(data, feat, key, y, debug=False):
    arr1 = data[data[feat].str.contains(key) == True][y]
    arr2 = data[data[feat].str.contains(key) == False][y]
    ret1 = stats.ttest_ind(arr1, arr2, equal_var = False)
    ret2 = stats.levene(arr1, arr2)
    if ret1.pvalue < 0.05 or ret2.pvalue < 0.05:
        return True    
    return False

def check(data, feat, y):
    ret = []
    keywords = get_keywords(data, feat)
    arr = check_freq(data, feat, keywords, 5)
    for word in arr:
        if do_test(data, feat, word, y):
            ret.append(word)
    return ret

# 讀取數據文件的前500條數據，其中第6個字段是微博內容，第5個字段爲點贊次數。
data = pd.read_csv('data/weibo_train_data.txt', sep='\t', 
                   header=None, nrows=500)
print(check(data, 6, 5))


print('------------------------------------------------------------')	#60個

from scipy import stats
import numpy as np

arr1 = [96,95,95,95,95,95,95,95,95,95,95,95,95,95,95,
        95,95,95,95,95,95,95,95,95,95,95,95,95,95,95]
arr2 = [90,91,92,93,94,90,91,92,93,94,90,91,92,93,94,
        90,91,92,93,94,90,91,92,93,94,90,91,92,93,94]
print(stats.ttest_1samp(arr1, 92))
print(stats.ttest_1samp(arr2, 92))
print((np.mean(arr1)-92)/(np.std(arr1)/np.sqrt(len(arr1)-1)))
print((np.mean(arr2)-92)/(np.std(arr2)/np.sqrt(len(arr2)-1)))

print('------------------------------------------------------------')	#60個




from scipy import stats
import numpy as np

print('正態性檢驗')

# 檢驗樣本是否服從某一分佈
np.random.seed(12345678)
x = stats.norm.rvs(loc=0, scale=1, size=300) # loc爲均值，scale爲方差
print(stats.kstest(x,'norm'))

# 數據的正態性檢驗
np.random.seed(12345678)
x = stats.norm.rvs(loc=10, scale=2, size=70) 
print(stats.shapiro(x))

# 作圖法檢驗正態分佈
import matplotlib.pyplot as plt
np.random.seed(12345678)
x = stats.norm.rvs(loc=10, scale=2, size=300) 
plt.hist(x)
plt.show()

print('------------------------------------------------------------')	#60個

print('方差齊性檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500)  
rvs2 = stats.norm.rvs(loc=25,scale=9,size=500)
print(stats.levene(rvs1, rvs2))


print('------------------------------------------------------------')	#60個

from scipy import stats
import numpy as np

print('兩獨立樣本T檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500)  
rvs2 = stats.norm.rvs(loc=6,scale=10,size=500)
print(stats.ttest_ind(rvs1,rvs2))


print('配對樣本T檢驗')
np.random.seed(12345678)
rvs1 = stats.norm.rvs(loc=5,scale=10,size=500) 
rvs2 = (stats.norm.rvs(loc=5,scale=10,size=500) + stats.norm.rvs(scale=0.2,size=500)) 
print(stats.ttest_rel(rvs1,rvs2))



print('------------------------------------------------------------')	#60個


from scipy import stats

a = [47,56,46,56,48,48,57,56,45,57]  # 分組1
b = [87,85,99,85,79,81,82,78,85,91]  # 分組2
c = [29,31,36,27,29,30,29,36,36,33]  # 分組3
print(stats.f_oneway(a,b,c))

'''
print('------------------------------------------------------------')	#60個

from scipy import stats

A = [6, 15, 22, 36, 40, 48, 53]
B = [3, 4, 5, 12, 17, 18, 21]
print(stats.ranksums(A, B))

C = [1, 2, 3, 4, 5, 6, 7]
print(stats.kruskal(A, B, C))


print('------------------------------------------------------------')	#60個

import statsmodels.api as sm 
import scipy.stats as stats
import pandas as pd

data = sm.datasets.anes96.load_pandas().data
contingency = pd.crosstab(data['vote'], [data['educ']])
print(stats.chi2_contingency(contingency)) # 卡方檢驗

print('------------------------------------------------------------')	#60個

from scipy import stats
import numpy as np

print('圖形描述相關性')
import statsmodels.api as sm
import matplotlib.pyplot as plt
data = sm.datasets.ccard.load_pandas().data
plt.scatter(data['INCOMESQ'], data['INCOME'])

plt.show()

print('正態資料的相關分析')
np.random.seed(12345678)
a = np.random.normal(0,1,100)
b = np.random.normal(2,2,100)
print(stats.pearsonr(a, b))


print('非正態資料的相關分析')
print(stats.spearmanr([1,2,3,4,5], [1,6,7,8,20]))


print('------------------------------------------------------------')	#60個

print('多元線性迴歸')
import statsmodels.api as sm 

data = sm.datasets.ccard.load_pandas().data
model = sm.OLS(endog = data['AVGEXP'],
     exog = data[['AGE','INCOME','INCOMESQ','OWNRENT']]).fit()
print(model.summary())

print('------------------------------------------------------------')	#60個

# 邏輯迴歸

import statsmodels.api as sm
data = sm.datasets.ccard.load_pandas().data
data['OWNRENT'] = data['OWNRENT'].astype(int)
model = sm.Logit(endog = data['OWNRENT'], 
     exog = data[['AVGEXP','AGE','INCOME','INCOMESQ']]).fit()
print(model.summary())


print('------------------------------------------------------------')	#60個

import statsmodels as sm
import tableone 

data = sm.datasets.anes96.load_pandas().data
categorical = ['TVnews', 'selfLR', 'ClinLR', 'educ', 'income'] 
groupby = 'vote'
mytable = tableone.TableOne(data, categorical=categorical, 
                            groupby=groupby, pval=True)
print(mytable)
mytable.to_excel("a.xlsx")



print('------------------------------------------------------------')	#60個


print('------------------------------------------------------------')	#60個


print('------------------------------------------------------------')	#60個

print('------------------------------------------------------------')	#60個


