import PySimpleGUI as sg
#--------------------vvv
#【1. import函式庫】
import calendar
import openpyxl

#【2. 設定於應用程式顯示的字串】
title = "製作Excel月曆"
label1, value1 = "年", "2022"
label2, value2 = "月", "12"
dayname = ["日","一","二","三","四","五","六"]

fontN = openpyxl.styles.Font(size=24)
fontB = openpyxl.styles.Font(size=24, color="0000FF")
fontR = openpyxl.styles.Font(size=24, color="FF0000")
fillB = openpyxl.styles.PatternFill(patternType="solid", fgColor="AAAAFF")
fillR = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFAAAA")

#【3.函數：在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = str(year)+"_"+str(month)+".xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    wb = openpyxl.Workbook()
    ws = wb.active
    for c in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[c].width = 20
    c = ws.cell(1,4)
    c.value = str(year)+"年"+str(month)+"月"
    c.font = fontN
    for row in range(7):
        c = ws.cell(2, row+1)
        c.value = dayname[row]
        c.font = fontN
        c.alignment = openpyxl.styles.Alignment("center")
        if row == 6:
            c.font = fontB
            c.fill = fillB
        if row == 0:
            c.font = fontR
            c.fill = fillR
    for (col, week) in enumerate(cal.monthdayscalendar(year, month)):
        ws.row_dimensions[col+3].height = 50
        for (row, day) in enumerate(week):
            if day > 0 :
                c = ws.cell((col + 3), row+1)
                c.value = day
                c.font = fontN
                if row == 6:
                    c.font = fontB
                if row == 0:
                    c.font = fontR
    wb.save(savefile)   #Excel轉存檔案
    return "轉存"+savefile+"了。"
#--------------------^^^
def execute():
    value1 = values["input1"]
    value2 = values["input2"]
    #--------------------vvv
    #【4.執行函數】
    msg = makecalendar(value1, value2)
    #--------------------^^^
    window["text1"].update(msg)
#應用程式的介面
layout = [[sg.Text(label1, size=(14,1)), sg.Input(value1, key="input1")],
          [sg.Text(label2, size=(14,1)), sg.Input(value2, key="input2")],
          [sg.Button("執行", size=(20,1), pad=(5,15), bind_return_key=True)],
          [sg.Multiline(key="text1", size=(60,10))]]
#執行應用程式的處理
window = sg.Window(title, layout, font=(None,14))
while True:
    event, values = window.read()
    if event == None:
        break
    if event == "執行":
      execute()
window.close()
