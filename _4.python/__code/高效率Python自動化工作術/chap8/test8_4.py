import openpyxl

value1 = "tmp_output1.xlsx"

wb = openpyxl.Workbook()
ws = wb.active 

c = ws.cell(1,1)
c.value = "第1列,第1欄"
c.font = openpyxl.styles.Font(size = 24, color="0000CC") # 設定儲存格的文字大小與文字顏色
c.fill = openpyxl.styles.PatternFill("solid", fgColor="66CCFF") # 設定儲存格的背景色
c = ws.cell(5,3)
c.value = "第5列,第3欄"
c.font = openpyxl.styles.Font(size = 24, color="0000CC")
c.fill = openpyxl.styles.PatternFill("solid", fgColor="66CCFF")

ws.column_dimensions["A"].width = 20 #設定欄寬
ws.column_dimensions["C"].width = 20
ws.row_dimensions[1].height = 50 #設定列高
ws.row_dimensions[5].height = 50

wb.save(value1)     #Excel轉存檔案
