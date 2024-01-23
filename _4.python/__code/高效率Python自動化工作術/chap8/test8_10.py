import calendar
import openpyxl

year = 2022
month = 12
dayname = ["日","一","二","三","四","五","六"]

#【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = str(year)+"_"+str(month)+".xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    wb = openpyxl.Workbook()
    ws = wb.active
    c = ws.cell(1,4)
    c.value = str(year)+"年"+str(month)+"月"
    for col in range(7):                        #一週的每一天
        c = ws.cell(2, col+1)
        c.value = dayname[col]
    for (col, week) in enumerate(cal.monthdayscalendar(year, month)):
        for (row, day) in enumerate(week):
            if day > 0 :
                c = ws.cell((col + 3), row+1)
                c.value = day
    wb.save(savefile)   #Excel轉存檔案
    return "轉存"+savefile+"了。"

msg = makecalendar(year, month)
print(msg)