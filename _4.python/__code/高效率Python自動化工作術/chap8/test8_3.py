import openpyxl

value1 = "tmp_output0.xlsx"

wb = openpyxl.Workbook()
ws = wb.active 

c = ws.cell(1,1)
c.value = "第1列,第1欄"
c = ws.cell(5,3)
c.value = "第5列,第3欄"

wb.save(value1)     #Excel轉存檔案
