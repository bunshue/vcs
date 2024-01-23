import openpyxl

infile = "test.xlsx"
try:
    wb = openpyxl.load_workbook(infile)
    for sheetname in wb.sheetnames:                 #所有工作表
        sheet = wb[sheetname]
        for c in range(1, sheet.max_column+1):      #欄1〜最後
            for r in range(1, sheet.max_row+1):     #列1〜最後
                cell = sheet.cell(row=r, column=c)  #儲存格
                if cell.value != None:
                    print(cell.value)
except:
    print("程式執行失敗。")
    