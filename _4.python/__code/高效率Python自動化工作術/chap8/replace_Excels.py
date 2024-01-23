from pathlib import Path
import openpyxl

infolder = "testfolder"
value1 = "這個是"
value2 = "那個是"
value3 = "outputfolder"
ext = "*.xlsx"

#【函數: 搜尋Excel檔案之内的字串】
def replacefile(readfile, findword, newword, savefolder):
    try:
        msg = ""
        wb = openpyxl.load_workbook(readfile)
        cnt = 0
        for sheetname in wb.sheetnames:             #依序開啟所有工作表
            sheet = wb[sheetname]
            for c in range(1, sheet.max_column+1):  #各欄的
                for r in range(1, sheet.max_row+1): #各列的
                    cell = sheet.cell(row=r, column=c)
                    if type(cell.value)==str:
                        new_text = cell.value.replace(findword, newword)
                        cell.value = new_text           #置換儲存格的值
        savedir = Path(savefolder)
        savedir.mkdir(exist_ok=True)            #建立轉存資料夾
        filename = Path(readfile).name
        newname = savedir.joinpath(filename)
        wb.save(newname)                        #Excel轉存檔案
        msg = "在" + savefolder+"轉存"+ filename + "了喲。\n"
        return msg
    except:
        return readfile + "：程式執行失敗。"
#【函數：置換資料夾之內的文字檔】
def replacefiles(infolder, findword, newword, savefolder):
    msg = ""
    filelist = []
    for p in Path(infolder).glob(ext):  #將這個資料夾的檔案
        filelist.append(str(p))         #新增至列表
    for filename in sorted(filelist):   #再替每個檔案排序
        msg += replacefile(filename, findword, newword, savefolder)
    return msg

#【執行函數】
msg = replacefiles(infolder, value1, value2, value3)
print(msg)
