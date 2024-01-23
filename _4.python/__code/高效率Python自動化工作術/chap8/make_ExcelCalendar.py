import PySimpleGUI as sg

import calendar
import openpyxl

value1 = "2022"
value2 = "12"
dayname = ["日","一","二","三","四","五","六"]

fontN = openpyxl.styles.Font(size=24)
fontB = openpyxl.styles.Font(size=24, color="0000FF")
fontR = openpyxl.styles.Font(size=24, color="FF0000")
fillB = openpyxl.styles.PatternFill(patternType="solid", fgColor="AAAAFF")
fillR = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFAAAA")

#【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = str(year)+"_"+str(month)+".xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    wb = openpyxl.Workbook()
    ws = wb.active
    for c in ["A","B","C","D","E","F","G"]:
        ws.column_dimensions[c].width = 20
    c = ws.cell(1,4)
    c.value = str(year)+"年"+str(month)+"月"
    c.font = fontN
    for row in range(7):
        c = ws.cell(2, row+1)
        c.value = dayname[row]
        c.font = fontN
        c.alignment = openpyxl.styles.Alignment("center")
        if row == 6:
            c.font = fontB
            c.fill = fillB
        if row == 0:
            c.font = fontR
            c.fill = fillR
    for (col, week) in enumerate(cal.monthdayscalendar(year, month)):
        ws.row_dimensions[col+3].height = 50
        for (row, day) in enumerate(week):
            if day > 0 :
                c = ws.cell((col + 3), row+1)
                c.value = day
                c.font = fontN
                if row == 6:
                    c.font = fontB
                if row == 0:
                    c.font = fontR
    wb.save(savefile)   #Excel轉存檔案
    return "轉存"+savefile+"了。"

msg = makecalendar(value1, value2)
print(msg)