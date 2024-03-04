"""
相關抽出

各種檔案讀寫


"""

import os
import sys
import time
import random

print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxostudio.xlsx")  # 開啟 Excel 檔案

names = wb.sheetnames  # 讀取 Excel 裡所有工作表名稱
s1 = wb["工作表1"]  # 取得工作表名稱為「工作表1」的內容
s2 = wb.active  # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(names)
# 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(s1.title, s1.max_row, s1.max_column)
print(s2.title, s2.max_row, s2.max_column)


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb["工作表1"]
s2 = wb["工作表2"]

print(s1["A1"].value)  # 取出 A1 的內容
print(s1.cell(1, 1).value)  # 等同取出 A1 的內容
print(s2["B2"].value)  # 取出 B2 的內容
print(s2.cell(2, 2).value)  # 等同取出 B2 的內容


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb["工作表1"]
s2 = wb["工作表2"]


def get_values(sheet):
    arr = []  # 第一層串列
    for row in sheet:
        arr2 = []  # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr


print(get_values(s1))  # 印出工作表 1 所有內容
print(get_values(s2))  # 印出工作表 2 所有內容

"""
[[12, 34, 56, 78, 180, 180], [11, 22, 33, 44, 110, 110]]
[['a1', 'b1', 'c1'], ['a2', 'b2', 'c2'], ['a3', 'b3', 'c3'], ['a4', 'b4', 'c4'], ['a5', 'b5', 'c5']]
"""


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)

s1 = wb["工作表1"]
v = s1.iter_rows(min_row=1, min_col=1, max_col=2, max_row=2)  # 取出四格內容
print(v)
for i in v:
    for j in i:
        print(j.value)
"""
12
34
11
22
"""


print("------------------------------------------------------------")  # 60個

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(column_index_from_string("A"))  # 1
print(column_index_from_string("AA"))  # 27

print(get_column_letter(5))  # E
print(get_column_letter(100))  # CV


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
wb.save("empty.xlsx")  # 儲存檔案


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx")  # 開啟現有的 Excel 活頁簿物件
wb.save("new.xlsx")  # 儲存檔案


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx")  # 開啟 Excel 檔案

s1 = wb["工作表1"]  # 取得工作表名稱為「工作表1」的內容
s2 = wb.active  # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(
    s1.title, s1.max_row, s1.max_column
)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(
    s2.title, s2.max_row, s2.max_column
)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數

print(s1.sheet_properties)  # 印出工作表屬性


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]  # 開啟工作表 1
s2 = wb["工作表2"]  # 開啟工作表 2
s1.sheet_properties.tabColor = "ff0000"  # 修改工作表 1 頁籤顏色為紅色
s2.sheet_properties.tabColor = "ffff00"  # 修改工作表 2 頁籤顏色為黃色

wb.create_sheet("工作表3")  # 插入工作表 3 在最後方
wb.create_sheet("工作表1.5", 1)  # 插入工作表 1.5 在第二個位置 ( 工作表 1 和 2 的中間 )
wb.create_sheet("工作表0", 0)  # 插入工作表 0 在第一個位置

wb.copy_worksheet(s2)  # 複製工作表 2 放到最後方

s1.title = "oxxo"  # 修改工作表 1 的名稱為 oxxo
s2.title = "studio"  # 修改工作表 2 的名稱為 studio

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]  # 開啟工作表 1
s1["A1"].value = "apple"  # 儲存格 A1 內容為 apple
s1["A2"].value = "orange"  # 儲存格 A2 內容為 orange
s1["A3"].value = "banana"  # 儲存格 A3 內容為 banana
s1.cell(1, 2).value = 100  # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
s1.cell(2, 2).value = 200  # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
s1.cell(3, 2).value = 300  # 儲存格 B3 內容 ( row=3, column=2 ) 為 300

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s3 = wb.create_sheet("工作表3")  # 新增工作表 3
data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]  # 二維陣列資料
for i in data:
    s3.append(i)  # 逐筆添加到最後一列

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s2 = wb["工作表2"]  # 開啟工作表 2
data = [[1, 2], [3, 4]]  # 二維陣列資料
for y in range(len(data)):
    for x in range(len(data[y])):
        row = 2 + y  # 寫入資料的範圍從 row=2 開始
        col = 2 + x  # 寫入資料的範圍從 column=2 開始
        s2.cell(row, col).value = data[y][x]

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s2 = wb["工作表2"]
s2["d1"] = "=sum(a1:c1)"  # 寫入公式
s2["d2"] = "=sum(a2:c2)"  # 寫入公式
s2["d3"] = "=sum(a3:c3)"  # 寫入公式
s2["d4"] = "=sum(a4:c4)"  # 寫入公式
s2["d5"] = "=sum(a5:c5)"  # 寫入公式

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import openpyxl
from openpyxl.styles import Font, PatternFill  # 載入 Font 和 PatternFill 模組

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]
s1["e1"].font = Font(name="Arial", color="ff0000", size=30, bold=True)  # 設定 g1 儲存格的文字樣式
s1["f1"].fill = PatternFill(fill_type="solid", fgColor="DDDDDD")  # 設定 f1 儲存格的背景樣式

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

import csv

csvfile = open("csv-demo.csv")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列
print(data)
"""
[['name', 'id', 'color', 'price'],
 ['apple', '1', 'red', '10'],
 ['orange', '2', 'orange', '15'],
 ['grap', '3', 'purple', '20'],
 ['watermelon', '4', 'green', '30']]
"""


print("------------------------------------------------------------")  # 60個

import csv
import openpyxl

csvfile = open("csv-demo.csv")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列

wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
sheet = wb.create_sheet("csv")  # 建立空白的工作表
for i in data:
    sheet.append(i)  # 逐筆添加到最後一列

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

filename = "C:/_git/vcs/_1.data/______test_files1/picture1.jpg"

show = "不存在，可能檔案或路徑有誤"
if os.path.exists(filename):  # 判斷檔案或路徑是否存在
    show = "存在"
print("%s %s" % (filename, show))

print("------------------------------------------------------------")  # 60個

show = "不存在，可能檔案或路徑有誤"
if os.path.exists(filename):
    show = "存在"
    if os.path.isfile(filename):
        show += "，為檔案路徑"
    if os.path.isdir(filename):
        show += "，為目錄路徑"
print("%s %s" % (filename, show))

print("------------------------------------------------------------")  # 60個

dirPath = "tmp_PythonHw"
if os.path.exists(dirPath):
    print("目錄已存在或路徑有誤，無法建立")
else:
    os.mkdir(dirPath)
    pathname = os.path.split(dirPath)[0]  # 取得路徑
    filename = os.path.split(dirPath)[1]  # 取得目錄名稱
    print("於 %s 建立 %s 目錄" % (pathname, filename))

print("------------------------------------------------------------")  # 60個

dirPath = "tmp_PythonHw"
if os.path.exists(dirPath):
    os.rmdir(dirPath)
    pathname = os.path.split(dirPath)[0]  # 取得路徑
    filename = os.path.split(dirPath)[1]  # 取得目錄名稱
    print("刪除 %s 下的 %s 目錄" % (pathname, filename))
else:
    print("目錄不存在或路徑有誤，無法刪除")

print("------------------------------------------------------------")  # 60個

filename = "tmp_product.txt"
if os.path.exists(filename):
    f = open(filename, "r")
    print(f.readline(), end="")
    print(f.readline(), end="")
    print(f.readline(), end="")
    print("檔案讀取完成")
    f.close()
else:
    print("%s 檔案不存在" % (filename))

print("------------------------------------------------------------")  # 60個

filename = "tmp_product.txt"
if os.path.exists(filename):
    f = open(filename, "r")
    while True:
        row = f.readline()
        if row != "" and row != "\n":
            print(row.strip())
        else:
            print("檔案讀取完成")
            f.close()
            break
else:
    print("%s 檔案不存在" % (filename))

print("------------------------------------------------------------")  # 60個

filename = "tmp_product.txt"
if os.path.exists(filename):
    f = open(filename, "r")
    while True:
        row = f.readline()
        if row != "" and row != "\n":
            data = row.strip()
            listcol = data.split(",")
            colNum = 1
            for col in listcol:
                print(col, end="\t")
                colNum += 1
                if colNum == 4:
                    print("$%.2f" % (float(col) * 0.9), end="\t")
            print()
        else:
            print("檔案讀取完成")
            f.close()
            break
else:
    print("%s 檔案不存在" % (filename))

print("------------------------------------------------------------")  # 60個

filename = "tmp_product.txt"
if os.path.exists(filename):
    f = open(filename, "r")
    listProduct = f.readlines()
    for row in listProduct:
        data = row.strip()
        listcol = data.split(",")
        colNum = 1
        for col in listcol:
            print(col, end="\t")
            colNum += 1
            if colNum == 4:
                print("$%.2f" % (float(col) * 0.9), end="\t")
        print()
    print("檔案讀取完成")
    f.close()
else:
    print("%s 檔案不存在" % (filename))

print("------------------------------------------------------------")  # 60個

filename = "tmp_product.txt"
print("請輸入產品記錄：")
f = open(filename, "a")  # 以附加資料模式開啟檔案
pid = 123
name = "aaaaaa"
price = 456
data = str(pid) + "," + name + "," + str(price) + "\n"
f.write(data)
print("產品記錄新增成功")
f.close()

print("------------------------------------------------------------")  # 60個


#   讀檔作業
def fnRead():
    if os.path.exists(filename):
        f = open(filename, "r")
        listProduct = f.readlines()
        print("編號\t品名\t單價\t9折折扣")
        for row in listProduct:
            data = row.strip()
            listcol = data.split(",")
            colNum = 1
            for col in listcol:
                print(col, end="\t")
                colNum += 1
                if colNum == 4:
                    print("$%.2f" % (float(col) * 0.9), end="\t")
            print()
        print("產品讀取成功")
        f.close()
    else:
        print("%s 檔案不存在，請先使用新增功能建檔" % (filename))


# 新增作業
def fnWrite():
    f = open(filename, "a")  # 以附加資料模式開啟檔案
    pid = input("編號：")  # 輸入編號並指定給pid
    name = input("品名：")  # 輸入品名並指定給name
    price = int(input("單價："))  # 輸入單價並指定給price
    data = pid + "," + name + "," + str(price) + "\n"
    f.write(data)
    f.close()
    print("產品新增成功")


# 主程式
filename = "tmp_product.txt"

print("==DTC產品管理系統==")

# fnWrite()
fnRead()

print("------------------------------------------------------------")  # 60個

import csv

f = open("data/tScore.csv")  # 建立檔案物件f，此物件操作StudentScore.csv
csvReader = csv.reader(f, delimiter="\t")  # 建立reader物件使用 '\t' 當分隔符號
listData = list(csvReader)  # 使用list()函數將data轉成串列再指定給Reader
for row in listData:  # 使用巢狀迴圈將ListData串列逐欄印出
    for col in row:
        print(col, "   ", end="")
    print()
f.close()

print("------------------------------------------------------------")  # 60個

import csv

f = open("data/StudentScore.csv")
data = csv.DictReader(f)  # 使用DictReader ()方法取得csv檔資料並傳回data字典型別
for row in data:  # 逐一印出字典的內容
    print(row)
f.close()

print("------------------------------------------------------------")  # 60個

import csv

f = open("data/StudentScore.csv")
data = csv.DictReader(f)  # 使用DictReader ()方法取得csv檔資料並傳回data字典型別
print("學號\t姓名\t國文\t英語\t數學\t總分")
for row in data:  # 逐一印出字典的內容，並計算總分
    print(
        "{}\t{}\t{}\t{}\t{}\t{}".format(
            row["學號"],
            row["姓名"],
            row["國文"],
            row["英語"],
            row["數學"],
            (int(row["國文"]) + int(row["英語"]) + int(row["數學"])),
        )
    )
f.close()

print("------------------------------------------------------------")  # 60個

import csv

# searchName=input('請輸入學生姓名進行查詢成績：') #輸入查詢姓名

searchName = "david"

f = open("data/StudentScore.csv")
data = csv.DictReader(f)  # 使用DictReader ()方法取得csv檔資料並傳回data字典型別
for row in data:  # 逐一比對姓名是否符合searchName
    if row["姓名"] == searchName:
        print("{}成績資訊如下：".format(row["姓名"]))
        print("學號：{}".format(row["學號"]))
        print("國文：{}".format(row["國文"]))
        print("英語：{}".format(row["英語"]))
        print("數學：{}".format(row["數學"]))
        print("總分：{}".format((int(row["國文"]) + int(row["英語"]) + int(row["數學"]))))
        break  # 離開迴圈
else:  # 當迴圈沒有執行break，即會執行else區域，表示沒有找到符合姓名
    print("查無{}成績".format(searchName))
f.close()

print("------------------------------------------------------------")  # 60個

import csv

f = open("tmp_dictWriterProduct.csv", "w", newline="")
# 建立writer物件，同時指定欄位名稱
csvWriter = csv.DictWriter(f, fieldnames=["產品編號", "品名", "單價"])
csvWriter.writeheader()  # 寫入欄位名稱
csvWriter.writerow({"產品編號": "A02", "品名": "黑松沙士", "單價": 90})
# 寫入兩筆產品記錄到csv檔中
csvWriter.writerow({"產品編號": "A02", "品名": "草苺蛋糕", "單價": 120})
f.close()

print("------------------------------------------------------------")  # 60個

import json

listProduct = [
    {"編號": "P01", "品名": "五香豆干", "單價": 89},
    {"編號": "P02", "品名": "龍哥可樂", "單價": 20},
    {"編號": "P03", "品名": "阿才紅茶", "單價": 15},
]
f = open("tmp_product.json", "w", encoding="utf_8")
json.dump(listProduct, f, ensure_ascii=False, indent=4)
f.close()
print("JSON產品資料存檔成功")

print("------------------------------------------------------------")  # 60個

import json

listScore = [89, 100, 23, 78, 89]
print("listScore串列：", listScore)
print("listScore型別：", type(listScore))

jsonScore = json.dumps(listScore)
print("jsonScore字串：", jsonScore)
print("jsonScore型別：", type(jsonScore))

dictEmp = {"編號": "P01", "品名": "五香豆干", "單價": 89}
print("dictEmp字典：", dictEmp)
print("dictEmp型別：", type(dictEmp))

jsonEmp = json.dumps(dictEmp, ensure_ascii=False)
print("jsonEmp字串：", jsonEmp)
print("jsonEmp型別：", type(jsonEmp))

print("------------------------------------------------------------")  # 60個

import json

fruit = {"banana": "香蕉", "papaya": "木瓜", "apple": "蘋果"}
print(json.dumps(fruit, ensure_ascii=False))
print(json.dumps(fruit, ensure_ascii=False, sort_keys=True))
print(json.dumps(fruit, ensure_ascii=False, sort_keys=True, indent=4))

print("------------------------------------------------------------")  # 60個

import json

f = open("tmp_product.json", "r", encoding="utf_8")
pObj = json.load(f)
f.close()
print("====== DTC商店 ======")
for product in pObj:
    for key in product:
        print(key, "：", product[key])
    print("=" * 20)

print("------------------------------------------------------------")  # 60個

import json

jsonStr = """
{"編號":"E01","姓名": "王小明",
"性別": true, "電話":["0912345678","0978123321"]}
"""
print("jsonStr字串：", jsonStr)
print("jsonStr型別：", type(jsonStr))
print()
pObj = json.loads(jsonStr)
print("pObj物件：", pObj)
print("pObj型別：", type(pObj))
for key in pObj:
    print(key, ":", pObj[key], " value的型別：", type(pObj[key]))

print("------------------------------------------------------------")  # 60個

import json


# 新增員工記錄函式
def fnCreate():
    uid = input("編號：")
    if uid in listUid:
        print("編號重複，無法在記憶體中新增員工記錄")
        return
    name = input("姓名︰")
    salary = int(input("薪資︰"))
    newMember = {"編號": uid, "姓名": name, "薪資": salary}
    listMember.append(newMember)
    listUid.append(uid)
    print("記憶體新增編號 %s 的員工記錄" % (uid))


# 修改員工記錄函式
def fnUpdate():
    uid = input("編號：")
    for member in listMember:
        if member["編號"] == uid:
            name = input("姓名︰")
            salary = int(input("薪資︰"))
            newMember = {"編號": uid, "姓名": name, "薪資": salary}
            cIndex = listMember.index(member)
            listMember[cIndex] = newMember
            print("記憶體修改編號 %s 的員工記錄" % (uid))
            break
    else:
        print("查無編號，無法修改記憶體中的員工記錄")


# 刪除員工記錄函式
def fnDelete():
    uid = input("編號：")
    for member in listMember:
        if member["編號"] == uid:
            listMember.remove(member)
            listUid.remove(uid)
            print("記憶體刪除編號 %s 的員工記錄" % (uid))
            break
    else:
        print("查無編號，無法刪除記憶體中的員工記錄")


# 顯示員工記錄函式
def fnPrintMember():
    if len(listMember) == 0:
        print("記憶體中目前無員工記錄")
        return
    for member in listMember:
        for key in member:
            print(member[key], end="\t")
        print()


# 員工記錄儲存至MemberInfo.json的函式
def fnSaveJSONFile():
    f = open(jsonfile, "w", encoding="utf_8")
    json.dump(listMember, f, ensure_ascii=False, indent=4)
    f.close()
    print("記憶體中的員工記錄成功儲存至 %s 檔案" % (jsonfile))


jsonfile = "MemberInfo.json"
listMember = []
listUid = []
if os.path.exists(jsonfile):
    f = open("MemberInfo.json", "r", encoding="utf_8")
    listMember = json.load(f)
    listUid = []
    for member in listMember:
        listUid.append(member["編號"])
    f.close()

"""
# 主程式
print("======= DTC員工管理系統 =======")
while True:
   option=int(input('系統功能->1.新增 2.修改 3.刪除 4.查詢 5.儲存JSON檔案 其他.離開：'))
   if option==1:
       fnCreate()
   elif option==2:
       fnUpdate()
   elif option==3:
       fnDelete()
   elif option==4:
       fnPrintMember()
   elif option==5:
       fnSaveJSONFile()
   else:
       print("離開系統")
       break;
"""

print("------------------------------------------------------------")  # 60個

import csv

f = open("data/StudentScore.csv")  # 建立檔案物件f，此物件操作StudentScore.csv
csvReader = csv.reader(f)  # 使用csv的reader()方法取得檔案物件f的資料並傳回Reader物件
listData = list(csvReader)  # 使用list()函數將data轉換串列再指定給listData
for row in listData:  # 將二維串列listData 逐列印出
    print(row)
f.close()  # 關閉檔案

print("------------------------------------------------------------")  # 60個

import csv

f = open("data/StudentScore.csv")  # 建立檔案物件f，此物件操作StudentScore.csv
csvReader = csv.reader(f)  # 使用csv的reader()方法取得檔案物件f的資料並傳回Reader物件
listData = list(csvReader)  # 使用list()函數將csvReader轉成串列再指定給listData
for row in listData:  # 使用巢狀迴圈將ListData串列逐欄印出
    for col in row:
        print(col, "  ", end="")
    print()
f.close()

print("------------------------------------------------------------")  # 60個

import csv

f = open("tmp_writerProduct.csv", "w", newline="")  # 開啟writerProduct.csv檔案
csvWriter = csv.writer(f)  # 建立writer物件，物件名稱為csvWriter
# 建立listProduct二維串列有兩筆產品
listProduct = [["B01", "小林煎餅", "78"], ["B02", "五香豆干", "90"]]
# 寫入一維串列當做標題
csvWriter.writerow(["編號", "品名", "單價"])
csvWriter.writerows(listProduct)  # 將二維串列的兩筆產品寫入csv內
f.close()  # 關閉檔案

print("------------------------------------------------------------")  # 60個

"""
import csv

listProduct=["","",""] # 建立listProduct串列，用來存放一筆產品記錄
while True:
     option = input("功能選單：1.新增 2.查詢 3.離開：")
     if option=="1":
         # 以附加模式開啟tProduct.csv檔案
         f=open('tProduct.csv','a', newline='') 
         csvWriter=csv.writer(f) 
         listProduct[0] = input("編號：")   #listProduct[0]存放編號
         listProduct[1] = input("品名：")   #listProduct[1]存放品名
         listProduct[2] = input("單價：")   #listProduct[2]存放單價
         csvWriter.writerow(listProduct)
         print("新增成功")
         f.close()
     elif option=="2" :
         # 以讀檔模式開啟tProduct.csv檔案
         f=open ('tProduct.csv')
         data=csv.DictReader(f)  
         print("編號\t品名\t單價")
         for row in data:   		
             print("{}\t{}\t{}".format(row['編號'],row['品名'],row['單價']))
         f.close()
     else:
         break
"""
print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個
