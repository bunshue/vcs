"""
一本精通 - Python 範例應用大全
"""

import os
import sys
import time
import random

# 有順序

print("------------------------------------------------------------")  # 60個
print("Python 常用標準函式庫")
print("------------------------------------------------------------")  # 60個


# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code001.py

# Copyright © https://steam.oxxostudio.tw

import time


def aa():
    i = 0
    while i < 5:
        i = i + 1
        time.sleep(0.5)
        print("A:", i)


def bb():
    i = 0
    while i < 50:
        i = i + 10
        time.sleep(0.5)
        print("B:", i)


aa()  # 先執行 aa()
bb()  # aa() 結束後才會執行 bb()

"""
A: 1
A: 2
A: 3
A: 4
A: 5
B: 10
B: 20
B: 30
B: 40
B: 50
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code002.py

# Copyright © https://steam.oxxostudio.tw

import threading
import time


def aa():
    i = 0
    while i < 5:
        i = i + 1
        time.sleep(0.5)
        print("A:", i)


def bb():
    i = 0
    while i < 50:
        i = i + 10
        time.sleep(0.5)
        print("B:", i)


a = threading.Thread(target=aa)  # 建立新的執行緒
b = threading.Thread(target=bb)  # 建立新的執行緒

a.start()  # 啟用執行緒
b.start()  # 啟用執行緒

"""
A: 1
B: 10
A: 2
B: 20
A: 3
B: 30
A: 4
B: 40
A: 5
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code003.py

import threading
import time


def aa():
    i = 0
    while i < 5:
        i = i + 1
        time.sleep(0.5)
        print("A:", i)


def bb():
    i = 0
    while i < 50:
        i = i + 10
        time.sleep(0.5)
        print("B:", i)


def cc():
    i = 0
    while i < 500:
        i = i + 100
        time.sleep(0.5)
        print("C:", i)


a = threading.Thread(target=aa)
b = threading.Thread(target=bb)
c = threading.Thread(target=cc)

a.start()
b.start()
a.join()  # 加入等待 aa() 完成的方法
b.join()  # 加入等待 bb() 完成的方法
c.start()  # 當 aa() 與 bb() 都完成後，才會開始執行 cc()

"""
A: 1
B: 10
A: 2
B: 20
A: 3
B: 30
A: 4
B: 40
A: 5
B: 50
C: 100 <--- A B 都結束後才開始
C: 200
C: 300
C: 400
C: 500
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code004.py

# Copyright © https://steam.oxxostudio.tw

import time


def aa():
    i = 0
    while i < 5:
        i = i + 1
        time.sleep(0.5)
        print("A:", i)


def bb():
    i = 0
    while i < 100:
        i = i + 10
        time.sleep(0.5)
        print("B:", i)


def cc():
    i = 0
    while i < 500:
        i = i + 100
        time.sleep(0.5)
        print("C:", i)


a = threading.Thread(target=aa)
b = threading.Thread(target=bb)
c = threading.Thread(target=cc)

a.start()
b.start()
a.join()  # 加入等待 aa() 完成的方法
c.start()  # 當 aa() 完成後，就會開始執行 cc()

"""
A: 1
B: 10
A: 2
B: 20
A: 3
B: 30
A: 4
B: 40
A: 5
B: 50
C: 100 <--- A 結束就開始
B: 60
C: 200
B: 70
C: 300
B: 80
C: 400
B: 90
C: 500
B: 100
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code005.py

# Copyright © https://steam.oxxostudio.tw

import threading
import time


def aa():
    lock.acquire()  # 鎖定
    i = 0
    while i < 5:
        i = i + 1
        time.sleep(0.5)
        print("A:", i)
        if i == 2:
            lock.release()  # i 等於 2 時解除鎖定


def bb():
    lock.acquire()  # 鎖定
    i = 0
    while i < 50:
        i = i + 10
        time.sleep(0.5)
        print("B:", i)
    lock.release()


lock = threading.Lock()  # 建立 Lock
a = threading.Thread(target=aa)
b = threading.Thread(target=bb)

a.start()
b.start()

"""
A: 1
A: 2
B: 10
A: 3
B: 20
A: 4
B: 30
A: 5
B: 40
B: 50
"""

print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code006.py

# Copyright © https://steam.oxxostudio.tw

import threading
import time


def aa():
    event.wait()  # 等待事件被觸發
    event.clear()  # 觸發後將事件回歸原本狀態
    for i in range(1, 6):
        print("A:", i)
        time.sleep(0.5)


def bb():
    for i in range(10, 60, 10):
        if i == 30:
            event.set()  # 觸發事件
        print("B:", i)
        time.sleep(0.5)


event = threading.Event()  # 註冊事件
a = threading.Thread(target=aa)
b = threading.Thread(target=bb)

a.start()
b.start()

"""
B: 10
B: 20
B: 30
A: 1
B: 40
A: 2
B: 50
A: 3
A: 4
A: 5
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code007.py

# Copyright © https://steam.oxxostudio.tw

import threading
import time


def aa():
    i = 0
    while True:
        event_a.wait()  # 等待 event_a 被觸發
        event_a.clear()  # 還原 event_a 狀態
        for i in range(1, 6):
            print(i)
            time.sleep(0.5)
        event_b.set()  # 觸發 event_b


def bb():
    while True:
        input("輸入任意內容")
        event_a.set()  # 觸發 event_a
        event_b.wait()  # 等待 event_b 被觸發
        event_b.clear()  # 還原 event_b 狀態


event_a = threading.Event()  # 註冊 event_a
event_b = threading.Event()  # 註冊 event_b
a = threading.Thread(target=aa)
b = threading.Thread(target=bb)

a.start()
b.start()

"""
輸入任意內容a
1
2
3
4
5
輸入任意內容b
1
2
3
4
5
輸入任意內容
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch09\code008.py

# Copyright © https://steam.oxxostudio.tw

import time
from concurrent.futures import ThreadPoolExecutor

a = True  # 定義 a 為 True


def run():
    global a  # 定義 a 是全域變數
    while a:  # 如果 a 為 True
        print(123)  # 不斷顯示 123
        time.sleep(1)  # 每隔一秒


def keyin():
    global a  # 定義 a 是全域變數
    if input() == "a":
        a = False  # 如果輸入的是 a，就讓 a 為 False，停止 run 函式中的迴圈


executor = ThreadPoolExecutor()
e1 = executor.submit(run)
e2 = executor.submit(keyin)
executor.shutdown()


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("Python 基礎範例")
print("------------------------------------------------------------")  # 60個


# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code001.py

# Copyright © https://steam.oxxostudio.tw

while True:
    try:
        num = float(input("請輸入用電度數："))
        output = 0
        if num <= 200:
            output = num * 3.2
        elif num > 200 and num <= 300:
            output = 200 * 3.2 + (num - 200) * 3.4
        else:
            output = 200 * 3.2 + 100 * 3.4 + (num - 300) * 3.6
        print(f"用電 {num} 度共 {output} 元")
    except:
        break


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code002.py

# Copyright © https://steam.oxxostudio.tw

c = int(input("輸入 1 ( 攝氏 ) 或 2 ( 華氏 )："))  # 使用變數 c 記錄攝氏還是華氏
t = int(input("輸入溫度數值："))  # 使用變數 t 記錄要轉換的數值

if c == 1:
    print(f"攝氏 {t} 度等於華氏 {9/5*t+32} 度")  # 套用攝氏轉華氏公式
else:
    print(f"華氏 {t} 度等於攝氏 {(t-32)*5/9} 度")  # 套用華氏轉攝氏公式


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code003.py

# Copyright © https://steam.oxxostudio.tw

c = int(input("輸入 1 ( 公分 ) 或 2 ( 英吋 )："))  # 使用變數 c 記錄公分還是英吋
length = int(input("輸入長度數值："))  # 使用變數 length 記錄數值

if c == 1:
    # 套用轉換公式
    print(f"{length} 公分等於 {length*0.394} 英吋 ( {length*0.03281} 英尺、{length*0.01094} 碼 )")
else:
    # 套用轉換公式
    print(f"{length} 英吋等於 {length*2.54} 公分 ( {length/12} 英尺、{length/36} 碼 )")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code004.py

# Copyright © https://steam.oxxostudio.tw

c = int(input("輸入 1 ( 公分 ) 或 2 ( 英吋 )："))
length = int(input("輸入長度數值："))

print("|cm   |m    |ich  |foot |yd   |")  # 印出說明
print("|-----|-----|-----|-----|-----|")  # 印出分隔線

if c == 1:
    print("|", end="")  # 印出表格左側的框線
    print(f"{str(length):<5.5s}", end="|")
    print(f"{str(length*0.01):<5.5s}", end="|")
    print(f"{str(length*0.394):<5.5s}", end="|")
    print(f"{str(length*0.03281):<5.5s}", end="|")
    print(f"{str(length*0.01094):<5.5s}", end="|")
else:
    print("|", end="")
    print(f"{str(length*2.54):<5.5s}", end="|")
    print(f"{str(length*0.0254):<5.5s}", end="|")
    print(f"{str(length):<5.5s}", end="|")
    print(f"{str(length/12):<5.5s}", end="|")
    print(f"{str(length/36):<5.5s}", end="|")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code005.py

# Copyright © https://steam.oxxostudio.tw

year = int(input(">"))  # 使用變數 year 紀錄使用者輸入的年份
if year % 4 == 0:  # 如果除以 4 能整除
    if year % 100 == 0:  # 如果除以 100 能整除
        if year % 400 == 0:  # 如果除以 400 能整除，就是閏年
            print(f"{year} 是閏年")
        else:
            print(f"{year} 是平年")
    else:
        print(f"{year} 是閏年")
else:
    print(f"{year} 是平年")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code006.py

# Copyright © https://steam.oxxostudio.tw

year = int(input(">"))
text = "平年"  # 新增變數 text 預設平年
if year % 4 == 0:
    text = "閏年"  # 如果除以 4 能整除，將 text 改為閏年
if year % 100 == 0:
    text = "平年"  # 如果除以 100 能整除，將 text 改為平年
if year % 400 == 0:
    text = "閏年"  # 如果除以 400 能整除，將 text 改為閏年
print(f"{year} 是{text}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code007.py

# Copyright © https://steam.oxxostudio.tw

text = input("請輸入一串英文或數字：")  # 新增 text 變數，記錄輸入的字串
repeat = []  # 新增 repeat 變數為空串列
not_repeat = []  # 新增 not_repeat 變數為空串列
for i in text:  # 使用 for 迴圈，依序取出每個字元
    a = text.count(i, 0, len(text))  # 判斷字元在字串中出現的次數
    if a > 1 and i not in repeat:  # 如果次數大於 1，且沒有存在 repeat 串列中
        repeat.append(i)  # 將字元加入 repeat 串列
    if a == 1 and i not in not_repeat:  # 如果次數等於 1，且沒有存在 not_repeat 串列中
        not_repeat.append(i)  # 將字元加入 not_repeat 串列

print(repeat)
print(not_repeat)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code008.py

# Copyright © https://steam.oxxostudio.tw

import math  # import math 標準函式模組

text = input("輸入文字：")  # 讓使用者輸入文字
length = len(text)  # 取得輸入的文字長度
center = math.floor(length / 2)  # 取出中間值
if length % 2 == 0:  # 如果除以 2 餘數為 0，表示偶數
    print(f"{text[center-1:center+1]}")  # 取出中間兩個字元
else:
    print(f"{text[center]}")  # 如果是奇數，取出中間一個字元


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code009.py

# Copyright © https://steam.oxxostudio.tw

import re

# 要轉換的字串
text = """請 求 您 幫 我 oxxo.studio 去 除 空 白 ok ？
但是要保留換行 可以 嗎 ，(        哈哈哈 )( 啊哈)
統一便利超商 (711) 的括號之間也要有空白喔！
寫作規    範就是這 麼 100% 的龜毛～
"""

# 取得中文字和英文單字的正規表達式
# [a-zA-Z0-9]+ 表示開頭是英文字母後面連接一串字母或數字
regex = re.compile(r"[\u4E00-\u9FFF\uFF00-\uFFFF\u0021-\u002F\n]|[a-zA-Z0-9]+")

# 根據正規表達式，將每個中文字、標點符號和英文單字變成串列
arr = re.findall(regex, text)

# 使用空格合併串列
text = " ".join(arr)
print(text)

"""
請 求 您 幫 我 oxxo . studio 去 除 空 白 ok ？
但 是 要 保 留 換 行 可 以 嗎 ， ( 哈 哈 哈 ) ( 啊 哈 )
統 一 便 利 超 商 ( 711 ) 的 括 號 之 間 也 要 有 空 白 喔 ！
寫 作 規 範 就 是 這 麼 100 % 的 龜 毛 ～
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code010.py

# Copyright © https://steam.oxxostudio.tw

import re

# 輸入字符串
text = """請 求 您 幫 我 oxxo.studio 去 除 空 白 ok ？
但是要保留換行 可以 嗎 ，(        哈哈哈 )( 啊哈)
統一便利超商 (711) 的括號之間也要有空白喔！
寫作規    範就是這 麼 100% 的龜毛～
"""

regex = re.compile(r"[\u4E00-\u9FFF\uFF00-\uFFFF\u0021-\u002F\n]|[a-zA-Z0-9]+")
arr = re.findall(regex, text)
text = " ".join(arr)

regex = re.compile(r"(?<=[^a-zA-Z0-9\u0021-\u002E])(\x20)(?=[^a-zA-Z0-9\u0021-\u002E])")
text = re.sub(regex, "", text)

regex = re.compile(r"(\x20)(?=[\(\%\uFF00-\uFFFF])")
text = re.sub(regex, "", text)

text = text.replace(" . ", ".")
print(text)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code011.py

# Copyright © https://steam.oxxostudio.tw

import random

a = []  # 建立空串列
while len(a) < 6:  # 使用 while 迴圈，直到串列的長度等於 6 就停止
    b = random.randint(1, 49)  # 取出 1～49 得隨機整數
    if b not in a:  # 判斷如果 a 裡面沒有 b
        a.append(b)  # 將 b 放入 a
print(a)  # [34, 18, 31, 11, 47, 46]


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code012.py

# Copyright © https://steam.oxxostudio.tw

import random

a = set()  # 建立空集合
while len(a) < 6:  # 使用 while 迴圈，直到集合的長度等於 6 就停止
    b = random.randint(1, 49)  # 取出 1～49 得隨機整數
    a.add(b)  # 將隨機數加入集合
print(a)  # {34, 41, 48, 49, 19, 30}


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code013.py

# Copyright © https://steam.oxxostudio.tw

import random

a = random.sample(range(1, 50), 6)
# 從包含 1～49 數字的串列中，取出六個不重複的數字變成串列
print(a)  # [9, 39, 10, 8, 25, 43]


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code014.py

# Copyright © https://steam.oxxostudio.tw

import time

n = 10
for i in range(n + 1):
    print(f"\r倒數 {n-i} 秒", end="")
    time.sleep(1)
print("\r時間到", end="")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code015.py

# Copyright © https://steam.oxxostudio.tw

import time

n = 20  # 設定進度條總長
for i in range(n + 1):
    print(f'\r[{"█"*i}{" "*(n-i)}] {i*100/n}%', end="")  # 輸出不換行的內容
    time.sleep(0.5)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code016.py

# Copyright © https://steam.oxxostudio.tw

import time

n = 100
icon = "⋮⋰⋯⋱"  # 建立旋轉的符號清單
for i in range(n + 1):
    print(f"\r{icon[i%4]} {i*100/n}%", end="")
    time.sleep(0.1)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code017.py

# Copyright © https://steam.oxxostudio.tw

a = 15  # 新增變數 a，設定金字塔有幾層
b = a * 2 + 1  # 新增變數 b，計算底部有幾個星星
for i in range(1, b, 2):  # 使用 for 迴圈，從 1～b，每隔 2 個一數
    move = round((b - i) / 2) - 1  # 計算星星的位移空白 ( 要將星星都置中 )
    print(f" " * move, end="")  # 印出星星前方的位移空白 ( 不換行 )
    print("*" * i)  # 加上「幾個星星」( 乘以 i )


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code018.py

# Copyright © https://steam.oxxostudio.tw

a = 15
b = a * 2 + 1
for i in range(1, b, 4):  # 改成 4 個一數，金字塔每一層就會增加 2，高度也會減半
    move = round((b - i) / 2) - 1
    print(f" " * move, end="")
    print("*" * i)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code019.py

# Copyright © https://steam.oxxostudio.tw

a = 15  # 新增變數 a，設定金字塔有幾層
for i in range(1, a + 1):  # 使用 for 迴圈，重複指定的層數
    print(" " * (a - i) + "*" * (2 * i - 1))
    # ' ' * (a-i) 表示星星數越少，前面空白越多
    # '*' * (2*i-1) 串接後方星星的數量


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code020.py

# Copyright © https://steam.oxxostudio.tw

a = 10  # 要產生的金字塔層數
for i in range(1, a + 1):  # 使用 for 迴圈，重複 1～10 ( a+1 ) 的數字
    print(" " * 3 * (a - i), end="")  # 根據不同的層數，讓第一個數字前方增加指定的空白字元 ( 後方不換行 )
    for j in range(1, i + 1):  # 第二層 for 迴圈，重複不同層內的數字
        if j == 1:  # 如果是第一個數字
            print(j, end="")  # 單純印出數字即可 ( 後方不換行 )
        else:  # 如果是第二個以後的數字
            print(f"{j:>3d}", end="")  # 格式化數字，使其寬度為 3，並靠右對齊 ( 後方不換行 )
    for j in range(i - 1, 0, -1):  # 剛剛的 for 迴圈是由小到大，加入另外一個由大到小的迴圈
        print(f"{j:>3d}", end="")  # 格式化數字，使其寬度為 3，並靠右對齊 ( 後方不換行 )
    print("")  # 最後執行換行的 print


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code021.py

# Copyright © https://steam.oxxostudio.tw

a = 10
for i in range(1, a + 1):
    print(" " * 3 * (a - i), end="")
    for j in range(0, i):  # ragne 改成從 0 開始，因為 2 的 0 次方等於 1
        k = 2**j  # 計算 2 的幾次方
        if k == 1:
            print(k, end="")
        else:
            print(f"{k:>3d}", end="")
    for j in range(i - 2, -1, -1):  # 修改 range，使其最後一位數為 0
        k = 2**j  # 計算 2 的幾次方
        print(f"{k:>3d}", end="")
    print("")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code022.py

# Copyright © https://steam.oxxostudio.tw

a = 10  # 要產生的金字塔層數
b = 1  # 提供 while 迴圈停止的依據
while b <= a:  # 如果 b <= a 就讓 while 迴圈繼續
    n = 1  # 設定從 1 開始
    print(" " * 3 * (a - b), end="")  # 根據不同的層數，讓第一個數字前方增加指定的空白字元 ( 後方不換行 )
    while n <= b:  # 第二層 while 迴圈，如果 n <= b 就讓 while 迴圈繼續
        if n == 1:  # 如果是第一個數字
            print(n, end="")  # 單純印出數字即可 ( 後方不換行 )
        else:  # 如果是第二個以後的數字
            print(f"{n:>3d}", end="")  # 格式化數字，使其寬度為 3，並靠右對齊 ( 後方不換行 )
        n = n + 1  # 將 n 增加 1
    while n > 2:  # 剛剛的 while 迴圈是由小到大，加入另外一個由大到小的迴圈
        print(f"{n-2:>3d}", end="")  # 格式化數字，使其寬度為 3，並靠右對齊 ( 後方不換行 )
        n = n - 1  # 將 n 減少 1
    print("")  # 最後執行換行的 print
    b = b + 1  # 將 b 增加 1


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code023.py

# Copyright © https://steam.oxxostudio.tw

import random

a = random.randint(1, 99)  # 產生 1～99 的隨機整數
b = int(input("輸入 1～99 的數字："))  # 讓使用者輸入數字，使用 int 轉換成數字
while a != b:  # 使用 while 迴圈，如果 a 不等於 b，就不斷繼續
    if b < a:
        b = int(input("數字太小囉！再試一次吧："))  # 如果 b<a，提示數字太小
    elif b > a:
        b = int(input("數字太大囉！再試一次吧："))  # 如果 b>a，提示數字太大
print("答對囉！")  # 如果 b=a 會停止 while 迴圈，顯示正確答案


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code024.py

# Copyright © https://steam.oxxostudio.tw

import random

a = random.randint(1, 99)
b = int(input("輸入 1～99 的數字："))
while True:
    if b < a:
        b = int(input("數字太小囉！再試一次吧："))
    elif b > a:
        b = int(input("數字太大囉！再試一次吧："))
    else:
        print("答對囉！")
        break


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code025.py

# Copyright © https://steam.oxxostudio.tw

import random

answer = random.sample(range(1, 10), 4)
print(answer)
a = b = n = 0  # 設定 a、b、n 三個變數，預設值 0
while a != 4:  # 使用 while 迴圈，直到 a 等於 4 才停止
    a = b = n = 0  # 每次重複時將 a、b、n 三個變數再次設定為 0
    user = list(input("輸入四個數字："))  # 讓使用者輸入數字，並透過 list 轉換成串列
    for i in user:  # 使用 for 迴圈，將使用者輸入的數字一一取出
        if int(user[n]) == answer[n]:  # 因為使用者輸入的是「字串」，透過 int 轉換成數字，和答案串列互相比較
            a += 1  # 如果位置和內容都相同，就將 a 增加 1
        else:
            if int(i) in answer:  # 如果位置不同，但答案裡有包含使用者輸入的數字
                b += 1  # 就將 b 增加 1
        n += 1  # 因為輸入的每個數字都要判斷，將 n 增加 1
    output = ",".join(user).replace(",", "")  # 四個數字都判斷後，使用 join 將串列合併成字串
    print(f"{output}: {a}A{b}B")
print("答對了！")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code026.py_

import random
import time  # import time 模組

answer = random.sample(range(1, 10), 4)
print(answer)
a = b = n = 0
num = 0  # 新增 num 變數為 0，作為計算次數使用
t = time.time()  # 新增 t 變數為現在的時間
while a != 4:
    num += 1  # 每次重複時將 num 增加 1
    a = b = n = 0
    user = list(input("輸入四個數字："))
    for i in user:
        if int(user[n]) == answer[n]:
            a += 1
        else:
            if int(i) in answer:
                b += 1
        n += 1
    output = ",".join(user).replace(",", "")
    print(f"{output}: {a}A{b}B")
t = round((time.time() - t), 3)  # 當 a 等於 4 時，計算結束和開始的時間差
print(f"答對了！總共猜了 {num} 次，用了 {t} 秒")  # 印出對應的文字


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code027.py

# Copyright © https://steam.oxxostudio.tw

import datetime

now = datetime.datetime.now().strftime("%H:%M:%S")
print(now)  # 14:30:23


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code028.py

# Copyright © https://steam.oxxostudio.tw

import datetime
import time

while True:
    now = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"\r{now}", end="")  # 前方加上 \r
    time.sleep(1)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code029.py

# Copyright © https://steam.oxxostudio.tw

import datetime
import time


def timezone(h):
    GMT = datetime.timezone(datetime.timedelta(hours=h))  # 取得時區
    return datetime.datetime.now(tz=GMT).strftime("%H:%M:%S")  # 回傳該時區的時間


# 六個時區的名稱與時差
local = {"倫敦": 1, "台灣": 8, "日本": 9, "紐約": -4, "洛杉磯": -7, "紐西蘭": 12}

while True:
    print("\r", end="")  # 開始時將游標移到開頭
    # 讀取 local 的 key
    for i in local:
        now = timezone(local[i])  # 根據 key 的 value 取得時間
        print(f"{i}>{timezone(local[i])}  ", end="")
    time.sleep(1)
    # 倫敦>08:43:09  台灣>15:43:09  日本>16:43:09  紐約>03:43:09  洛杉磯>00:43:09  紐西蘭>19:43:09


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code030.py

# Copyright © https://steam.oxxostudio.tw

h = float(input("請輸入身高(cm)：")) / 100
# 使用 float 轉換成浮點數後除以 100 ( 因為身高可能會有小數點 )

w = float(input("請輸入體重(kg)："))
# 使用 float 轉換成浮點數 ( 因為體重可能會有小數點 )

bmi = w / (h * h)  # 套用公式計算
print(f"你的 BMI 數值為：{bmi}")  # 你的 BMI 數值為：23.044982698961938


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code031.py

# Copyright © https://steam.oxxostudio.tw

h = float(input("請輸入身高(cm)：")) / 100
w = float(input("請輸入體重(kg)："))
bmi = round(w / (h * h), 3)  # 使用 round 四捨五入到小數點三位
if bmi < 18.5:  # 使用邏輯判斷
    note = "你太輕囉！"
elif bmi >= 18.5 and bmi <= 25:
    note = "你的體重正常！"
else:
    note = "你有點太重囉～"
print(f"你的 BMI 數值為：{bmi}，{note}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code032.py

# Copyright © https://steam.oxxostudio.tw

import datetime  # import datetime 標準函式

today = datetime.date.today()  # 使用 datetime.date 取得今天的日期
age = input("輸入生日 ( YYYY/MM/DD )：")  # 讓使用者輸入生日，格式為 YYYY/MM/DD
age_list = age.split("/")  # 將使用者輸入的日期，使用「/」拆成串列
year = today.year - int(age_list[0])  # 用今天的年份，減去使用者的生日年份 ( 年份差 )
month = today.month - int(age_list[1])  # 用今天的月份，減去使用者生日的月份 ( 月份差 )
if month < 0:  # 如果月份差的數字小於零，表示生日還沒到
    year = year - 1  # 將年份差減少 1 ( 表示跨了一年 )
    month = 12 + month  # 將月份差改成 12 + 月份差
day_list = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]  # 建立一個每個月有多少天的串列
day = today.day - int(age_list[2])  # 用今天的日期，點去使用者生日的日期 ( 月份差 )
if day < 0:  # 如果月份差的數字小於 0，表示生日還沒到
    month = month - 1  # 將月份差減少 1
    if month < 0:  # 如果月份差減少後小於 0
        year = year - 1  # 再將年份差減少 1 ( 表示跨了一年 )
        month = 12 + month  # 將月份差改成 12 + 月份差
    day = day_list[month] + day  # 將日期差改成該月的天數 + 日期差

print(f"{year} 歲 {month} 個月 {day} 天")  # 印出現在幾歲幾個月又幾天


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code033.py

# Copyright © https://steam.oxxostudio.tw

import datetime
import calendar  # import calendar 模組

today = datetime.date.today()
age = input("輸入生日 ( YYYY/MM/DD )：")
age_list = age.split("/")
year = today.year - int(age_list[0])
month = today.month - int(age_list[1])
if month < 0:
    year = year - 1
    month = 12 + month
day_list = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
if calendar.isleap(today.year):  # 判斷如果是閏年
    day_list[1] = 29  # 就將二月份的天數改成 29 天
day = today.day - int(age_list[2])
if day < 0:
    month = month - 1
    if month < 0:
        year = year - 1
        month = 12 + month
    day = day_list[month] + day

print(f"{year} 歲 {month} 個月 {day} 天")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code034.py

# Copyright © https://steam.oxxostudio.tw

import random

local_table = {
    "A": 10,
    "B": 11,
    "C": 12,
    "D": 13,
    "E": 14,
    "F": 15,
    "G": 16,
    "H": 17,
    "I": 34,
    "J": 18,
    "K": 19,
    "L": 20,
    "M": 21,
    "N": 22,
    "O": 35,
    "P": 23,
    "Q": 24,
    "R": 25,
    "S": 26,
    "T": 27,
    "U": 28,
    "V": 29,
    "W": 32,
    "X": 30,
    "Y": 31,
    "Z": 33,
}
for j in range(20):  # 使用 20 次的 for 迴圈
    local = random.choice(list(local_table.keys()))

    check_arr = list(str(local_table[local]))
    check_arr[0] = int(check_arr[0])
    check_arr[1] = int(check_arr[1]) * 9

    sex = random.choice([1, 2])
    check_arr.append(sex * 8)

    nums_str = ""
    for i in range(7):
        nums = random.randint(0, 9)
        nums_str = nums_str + str(nums)
        check_arr.append(nums * (7 - i))

    check_num = 10 - sum(check_arr) % 10
    if check_num == 10:
        check_num = 0

    id_number = str(local) + str(sex) + nums_str + str(check_num)
    print(id_number)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code035.py

# Copyright © https://steam.oxxostudio.tw

local_table = {
    "A": 10,
    "B": 11,
    "C": 12,
    "D": 13,
    "E": 14,
    "F": 15,
    "G": 16,
    "H": 17,
    "I": 34,
    "J": 18,
    "K": 19,
    "L": 20,
    "M": 21,
    "N": 22,
    "O": 35,
    "P": 23,
    "Q": 24,
    "R": 25,
    "S": 26,
    "T": 27,
    "U": 28,
    "V": 29,
    "W": 32,
    "X": 30,
    "Y": 31,
    "Z": 33,
}
id_number = input("輸入身分證字號：")
check = False  # 新增 check=False 變數，與 while 迴圈搭配
while True:  # 使用 while 迴圈
    id_arr = list(id_number)  # 新增 id_arr 變數，將身分證字號轉換成串列存入
    if len(id_arr) != 10:
        break  # 判斷如果 id_arr 長度不等於 10，就跳出 while 迴圈
    local = str(local_table[id_arr[0]])  # 將對應的二位數字轉換成字串
    check_arr = list(local)  # 將字串轉換成陣列，例如 '10' 會轉換成 ['1','0']
    check_arr[0] = int(check_arr[0])  # 將串列中的第一個項目轉換成數字
    check_arr[1] = int(check_arr[1]) * 9  # 將串列中的第二個項目轉換成數字
    sex = id_arr[1]  # 取得第二碼數字
    if sex != "1" and sex != "2":
        break  # 判斷如果不是 '1' 也不是 '2' 就跳出 while 迴圈
    check_arr.append(int(sex) * 8)  # 將 sex 內容轉換成數字並乘以 8，存入 check_arr 裡
    for i in range(7):  # 使用 for 迴圈，重複七次
        check_arr.append(int(id_arr[i + 2]) * (7 - i))  # 每次重複，按照檢查碼程式，將數字乘以對應的數值
    check_num = 10 - sum(check_arr) % 10  # 計算使用者輸入的檢查碼
    if check_num != int(id_arr[9]):
        break  # 如果檢查碼不相同，跳出 while 迴圈
    check = True  # 如果迴圈都沒有跳出，讓 check 等於 True。
    break  # 結束後跳出迴圈

if check == False:  # while 迴圈結束後，如果 check 等於 Fasle，表示身分證字號錯誤
    print("身分證字號格式錯誤")
else:
    print("身分證字號正確")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code036.py

# Copyright © https://steam.oxxostudio.tw

local_table = {
    "A": 10,
    "B": 11,
    "C": 12,
    "D": 13,
    "E": 14,
    "F": 15,
    "G": 16,
    "H": 17,
    "I": 34,
    "J": 18,
    "K": 19,
    "L": 20,
    "M": 21,
    "N": 22,
    "O": 35,
    "P": 23,
    "Q": 24,
    "R": 25,
    "S": 26,
    "T": 27,
    "U": 28,
    "V": 29,
    "W": 32,
    "X": 30,
    "Y": 31,
    "Z": 33,
}
while True:  # 新增 while 迴圈，就可以重複輸入
    id_number = input("輸入身分證字號：")
    check = False
    while True:
        try:  # 使用 try
            id_arr = list(id_number)
            if len(id_arr) != 10:
                break
            local = str(local_table[id_arr[0]])
            check_arr = list(local)
            check_arr[0] = int(check_arr[0])
            check_arr[1] = int(check_arr[1]) * 9
            sex = id_arr[1]
            if sex != "1" and sex != "2":
                break
            check_arr.append(int(sex) * 8)
            for i in range(7):
                check_arr.append(int(id_arr[i + 2]) * (7 - i))
            check_num = 10 - sum(check_arr) % 10
            if check_num != int(id_arr[9]):
                break
            check = True
            break
        except:  # 使用 except，如果發生例外狀況，跳出迴圈
            break

    if check == False:
        print("身分證字號格式錯誤")
    else:
        print("身分證字號正確")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code037.py

# Copyright © https://steam.oxxostudio.tw

num = input("輸入你的發票號碼：")
ns = "05701942"  # 特別獎
n1 = "97718570"  # 特獎
n2 = ["88400675", "73475574", "53038222"]  # 頭獎
if num == ns:
    print("對中 1000 萬元！")  # 對中特別獎
if num == n1:
    print("對中 200 萬元！")  # 對中特獎
# 頭獎判斷
for i in n2:
    if num == i:
        print("對中 20 萬元！")  # 對中頭獎
        break
    if num[-7:] == i[-7:]:
        print("對中 4 萬元！")  # 末七碼相同
        break
    if num[-6:] == i[-6:]:
        print("對中 1 萬元！")  # 末六碼相同
        break
    if num[-5:] == i[-5:]:
        print("對中 4000 元！")  # 末五碼相同
        break
    if num[-4:] == i[-4:]:
        print("對中 1000 元！")  # 末四碼相同
        break
    if num[-3:] == i[-3:]:
        print("對中 200 元！")  # 末三碼相同
        break


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code038.py

# Copyright © https://steam.oxxostudio.tw

table = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}  # 轉換對照表
roman = [i for i in input()]  # 將輸入的羅馬數字變成串列
r = roman[::-1]  # 反轉串列
output = table[r[0]]  # 讓 output 先等於第一個數字
for i in range(1, len(r)):  # 從第二個數字開始依序取到最後一個數字
    if table[r[i]] < table[r[i - 1]]:  # 如果後面數字比較小
        output = output - table[r[i]]  # 讓 output 減去後面的數字
    else:
        output = output + table[r[i]]  # 如果後面數字比較大，讓 output 加上後面的數字
print(output)

# 輸入 IVMVIIMVVMVM 就會得到 3994

print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch10\code039.py

# Copyright © https://steam.oxxostudio.tw

num_table = [
    [1000, "M"],
    [900, "CM"],
    [500, "D"],
    [400, "CD"],
    [100, "C"],
    [90, "XC"],
    [50, "L"],
    [40, "XL"],
    [10, "X"],
    [9, "IX"],
    [5, "V"],
    [4, "IV"],
    [1, "I"],
]  # 建立對照表
num = int(input())  # 將輸入的文字轉換成數字
output = ""  # 設定輸出的 output 字串
for i in num_table:  # 依序判斷對照表中每個數字
    a = divmod(num, i[0])  # 取得商 ( a[0] ) 和餘數 ( a[1] )
    if a[0] != 0:  # 如果 a[0] 不為 0
        num = a[1]  # 取得餘數繼續往下除
        output = output + i[1] * a[0]  # 組合字串
print(output)


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("Python 數學範例")
print("------------------------------------------------------------")  # 60個


# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code001.py

# Copyright © https://steam.oxxostudio.tw

a = input("請輸入兩個數字 ( 格式 a,b )：")  # 新增變數 a，內容是使用者輸入的兩個數字，數字以逗號分隔
b = a.split(",")  # 新增變數 b，內容使用 split 根據逗號將數字拆開為串列
b1 = int(b[0])  # 使用 int 將第一個值轉換為「數字」
b2 = int(b[1])  # 使用 int 將第二個值轉換為「數字」
print(f"{b1} + {b2} = {b1+b2}")  # 印出四則運算的結果
print(f"{b1} - {b2} = {b1-b2}")
print(f"{b1} x {b2} = {b1*b2}")
print(f"{b1} / {b2} = {b1/b2}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code002.py

# Copyright © https://steam.oxxostudio.tw

a = input("請輸入兩個數字 ( 格式 a,b )：")
b = a.split(",")
b1 = int(b[0])
b2 = int(b[1])
print(f"{b1} + {b2} = {b1+b2}")
print(f"{b1} - {b2} = {b1-b2}")
print(f"{b1} x {b2} = {b1*b2}")
print(f"{b1} / {b2} = {round(b1/b2,3)}")  # 使用 round 四捨五入到小數點三位
print(f"{b2} + {b1} = {b2+b1}")
print(f"{b2} - {b1} = {b2-b1}")
print(f"{b2} x {b1} = {b2*b1}")
print(f"{b2} / {b1} = {round(b2/b1,3)}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code003.py

# Copyright © https://steam.oxxostudio.tw

a = input("請輸入數字 ( 格式 a,b,c... )：")  # 新增變數 a，內容是使用者輸入的多個數字，數字以逗號分隔
b = a.split(",")  # 新增變數 b，內容使用 split 根據逗號將數字拆開為串列
output = 0  # 設定 output 從 0 開始
for i in b:  # 使用 for 迴圈，依序取出 b 串列的每個項目
    output += int(i)  # 將 output 的數值加上每個項目 ( 使用 int 將項目轉換成數字 )

print(f"數字總和為：{output}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code004.py

# Copyright © https://steam.oxxostudio.tw

while output != 0:  # 使用 while 迴圈，如果 output 等於 0 才會停止
    a = input("請輸入數字 ( 格式 a,b,c... )：")
    b = a.split(",")
    output = 0
    for i in b:
        output += int(i)
    print(f"數字總和為：{output}")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code005.py

# Copyright © https://steam.oxxostudio.tw

nums = [int(i) for i in input().split(",")]  # 使用串列生成式，將輸入的數字轉換成串列
result = sum(nums)  # 將串列內的數字加總
print(result)  # 印出結果


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code006.py

# Copyright © https://steam.oxxostudio.tw


def fib(n):  # 建立函式 fib，帶有參數 n
    if n > 1:  # 如果 n 大於 1
        return fib(n - 1) + fib(n - 2)  # 使用遞迴
    return n


for i in range(20):  # 產生 20 個數字
    print(
        fib(i), end=","
    )  # 0,1,1,2,3,5,8,13,21,34,55,89,144,233,377,610,987,1597,2584,4181


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code007.py

# Copyright © https://steam.oxxostudio.tw

n = int(input())  # 輸入要產生的數字數量
arr = []  # 建立一個空串列，記錄數字
for i in range(n):  # 使用 for 迴圈，重複指定的數字
    if i == 0:  # 如果 i 等於 0，a 為 0
        a = 0
    elif i == 1:  # 如果 i 等於 1，a 為 1
        a = 1
        arr = [0, 1]  # 將串列設定為 [0, 1]
    else:  # 如果 i 大於 1
        a = arr[0] + arr[1]  # a 等於串列的兩個數字相加
        del arr[0]  # 刪除串列的第一個項目
        arr.append(a)  # 將 a 加入串列成為第二個項目
    print(a, end=",")  # 0,1,1,2,3,5,8,13,21,34,55,89,144,233,377,610,987,1597,2584,4181


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code008.py

# Copyright © https://steam.oxxostudio.tw

for a in range(1, 10):  # 讓 a 從 1 執行到 9
    for b in range(1, 10):  # 讓 b 從 1 執行到 9
        print(f"{a}x{b}={a*b}", end=" ")
    # 使用格式化字串，印出產生對應的字串，最後加上 end=' '表示不換行
    print("")
    # 內層迴圈執行結束後，執行 print('') 會換行顯示
"""
1x1=1 1x2=2 1x3=3 1x4=4 1x5=5 1x6=6 1x7=7 1x8=8 1x9=9
2x1=2 2x2=4 2x3=6 2x4=8 2x5=10 2x6=12 2x7=14 2x8=16 2x9=18
3x1=3 3x2=6 3x3=9 3x4=12 3x5=15 3x6=18 3x7=21 3x8=24 3x9=27
4x1=4 4x2=8 4x3=12 4x4=16 4x5=20 4x6=24 4x7=28 4x8=32 4x9=36
5x1=5 5x2=10 5x3=15 5x4=20 5x5=25 5x6=30 5x7=35 5x8=40 5x9=45
6x1=6 6x2=12 6x3=18 6x4=24 6x5=30 6x6=36 6x7=42 6x8=48 6x9=54
7x1=7 7x2=14 7x3=21 7x4=28 7x5=35 7x6=42 7x7=49 7x8=56 7x9=63
8x1=8 8x2=16 8x3=24 8x4=32 8x5=40 8x6=48 8x7=56 8x8=64 8x9=72
9x1=9 9x2=18 9x3=27 9x4=36 9x5=45 9x6=54 9x7=63 9x8=72 9x9=81
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code009.py

# Copyright © https://steam.oxxostudio.tw

for a in range(1, 10):
    for b in range(1, 10):
        print("{}x{}={}".format(a, b, a * b), end=" ")
    print("")

for a in range(1, 10):
    for b in range(1, 10):
        print("%dx%d=%d" % (a, b, a * b), end=" ")
    print("")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code010.py

# Copyright © https://steam.oxxostudio.tw

for a in range(1, 10):
    for b in range(1, 10):
        if (a * b) < 10:
            print(f"{a}x{b}={a*b}", end="  ")  # 如果 axb<10，讓結尾增加一個空白
        else:
            print(f"{a}x{b}={a*b}", end=" ")
    print("")

    """
    1x1=1  1x2=2  1x3=3  1x4=4  1x5=5  1x6=6  1x7=7  1x8=8  1x9=9
    2x1=2  2x2=4  2x3=6  2x4=8  2x5=10 2x6=12 2x7=14 2x8=16 2x9=18
    3x1=3  3x2=6  3x3=9  3x4=12 3x5=15 3x6=18 3x7=21 3x8=24 3x9=27
    4x1=4  4x2=8  4x3=12 4x4=16 4x5=20 4x6=24 4x7=28 4x8=32 4x9=36
    5x1=5  5x2=10 5x3=15 5x4=20 5x5=25 5x6=30 5x7=35 5x8=40 5x9=45
    6x1=6  6x2=12 6x3=18 6x4=24 6x5=30 6x6=36 6x7=42 6x8=48 6x9=54
    7x1=7  7x2=14 7x3=21 7x4=28 7x5=35 7x6=42 7x7=49 7x8=56 7x9=63
    8x1=8  8x2=16 8x3=24 8x4=32 8x5=40 8x6=48 8x7=56 8x8=64 8x9=72
    9x1=9  9x2=18 9x3=27 9x4=36 9x5=45 9x6=54 9x7=63 9x8=72 9x9=81
    """


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code011.py

# Copyright © https://steam.oxxostudio.tw

a = b = int(input("請輸入一個正整數："))  # 新增 a 和 b 變數，等於使用者輸入的數字
output = ""  # 新增 output 變數，作為輸出的文字
while True:  # 使用 while 迴圈
    for i in range(2, (a + 1)):  # 使用 for 迴圈
        if i == b:  # 如果 i 等於 b，表示是質數，跳出 for 迴圈
            break
        if a % i == 0:  # 如果可以被 i 整除，表示不是質數
            output += f"{i}"  # 設定 output 輸出的內容
            a = int(a / i)  # 重新將 a 設定為商
            break  # 跳出 for 迴圈
    if a == 1 or a == b:  # 如果商等於 1 或是質數，跳出 while 迴圈
        break
    else:
        output += "*"  # 否則在 output 後方加上 * 號，繼續 while 迴圈
if b == a and b != 1:
    print(f"{b} 是質數")  # while 迴圈結束後，如果 a 等於 b，印出質數的文字
elif b == 1:
    print(f"{b} 不是質數，也不能質因數分解")  # 如果輸入的是 1 或 2
else:
    print(f"{b}={output}")  # 否則印出質因數分解的結果


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code012.py

# Copyright © https://steam.oxxostudio.tw

a = range(2, 100)  # 產生 2～100 的串列
print(*a)
b = [i for i in a if i == a[0] or i % a[0] > 0]  # 找出第一個質數，並將串列裡該質數的倍數剔除
print(*b)
c = [i for i in b if i == b[1] or i % b[1] > 0]  # 找出第二個質數，並將串列裡該質數的倍數剔除
print(*c)
d = [i for i in c if i == c[2] or i % c[2] > 0]  # 找出第三個質數，並將串列裡該質數的倍數剔除
print(*d)


"""
2 3 4 5 6 7 8 9 10 11 12 13 14 15 16 17 18 19 20 21 22 23 24 25 26 27 28 29 30 31 32 33 34 35 36 37 38 39 40 41 42 43 44 45 46 47 48 49 50 51 52 53 54 55 56 57 58 59 60 61 62 63 64 65 66 67 68 69 70 71 72 73 74 75 76 77 78 79 80 81 82 83 84 85 86 87 88 89 90 91 92 93 94 95 96 97 98 99
2 3 5 7 9 11 13 15 17 19 21 23 25 27 29 31 33 35 37 39 41 43 45 47 49 51 53 55 57 59 61 63 65 67 69 71 73 75 77 79 81 83 85 87 89 91 93 95 97 99
2 3 5 7 11 13 17 19 23 25 29 31 35 37 41 43 47 49 53 55 59 61 65 67 71 73 77 79 83 85 89 91 95 97
2 3 5 7 11 13 17 19 23 29 31 37 41 43 47 49 53 59 61 67 71 73 77 79 83 89 91 97
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code013.py

# Copyright © https://steam.oxxostudio.tw

a = range(2, 100)  # 產生 2～100 的串列
p = 0  # 設定 p 從 0 開始 ( 從 a[p] 也就是第一個項目開始 )


def g():  # 定義一個函式 g
    global p, a  # 設定 p 和 a 是全域變數
    if p < len(a):  # 如果 p 小於 a 的長度 ( 依序取值到 a 的最後一個項目 )
        a = [i for i in a if i == a[p] or i % a[p] > 0]  # 重新設定 a 為移除倍數後的串列
        p = p + 1  # p 增加 1
        g()  # 重新執行函式 g


g()  # 執行函式 g
print(*a)  # 印出 a ( 使用 * 將串列打散印出 )

"""
2 3 5 7 11 13 17 19 23 29 31 37 41 43 47 53 59 61 67 71 73 79 83 89 97
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code014.py

# Copyright © https://steam.oxxostudio.tw


def gg(max):  # 定義一個 gg 函式
    s = set()  # 設定一個空集合
    for n in range(2, max):  # 從 range(2, max) 當中開始依序找質數
        if all(n % i > 0 for i in s):  # 判斷如果 i 已經存在於集合，且除以集合中的值會有餘數 ( 整除表示非質數 )
            s.add(n)  # 將該數字加入集合 ( 表示質數 )
            yield n  # 使用 yield 記錄狀態


print(*gg(100))  # 印出結果

"""
2 3 5 7 11 13 17 19 23 29 31 37 41 43 47 53 59 61 67 71 73 79 83 89 97
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code015.py

# Copyright © https://steam.oxxostudio.tw

input_str = input("輸入數字 ( 逗號分隔 )：")  # 讓使用者輸入數字，數字間用逗號分隔
nums = input_str.split(",")  # 將輸入的文字，用逗號拆分成串列
for i in range(len(nums)):  # 將串列的每個項目轉換成文字
    nums[i - 1] = int(nums[i - 1])
nums.sort(reverse=True)  # 將串列從大到小排序
result = nums[0]  # 設定「暫定的最小公倍數」為最大的數字
while True:  # 執行 while 迴圈
    a = 0  # 新增 a 變數，當作餘數使用
    for i in nums:  # 依序取出串列中的每個數字
        a = result % i  # 用「暫定的最小公倍數」除以每個數字，求出餘數
        if a != 0:  # 如果餘數不為 0，跳出 for 迴圈再來一次
            break
    if a == 0:  # 如果全部餘數都為 0，跳出 while 迴圈
        break
    else:
        result = result + nums[0]  # 如果餘數不為 0，就將「暫定的最小公倍數」加上最大的數字，然後再來一次
print(result)  # while 迴圈結束後，印出最小公倍數


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch11\code016.py

# Copyright © https://steam.oxxostudio.tw

input_str = input("輸入數字 ( 逗號分隔 )：")  # 讓使用者輸入數字，數字間用逗號分隔
nums_arr = input_str.split(",")  # 將輸入的文字，用逗號拆分成串列
for i in range(len(nums_arr)):  # 將串列的每個項目轉換成文字
    nums_arr[i - 1] = int(nums_arr[i - 1])
nums_arr.sort()  # 將串列從小到大排序
result = nums_arr[0]  # 建立變數 result，內容為輸入的第一個數字 ( 數字的最小值 )
arr = [result, 1]  # 建立一個變數 arr 為串列，內容預設為 [ 輸入的最小值, 1 ]
for i in range(2, result + 1):  # 使用 for 迴圈，找出 result 數字的每個因數
    if result % i == 0:  # 找因數的方法，將 result 依序除以 2、3、4...result
        result = int(result / i)  # 如果餘數為 0 ( 整除 )，表示這個數字為因數
        arr.append(i)  # 將因數加入 arr 串列中，並更新 result 為除以因數的數值
        arr.append(result)  # 也將 result 加入 arr 串列 ( 因為商也算是因數 )
arr.sort(reverse=True)  # 完成後將 arr 從大到小排序

for j in arr:  # 依序取出 arr 串列中的每個數字
    a = 0  # 建立 a 變數，記錄餘數
    output = 1  # 建立 output 變數，記錄最大公因數 ( 預設 1 )
    for i in nums_arr:  # 依序將輸入的數字除以 arr 串列中的數字
        a = a + i % j  # 將餘數加入 a 變數 ( 如果沒有餘數，a 就一直會是 0 )
        output = j  # 將 output 等於目前的因數
    if a == 0:  # 如果 a 為 0 表示都整除，將 result 等於 output
        result = output
        break
print(result)  # 印出最大公因數


print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個
print("Python 實際應用")
print("------------------------------------------------------------")  # 60個


# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code001.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui

myScreenshot = pyautogui.screenshot()
myScreenshot.save("圖片路徑\圖片名稱.png")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code002.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui

myScreenshot = pyautogui.screenshot(region=(x1, y1, x2, y2))
myScreenshot.save("圖片路徑\圖片名稱.png")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code003.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui
from time import sleep

for i in range(5):
    myScreenshot = pyautogui.screenshot()
    myScreenshot.save(f"./test{i}.png")
    sleep(2)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code004.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui
import requests

myScreenshot = pyautogui.screenshot()  # 截圖
myScreenshot.save("./test.png")  # 儲存為 test.png

url = "https://notify-api.line.me/api/notify"
token = "你的權杖"
headers = {"Authorization": "Bearer " + token}  # 設定 LINE Notify 權杖
data = {"message": "測試一下！"}  # 設定 LINE Notify message ( 不可少 )
image = open("./test.png", "rb")  # 以二進位方式開啟圖片
imageFile = {"imageFile": image}  # 設定圖片資訊
data = requests.post(url, headers=headers, data=data, files=imageFile)  # 發送 LINE Notify


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code005.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui
import requests
import time


# 定義截圖的函式
def screenshot():
    myScreenshot = pyautogui.screenshot()
    myScreenshot.save("./test.png")

    t = time.time()  # 取得到目前為止的秒數
    t1 = time.localtime(t)  # 將秒數轉換為 struct_time 格式的時間
    now = time.strftime("%Y/%m/%d %H:%M:%S", t1)  # 輸出為特定格式的文字
    sendLineNotify(now)  # 執行發送 LINE Notify 的函式，發送的訊息為時間


# 定義發送 LINE Notify 的函式
def sendLineNotify(msg):
    url = "https://notify-api.line.me/api/notify"
    token = "你的權杖"
    headers = {"Authorization": "Bearer " + token}
    data = {"message": msg}
    image = open("./test.png", "rb")
    imageFile = {"imageFile": image}
    data = requests.post(url, headers=headers, data=data, files=imageFile)


# 使用for 迴圈，每隔五秒截圖發送一次
for i in range(5):
    screenshot()
    time.sleep(5)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code006.py

# Copyright © https://steam.oxxostudio.tw

import glob

images = glob.glob("./demo/*")
print(images)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code007.py

# Copyright © https://steam.oxxostudio.tw

import glob
import os

images = glob.glob("./demo/*")
print(images)

n = 1  # 設定名稱從 1 開始
for i in images:
    os.rename(i, f"./demo/img-{n:03d}.jpg")  # 改名時，使用字串格式化的方式進行三位數補零
    n = n + 1  # 每次重複時將 n 增加 1


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code008.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import pdfplumber

pdf = pdfplumber.open("oxxostudio.pdf")  # 開啟 pdf
print(pdf.pages)  # [<Page:1>, <Page:2>, <Page:3>]，共有三頁


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code009.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import pdfplumber

pdf = pdfplumber.open("oxxostudio.pdf")
page = pdf.pages[0]  # 讀取第一頁
text = page.extract_text()  # 取出文字
print(text)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code010.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import pdfplumber

pdf = pdfplumber.open("oxxostudio.pdf")
page = pdf.pages[1]  # 讀取第二頁
table = page.extract_table()  # 取出表格
print(table)
pdf.close()


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code011.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import pdfplumber

pdf = pdfplumber.open("test.pdf", password="12345678")  # 輸入密碼
page = pdf.pages[0]
text = page.extract_text()
print(table)
pdf.close()


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code012.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import pdfplumber

pdf = pdfplumber.open("oxxostudio.pdf")
page = pdf.pages[0]
text = page.extract_text()
print(text)
pdf.close()

f = open("test.txt", "w+")  # 使用 w+ 模式開啟 test.txt
f.write(text)  # 寫入內容
f.close()  # 關閉 test.txt


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code013.py

# Copyright © https://steam.oxxostudio.tw

import pdfplumber

pdf = pdfplumber.open("oxxostudio.pdf")
page = pdf.pages[1]
table = page.extract_table()
print(table)
pdf.close()

import csv

csvfile = open("test-csv.csv", "w+")  # 建立 CSV 檔案
write = csv.writer(csvfile)  # 建立寫入物件
for i in table:
    write.writerow(i)  # 讀取表格每一列寫入 CSV
print("ok")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code014.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("oxxostudio.pdf", password="1234")  # 開啟 pdf
pdf_pwd = Pdf.open("oxxostudio-pwd.pdf", password="1234")  # 開啟需要密碼的 pdf
print(pdf)
print(pdf_pwd)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code015.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf, Permissions, Encryption

pdf = Pdf.open("oxxostudio-pwd.pdf", password="1234")  # 開啟密碼為 1234 的 pdf
no_extracting = Permissions(extract=False)
# 儲存為密碼是 qqqq 的 pdf
pdf.save(
    "new.pdf", encryption=Encryption(user="qqqq", owner="qqqq", allow=no_extracting)
)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code016.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("oxxostudio.pdf")  # 開啟 pdf
pages = pdf.pages  # 將每一頁的內容變成串列
output = Pdf.new()  # 建立新的 pdf 物件
output.pages.append(pages[0])  # 添加頁面內容
output.save("new.pdf")  # 儲存為新的 pdf


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code017.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("oxxostudio.pdf")
pages = pdf.pages
n = 1
for i in pages:
    output = Pdf.new()
    output.pages.append(i)
    output.save(f"new_{n}.pdf")  # 格式化檔案名稱
    n = n + 1  # 編號加 1


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code018.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("test.pdf")
pages = pdf.pages
output = Pdf.new()
output.pages.extend(pages[1:3])  # 改用 extend，放入特定範圍的頁面
output.save("new.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code019.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf1 = Pdf.open("oxxo_1.pdf")  # 讀取第一份 pdf
pdf2 = Pdf.open("oxxo_2.pdf")  # 讀取第二份 pdf
pdf3 = Pdf.open("oxxo_3.pdf")  # 讀取第三份 pdf

output = Pdf.new()  # 建立新的 pdf 物件
output.pages.append(pdf1.pages[0])  # 添加第一頁到第一份
output.pages.append(pdf2.pages[0])  # 添加第一頁到第二份
output.pages.append(pdf3.pages[0])  # 添加第一頁到第三份
output.save("output.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code020.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf1 = Pdf.open("oxxo_more_1.pdf")  # 讀取第一份多頁面 pdf
pdf2 = Pdf.open("oxxo_more_2.pdf")  # 讀取第一份多頁面 pdf
pdf3 = Pdf.open("oxxo_more_1.pdf")  # 讀取第一份多頁面 pdf

output = Pdf.new()
output.pages.extend(pdf1.pages)  # 添加所有頁面到第一份
output.pages.extend(pdf2.pages)  # 添加所有頁面到第二份
output.pages.extend(pdf3.pages)  # 添加所有頁面到第三份
output.save("output.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code021.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf1 = Pdf.open("oxxostudio.pdf")  # 開啟第一份 pdf
pdf2 = Pdf.open("new.pdf")  # 開啟第二份 pdf
pdf1.pages.insert(1, pdf2.pages[0])  # 在第一份的第一頁後方，插入第二份的第一頁
pdf1.save("output.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code022.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("oxxosudio.pdf")  # 開啟 pdf
del pdf.pages[1:2]  # 刪除第二頁
pdf.save("output.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code023.py_

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf1 = Pdf.open("oxxosudio.pdf")  # 開啟第一份 pdf
pdf2 = Pdf.open("new.pdf")  # 開啟第二份 pdf
pdf1.pages[2] = pdf2.pages[0]  # 將第一份的第三頁，換成第一份的第一頁
pdf1.save("output.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code024.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from pikepdf import Pdf

pdf = Pdf.open("output.pdf")
pdf.pages.reverse()  # 反轉 pdf
pdf.save("output2.pdf")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code025.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxostudio.xlsx")  # 開啟 Excel 檔案

names = wb.sheetnames  # 讀取 Excel 裡所有工作表名稱
s1 = wb["工作表1"]  # 取得工作表名稱為「工作表1」的內容
s2 = wb.active  # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(names)
# 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(s1.title, s1.max_row, s1.max_column)
print(s2.title, s2.max_row, s2.max_column)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code026.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb["工作表1"]
s2 = wb["工作表2"]

print(s1["A1"].value)  # 取出 A1 的內容
print(s1.cell(1, 1).value)  # 等同取出 A1 的內容
print(s2["B2"].value)  # 取出 B2 的內容
print(s2.cell(2, 2).value)  # 等同取出 B2 的內容


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code027.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)  # 設定 data_only=True 只讀取計算後的數值

s1 = wb["工作表1"]
s2 = wb["工作表2"]


def get_values(sheet):
    arr = []  # 第一層串列
    for row in sheet:
        arr2 = []  # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr


print(get_values(s1))  # 印出工作表 1 所有內容
print(get_values(s2))  # 印出工作表 2 所有內容

"""
[[12, 34, 56, 78, 180, 180], [11, 22, 33, 44, 110, 110]]
[['a1', 'b1', 'c1'], ['a2', 'b2', 'c2'], ['a3', 'b3', 'c3'], ['a4', 'b4', 'c4'], ['a5', 'b5', 'c5']]
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code028.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("test.xlsx", data_only=True)

s1 = wb["工作表1"]
v = s1.iter_rows(min_row=1, min_col=1, max_col=2, max_row=2)  # 取出四格內容
print(v)
for i in v:
    for j in i:
        print(j.value)
"""
12
34
11
22
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code029.py

# Copyright © https://steam.oxxostudio.tw

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

print(column_index_from_string("A"))  # 1
print(column_index_from_string("AA"))  # 27

print(get_column_letter(5))  # E
print(get_column_letter(100))  # CV


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code030.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
wb.save("empty.xlsx")  # 儲存檔案


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code031.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx")  # 開啟現有的 Excel 活頁簿物件
wb.save("new.xlsx")  # 儲存檔案


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code032.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx")  # 開啟 Excel 檔案

s1 = wb["工作表1"]  # 取得工作表名稱為「工作表1」的內容
s2 = wb.active  # 取得開啟試算表後立刻顯示的工作表 ( 範例為工作表 2 )

print(
    s1.title, s1.max_row, s1.max_column
)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數
print(
    s2.title, s2.max_row, s2.max_column
)  # 印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數

print(s1.sheet_properties)  # 印出工作表屬性


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code033.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]  # 開啟工作表 1
s2 = wb["工作表2"]  # 開啟工作表 2
s1.sheet_properties.tabColor = "ff0000"  # 修改工作表 1 頁籤顏色為紅色
s2.sheet_properties.tabColor = "ffff00"  # 修改工作表 2 頁籤顏色為黃色

wb.create_sheet("工作表3")  # 插入工作表 3 在最後方
wb.create_sheet("工作表1.5", 1)  # 插入工作表 1.5 在第二個位置 ( 工作表 1 和 2 的中間 )
wb.create_sheet("工作表0", 0)  # 插入工作表 0 在第一個位置

wb.copy_worksheet(s2)  # 複製工作表 2 放到最後方

s1.title = "oxxo"  # 修改工作表 1 的名稱為 oxxo
s2.title = "studio"  # 修改工作表 2 的名稱為 studio

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code034.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]  # 開啟工作表 1
s1["A1"].value = "apple"  # 儲存格 A1 內容為 apple
s1["A2"].value = "orange"  # 儲存格 A2 內容為 orange
s1["A3"].value = "banana"  # 儲存格 A3 內容為 banana
s1.cell(1, 2).value = 100  # 儲存格 B1 內容 ( row=1, column=2 ) 為 100
s1.cell(2, 2).value = 200  # 儲存格 B2 內容 ( row=2, column=2 ) 為 200
s1.cell(3, 2).value = 300  # 儲存格 B3 內容 ( row=3, column=2 ) 為 300

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code035.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s3 = wb.create_sheet("工作表3")  # 新增工作表 3
data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]  # 二維陣列資料
for i in data:
    s3.append(i)  # 逐筆添加到最後一列

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code036.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s2 = wb["工作表2"]  # 開啟工作表 2
data = [[1, 2], [3, 4]]  # 二維陣列資料
for y in range(len(data)):
    for x in range(len(data[y])):
        row = 2 + y  # 寫入資料的範圍從 row=2 開始
        col = 2 + x  # 寫入資料的範圍從 column=2 開始
        s2.cell(row, col).value = data[y][x]

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code037.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s2 = wb["工作表2"]
s2["d1"] = "=sum(a1:c1)"  # 寫入公式
s2["d2"] = "=sum(a2:c2)"  # 寫入公式
s2["d3"] = "=sum(a3:c3)"  # 寫入公式
s2["d4"] = "=sum(a4:c4)"  # 寫入公式
s2["d5"] = "=sum(a5:c5)"  # 寫入公式

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code038.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import openpyxl
from openpyxl.styles import Font, PatternFill  # 載入 Font 和 PatternFill 模組

wb = openpyxl.load_workbook("oxxo.xlsx", data_only=True)

s1 = wb["工作表1"]
s1["e1"].font = Font(name="Arial", color="ff0000", size=30, bold=True)  # 設定 g1 儲存格的文字樣式
s1["f1"].fill = PatternFill(fill_type="solid", fgColor="DDDDDD")  # 設定 f1 儲存格的背景樣式

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code039.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import csv

csvfile = open("csv-demo.csv")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列
print(data)
"""
[['name', 'id', 'color', 'price'],
 ['apple', '1', 'red', '10'],
 ['orange', '2', 'orange', '15'],
 ['grap', '3', 'purple', '20'],
 ['watermelon', '4', 'green', '30']]
"""


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code040.py

# Copyright © https://steam.oxxostudio.tw

import csv
import openpyxl

csvfile = open("csv-demo.csv")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列

wb = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
sheet = wb.create_sheet("csv")  # 建立空白的工作表
for i in data:
    sheet.append(i)  # 逐筆添加到最後一列

wb.save("test2.xlsx")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code041.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode

img = qrcode.make("https://steam.oxxostudio.tw")  # 要轉換成 QRCode 的文字
img.show()  # 顯示圖片 ( Colab 不適用 )
img.save("qrcode.png")  # 儲存圖片


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code042.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode

qr = qrcode.QRCode(
    version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4
)
qr.add_data("https://steam.oxxostudio.tw")  # 要轉換成 QRCode 的文字
qr.make(fit=True)  # 根據參數製作為 QRCode 物件

img = qr.make_image()  # 產生 QRCode 圖片
img.show()  # 顯示圖片 ( Colab 不適用 )
img.save("qrcode.png")  # 儲存圖片


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code043.py

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode
import qrcode.image.svg

img = qrcode.make(
    "https://steam.oxxostudio.tw", image_factory=qrcode.image.svg.SvgPathImage
)  # 要轉換成 QRCode 的文字
# img.show()                # SVG 無法使用
img.save("qrcode.svg")  # 儲存圖片，注意副檔名為 SVG

"""
下方的程式使用「進階設定」的方式產生 QRcode，額外載入 qrcode.image.svg，在 qrcode.QRCode 裡新增 image_factory=qrcode.image.svg.SvgPathImage 參數，就能產生 SVG 格式的 QRCode 圖片 ( 如果是 SVG 格式圖片無法改變顏色 )。
"""

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode
import qrcode.image.svg

qr = qrcode.QRCode(
    version=1,
    error_correction=qrcode.constants.ERROR_CORRECT_L,
    box_size=10,
    border=4,
    image_factory=qrcode.image.svg.SvgPathImage,
)
qr.add_data("https://steam.oxxostudio.tw")
qr.make(fit=True)

img = qr.make_image()
# img.show()               # SVG 無法使用
img.save("qrcode.svg")  # 儲存圖片，注意副檔名為 SVG


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code044.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode
from qrcode.image.styledpil import StyledPilImage
from qrcode.image.styles.moduledrawers import (
    VerticalBarsDrawer,
    RoundedModuleDrawer,
    HorizontalBarsDrawer,
    SquareModuleDrawer,
    GappedSquareModuleDrawer,
    CircleModuleDrawer,
)

qr = qrcode.QRCode(
    version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4
)
qr.add_data("https://steam.oxxostudio.tw")
qr.make(fit=True)

img1 = qr.make_image(image_factory=StyledPilImage, module_drawer=SquareModuleDrawer())
img2 = qr.make_image(
    image_factory=StyledPilImage, module_drawer=GappedSquareModuleDrawer()
)
img3 = qr.make_image(image_factory=StyledPilImage, module_drawer=CircleModuleDrawer())
img4 = qr.make_image(image_factory=StyledPilImage, module_drawer=RoundedModuleDrawer())
img5 = qr.make_image(image_factory=StyledPilImage, module_drawer=VerticalBarsDrawer())
img6 = qr.make_image(image_factory=StyledPilImage, module_drawer=HorizontalBarsDrawer())
img1.save("qrcode1.png")
img2.save("qrcode2.png")
img3.save("qrcode3.png")
img4.save("qrcode4.png")
img5.save("qrcode5.png")
img6.save("qrcode6.png")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code045.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode
from qrcode.image.styledpil import StyledPilImage
from qrcode.image.styles.moduledrawers import RoundedModuleDrawer
from qrcode.image.styles.colormasks import (
    ImageColorMask,
    SolidFillColorMask,
    RadialGradiantColorMask,
    SquareGradiantColorMask,
    VerticalGradiantColorMask,
    HorizontalGradiantColorMask,
)

qr = qrcode.QRCode(
    version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4
)
qr.add_data("https://steam.oxxostudio.tw")
qr.make(fit=True)

img1 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=SolidFillColorMask((255, 255, 255), (255, 0, 0)),
    module_drawer=RoundedModuleDrawer(),
)
img2 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=RadialGradiantColorMask((255, 255, 255), (255, 0, 0), (0, 0, 255)),
    module_drawer=RoundedModuleDrawer(),
)
img3 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=SquareGradiantColorMask((255, 255, 255), (255, 0, 0), (0, 0, 255)),
    module_drawer=RoundedModuleDrawer(),
)
img4 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=VerticalGradiantColorMask((255, 255, 255), (255, 0, 0), (0, 0, 255)),
    module_drawer=RoundedModuleDrawer(),
)
img5 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=HorizontalGradiantColorMask((255, 255, 255), (255, 0, 0), (0, 0, 255)),
    module_drawer=RoundedModuleDrawer(),
)
img6 = qr.make_image(
    image_factory=StyledPilImage,
    color_mask=ImageColorMask((255, 255, 255), "mona.jpg"),
    module_drawer=RoundedModuleDrawer(),
)

img1.save("qrcode1.png")
img2.save("qrcode2.png")
img3.save("qrcode3.png")
img4.save("qrcode4.png")
img5.save("qrcode5.png")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code046.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

import qrcode
from qrcode.image.styledpil import StyledPilImage

qr = qrcode.QRCode(
    version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4
)
qr.add_data("https://steam.oxxostudio.tw")
qr.make(fit=True)

img = qr.make_image(image_factory=StyledPilImage, embeded_image_path="mona.jpg")
img.save("qrcode.png")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code047.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from barcode import EAN13

number = "12345678987654321"  # 要轉換的數字
my_code = EAN13(number)  # 轉換成 barcode
my_code.save("oxxo")  # 儲存為 SVG


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code048.py

# Copyright © https://steam.oxxostudio.tw

import os

os.chdir("/content/drive/MyDrive/Colab Notebooks")  # Colab 換路徑使用，本機或 Jupyter 環境可刪除

from barcode import EAN13
from barcode.writer import ImageWriter  # 載入 barcode.writer 的 ImageWriter

number = "12345678987654321"
my_code = EAN13(number, writer=ImageWriter())  # 添加 writer=ImageWriter()
my_code.save("oxxo")


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code049.py

# Copyright © https://steam.oxxostudio.tw

import psutil

print(psutil.cpu_count())  # CPU 邏輯數量
print(psutil.cpu_count(logical=False))  # 實際物理 CPU 數量
print(psutil.cpu_percent(interval=0.5, percpu=True))  # CPU 使用率
# interval：每隔多少秒更新一次
# percpu：查看所有 CPU 使用率
print(psutil.cpu_freq())  # CPU 使用頻率


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code050.py

# Copyright © https://steam.oxxostudio.tw

import psutil

print(psutil.virtual_memory())  # 記憶體資訊


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code051.py

# Copyright © https://steam.oxxostudio.tw

import psutil

print(psutil.disk_partitions())  # 所有硬碟資訊
print(psutil.disk_usage("硬碟 device 名稱"))  # 指定硬碟資訊


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code052.py

# Copyright © https://steam.oxxostudio.tw

import psutil

print(psutil.net_io_counters())  # 網路封包
print(psutil.net_if_addrs())  # 網路卡的組態資訊, 包括 IP 地址、Mac地址、子網掩碼、廣播地址等等
print(psutil.net_connections())  # 目前機器的網路連線


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code053.py

# Copyright © https://steam.oxxostudio.tw

import psutil

print(psutil.users())  # 登陸的使用者資訊
print(psutil.boot_time())  # 系統啟動時間
print(datetime.datetime.fromtimestamp(psutil.boot_time()))  # 轉換成標準時間


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code054.py

# Copyright © https://steam.oxxostudio.tw

import psutil

for prcs in psutil.process_iter():
    print(prcs.name)  # 印出所有正在執行的應用程式 ( 從中觀察 pid )

p = psutil.Process(pid=3987)  # 讀取特定應用程式
print(p.name())  # 應用程式名稱
print(p.exe())  # 應用程式所在路徑
print(p.cwd())  # 應用程式執行路徑
print(p.status())  # 應用程式狀態
print(p.username())  # 執行應用程式的使用者
print(p.cpu_times())  # 應用程式的 CPU 使用時間
print(p.memory_info())  # 應用程式的 RAM 使用資訊


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code055.py

# Copyright © https://steam.oxxostudio.tw

import pyautogui

width, hwight = pyautogui.size()
print(width, hwight)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code056.py

# Copyright © https://steam.oxxostudio.tw

import tkinter as tk

root = tk.Tk()  # 產生 tkinter 視窗
width = root.winfo_screenwidth()
height = root.winfo_screenheight()
print(width, height)
root.destroy()  # 關閉視窗


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code057.py

from PyQt5 import QtWidgets
import sys

app = QtWidgets.QApplication(sys.argv)
screen = QtWidgets.QApplication.desktop()
width = screen.width()
height = screen.height()
print(width, height)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code058.py

# Copyright © https://steam.oxxostudio.tw

socket.socket(family, type, proto)
# family：IPv4 本機、IPv4 網路、IPv6 網路。
# type：使用 TCP 或 UDP 方式。
# protocol: 串接協定 ( 通常預設 0 )。


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code059.py

# Copyright © https://steam.oxxostudio.tw

import socket

s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
s.connect(("8.8.8.8", 80))
ip = s.getsockname()[0]
print(ip)
s.close()


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code060.py

# Copyright © https://steam.oxxostudio.tw

import requests

ip = requests.get("https://api.ipify.org").text

print(ip)


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code061.py

# Copyright © https://steam.oxxostudio.tw

import socket

hostname = "google.com"
print(socket.gethostbyname(hostname))


print("------------------------------------------------------------")  # 60個

# 檔案 : C:\_git\vcs\_4.python\__code\_oxxo\python\ch12\code062.py

# Copyright © https://steam.oxxostudio.tw

import os

hostname = "google.com"
response = os.system("ping -c 3 -i 1 " + hostname)
print(response)

response = os.popen(f"ping -c 3 -i 1 {hostname}").read()
print(response)


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個
