import openpyxl
from openpyxl.chart import AreaChart, Reference

wb = openpyxl.load_workbook(r"data\area_chart.xlsx")
sh = wb.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = AreaChart()
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
wb.save(r"tmp_area_chart.xlsx")



print("------------------------------------------------------------")  # 60個



#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\bubble_chart.py

import openpyxl
from openpyxl.chart import Series, Reference, BubbleChart

wb = openpyxl.load_workbook(r"data\bubble_chart.xlsx")
sh = wb.active

chart = BubbleChart()
chart.style = 18
for row in range(2,sh.max_row+1):
    xvalues = Reference(sh, min_col=3, min_row=row)
    yvalues = Reference(sh, min_col=2, min_row=row)
    size = Reference(sh, min_col=4, min_row=row)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size,  title=sh.cell(row,1).value)
    chart.series.append(series)

sh.add_chart(chart, "F2")
wb.save(r"tmp_bubble_chart.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\column_chart.py

import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("data\column_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=3, max_col=3, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #橫條圖
# 轉換成橫條圖之後，有部分的客戶名稱會消失，所以要指定chart整體的高度
#chart.height = 10

chart.style = 28    #1灰色、11藍色、28橙色、30黃色、37圖表的背景設定為淺灰色、45整體的背景設定為黑色
chart.title = "各客戶業績"
chart.y_axis.title = "營業額"
chart.x_axis.title = "客戶名稱"

chart.add_data(data,titles_from_data=True)  #以當月業績作為圖例
chart.set_categories(labels)
sh.add_chart(chart, "E3")

wb.save("tmp_column_chart.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\column_chart_stacked.py

import openpyxl
from openpyxl.chart import BarChart, Reference

wb = openpyxl.load_workbook("data\column_chart_stacked.xlsx")
sh = wb.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.overlap = 100
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
wb.save("tmp_column_chart_stacked.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\easy_bubble_chart.py

import openpyxl
from openpyxl.chart import Series, Reference, BubbleChart

wb = openpyxl.load_workbook(r"data\bubble_chart.xlsx")
sh = wb.active

chart = BubbleChart()
chart.style = 18
xvalues = Reference(sh, min_col=3, min_row=2, max_row=sh.max_row)
yvalues = Reference(sh, min_col=2, min_row=2, max_row=sh.max_row)
size = Reference(sh, min_col=4, min_row=2, max_row=sh.max_row)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size)
chart.series.append(series)

sh.add_chart(chart, "F2")
wb.save(r"tmp_bubble_chart.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\easy_pie_chart.py

import openpyxl
from openpyxl.chart import PieChart, Reference

wb = openpyxl.load_workbook("data\pie_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "D3")
wb.save("tmp_pie_chart.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\line_chart.py

import openpyxl
from openpyxl.chart import LineChart, Reference

wb = openpyxl.load_workbook("data\line_chart.xlsx")
sh = wb.active

data = Reference(sh, min_col=2, max_col=9, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = LineChart()
chart.title = "每月業績"
chart.y_axis.title = "銷售數量"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "A9")
wb.save("tmp_line_chart.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\06\python_prg\pie_chart.py

import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

wb = openpyxl.load_workbook("data\pie_chart.xlsx")
sh = wb.active
#print(sh.max_row)
data = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

#切出第一個扇形
slice = DataPoint(idx=0, explosion=10)
chart.series[0].data_points = [slice]

sh.add_chart(chart, "D3")
wb.save("tmp_pie_chart.xlsx")

print("------------------------------------------------------------")  # 60個

