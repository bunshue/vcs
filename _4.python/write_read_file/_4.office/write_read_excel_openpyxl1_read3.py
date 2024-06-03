"""
讀寫 Excel 檔案, 使用 openpyxl


"""

import os
import sys
import time
import openpyxl

print("------------------------------------------------------------")  # 60個

filename = "data/list-countries-world.xlsx"

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

print('共有 :', sheet.max_row, 'row')
print('共有 :', sheet.max_column, 'column')

list_values = list(sheet.values)
print(len(list_values))
print(list_values)

print('列出 第1列 標題')
print(list_values[0])

print('列出 第2列 開始的資料 10 筆')
row_index = 0
for value_tuple in list_values[1:]:
    col_index = 0
    for value in value_tuple:
        #print(row_index , col_index, str(value))
        print(str(value), end = " ")
        col_index += 1
    print()
    row_index += 1
    if row_index > 10:
        break
print()

print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print('------------------------------------------------------------')	#60個


print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

