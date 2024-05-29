"""
讀寫 Excel 檔案, 使用 openpyxl

python讀取/修改excel

pip3 install openpyxl

專有名詞

workbook（活頁簿）：也就是一個excel檔案
worksheet（工作表）：一個excel可以有很多個工作表，目前正在觀看或處理的工作表又稱作「活動中的工作表（active sheet）」
欄（column）：工作表中的直欄，以A、B、C…代表
行（row）：工作表中的橫列，以1、2、3…代表
儲存格（cell）：工作表中的每一個都是一個儲存格

1. 讀取檔案info
2 讀取檔案所有資料
3. 寫入檔案
    3.1 建立新檔 簡易資料
    3.2 建立新檔 完整資料 + 格式
    3.2 建立新檔 多工作頁
    3.3 建立新檔 寫入資料
    3.4 建立新檔 寫入資料 加 格式
    3.5 讀取檔案 修改、寫入資料
    3.6 插入圖片 表格

"""

import os
import sys
import time
import openpyxl

print("------------------------------------------------------------")  # 60個

print("openpyxl test 01 通用訊息")

filename_r = "data/python_ReadWrite_EXCEL.xlsx"

print("讀取 xlsx, 檔案 : " + filename_r)
workbook = openpyxl.load_workbook(filename_r)

names = workbook.sheetnames  # 讀取 excel檔案 裏所有工作表名稱
print("所有工作表名稱 :", names)

# 取得最後編輯的那個工作表
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print(workbook.active)
print(workbook.active.title)

# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 取得工作表參數
sheet_name = sheet.title
ROW = sheet.max_row
COL = sheet.max_column
print("此工作表名稱、總列數、總行數 :", sheet_name, ROW, COL)
print("工作表有資料最大欄數", COL)
print("工作表有資料最小欄數", sheet.min_column)
print("工作表有資料最大列數", ROW)
print("工作表有資料最小列數", sheet.min_row)

# 取得第 0 個工作表 same
sheet = workbook[workbook.sheetnames[0]]
print("第 0 個工作表名稱 :", sheet.title)

sheet1 = workbook["animals1"]  # 取得工作表名稱為「工作表1」的內容
sheet2 = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

print("印出 title ( 工作表名稱 )、max_row 最大列數、max_column 最大行數")
print(sheet1.title, sheet1.max_row, sheet1.max_column)
print(sheet2.title, sheet2.max_row, sheet2.max_column)

print(sheet2.sheet_properties)  # 印出工作表屬性

from openpyxl.utils import get_column_letter, column_index_from_string

print(column_index_from_string("A"))  # 1
print(column_index_from_string("Z"))  # 26
print(column_index_from_string("AA"))  # 27
print(column_index_from_string("ZZ"))  # 702

print("A = ", column_index_from_string('A'))
print("B = ", column_index_from_string('B'))

print(get_column_letter(5))  # E
print(get_column_letter(100))  # CV

ROW = sheet.max_row
COL = sheet.max_column

for i in range(1, COL+1):
    print(i, ' = ', get_column_letter(i))

workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print("目前工作表： ", workbook.active.title)

workbook.active = workbook["animals2"]
print("目前工作表： ", workbook.active.title)

print("------------------------------------------------------------")  # 60個

print("顯示最大、最小 欄數、列數")

workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

print("儲存格 欄名", sheet["A1"].column)
print("儲存格 列名", sheet["A1"].row)
print("儲存格名", sheet["A1"].coordinate)

print("------------------------------")  # 30個

print("顯示資料 方法一")

print("使用cell方法取得資料")
sheet1 = workbook["animals1"]
sheet2 = workbook["animals2"]

# 取得指定儲存格資料
print(sheet1["A1"], sheet1["A1"].value)

print(sheet1["A1"].value)  # 取出 A1 的內容
print(sheet1.cell(1, 1).value)  # 等同取出 A1 的內容

print(sheet2["B2"].value)  # 取出 B2 的內容
print(sheet2.cell(2, 2).value)  # 等同取出 B2 的內容

# 顯示 cell資料
for i in range(1, ROW + 1):
    for j in range(1, COL + 1):
        print(sheet.cell(row=i, column=j).value, end="   ")
    print()

print("------------------------------")  # 30個

print("顯示資料 方法二")

list_values = list(sheet.values)
print("所有資料")
print(list_values)
print()

print("標題欄")
cols = list_values[0]
print(cols)

cnt = 0
for value_tuple in list_values[1:]:
    print("第", cnt, "筆資料 :", value_tuple)
    # print(type(value_tuple))
    cnt += 1

print("------------------------------")  # 30個

print("顯示資料 方法三 所有工作表")

for sheet in workbook:
    print("工作表 sheet :", sheet)
    for row in sheet:
        # print('row')
        for cell in row:
            print(cell.value, end=" ")
        print()
    print()

print("------------------------------")  # 30個

print("顯示資料 方法四")

workbook = openpyxl.load_workbook(filename_r)
for sheetname in workbook.sheetnames:  # 所有工作表
    sheet = workbook[sheetname]
    ROW = sheet.max_row
    COL = sheet.max_column
    print("此工作表名稱、總列數、總行數 :", sheet_name, ROW, COL)
    for c in range(1, COL + 1):  # 欄1〜最後
        for r in range(1, ROW + 1):  # 列1〜最後
            cell = sheet.cell(row=r, column=c)  # 儲存格
            if cell.value != None:
                print(cell.value, end=" ")
        print()

print("------------------------------")  # 30個

print("顯示資料 方法五")


def get_values(sheet):
    arr = []  # 第一層串列
    for row in sheet:
        arr2 = []  # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr


sheet1 = workbook["animals1"]
sheet2 = workbook["animals2"]

print("印出工作表1 所有內容")
print(get_values(sheet1))

print("印出工作表2 所有內容")
print(get_values(sheet2))

print("------------------------------")  # 30個

print("顯示資料 方法六")

sheet1 = workbook["animals1"]
v = sheet1.iter_rows(min_row=1, min_col=1, max_col=5, max_row=4)  # 取出四格內容
#print(v)
for i in v:
    for j in i:
        print(j.value, end=" ")
    print()

print("------------------------------")  # 30個
print('之前先顯示工作表')
print("目前工作表： ", workbook.active.title)

print("顯示資料 方法七 讀取工作表所有內容")

filename_r = "data/python_ReadWrite_EXCEL.xlsx"
workbook = openpyxl.load_workbook(filename_r)

workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print("excel活動工作表： ", sheet)

for row in sheet:
    for cell in row:
        print(cell.value, end = " ")
    print()
print()

print('之後先顯示工作表')
print("目前工作表： ", workbook.active.title)

print("------------------------------")  # 30個

print("讀取工作表某欄某列的資料")

print("讀取 C 欄的所有資料")
for cell in sheet["C"]:  # 讀取C欄所有資料，並列印出來
    print(cell.value, end = " ")
print()

print("讀取 第 4 列的所有資料")
for cell in sheet["4"]:  # 讀取第2列所有資料，並列印出來
    print(cell.value, end = " ")
print()

print("------------------------------")  # 30個

print("顯示資料 方法八")

filename_r = "data/python_ReadWrite_EXCEL.xlsx"
workbook = openpyxl.load_workbook(filename_r)

workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print("excel活動工作表： ", sheet)

select_data = sheet["A2":"D7"]
for a, b, c, d in select_data:
    print("{0} {1} {2} {3}".format(a.value, b.value, c.value, d.value))
print()

print('sheet.dimensions', sheet.dimensions)

for a, b, c, d in sheet[sheet.dimensions]:
    print(a.value, b.value, c.value, d.value)

print("------------------------------")  # 30個

print("openpyxl test 03 循欄列印 ＆ 循列列印")

for cell in list(sheet.columns)[0]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.columns)[1]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.columns)[2]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.columns)[3]:
    print(cell.value, end = " ")
print()

print("------------------------------")  # 30個

for cell in list(sheet.rows)[0]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.rows)[1]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.rows)[2]:
    print(cell.value, end = " ")
print()

for cell in list(sheet.rows)[3]:
    print(cell.value, end = " ")
print()


print("------------------------------")  # 30個

print("openpyxl test 04 讀取指定區域內容")

for row in sheet["A2":"D5"]:
    for cell in row:
        print(cell.value, end=" ")
    print()

print("------------------------------------------------------------")  # 60個

print("openpyxl test 05 建立多工作表之Excel活頁簿")

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

# 預設為名為Sheet之工作表

#新增工作表
workbook.create_sheet("工作表3")  # 插入工作表 3 在最後方
workbook.create_sheet("工作表1.5", 1)  # 插入工作表 1.5 在第二個位置 ( 工作表 1 和 2 的中間 )
workbook.create_sheet("工作表0", 0)  # 插入工作表 0 在第一個位置
workbook.create_sheet("工作表aa")  # 插入工作表aa 在最後方
workbook.create_sheet("工作表bb")  # 插入工作表bb 在最後方
workbook.create_sheet("Mysheet1", 1)  # 新增工作表並指定放置位置
workbook.create_sheet("Mysheet0", 0)

# 建立新工作表
sheet = workbook.create_sheet("新animal")  # 建立新工作表 名為 新animal
data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]  # 二維陣列資料
for i in data:
    sheet.append(i)  # 逐筆添加到最後一列

# 將工作表的頁箋著色
sheet1 = workbook["工作表aa"]  # 開啟工作表aa
sheet1.sheet_properties.tabColor = "ff0000"  # 修改工作表 1 頁籤顏色為紅色

sheet2 = workbook["工作表bb"]  # 開啟工作表bb
sheet2.sheet_properties.tabColor = "ffff00"  # 修改工作表 2 頁籤顏色為黃色

workbook.copy_worksheet(sheet1)  # 複製工作表aa 放到最後方

# 修改工作表的名稱
sheet1.title = "新工作表A"  # 修改工作表aa 的名稱為 新工作表A
sheet2.title = "新工作表B"  # 修改工作表bb 的名稱為 新工作表B

"""
# 新增工作表，若名稱已經存在則原本名稱之後加數字
workbook.create_sheet(title="amos")

# 修改工作表
workbook["amos"].title = "carol"

# 刪除工作表
workbook.remove(workbook["carol"])
"""

"""
print("openpyxl test 06 新增工作表 ＆ 修改工作表名稱")
workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

"""

filename_w = "tmp_excel_openpyxl_a.xlsx"
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

print("openpyxl test 07a 建立新檔, 簡易資料")
print("開啟空白的活頁簿")
workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

"""
print("openpyxl test 07b 開啟舊檔, 修改資料, 存檔, 或另存新檔")
filename_r = "data/python_ReadWrite_EXCEL.xlsx"
print("讀取 xlsx, 檔案 : " + filename_r)
workbook = openpyxl.load_workbook(filename_r)
"""

print("開啟工作表")
# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料, 直接修改/設定工作表內的資料
sheet["A1"] = "中文名"
sheet["B1"] = "英文名"
sheet["C1"] = "體重"
sheet["D1"] = "全名"
sheet["A2"] = "鼠"
sheet["B2"] = "mouse"
sheet["C2"] = "3"
sheet["D2"] = "米老鼠"
sheet["A3"] = "牛"
sheet["B3"] = "ox"
sheet["C3"] = "48"
sheet["D3"] = "班尼牛"

filename_w = "tmp_excel_openpyxl_b1_new_simple.xlsx"
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個
print("openpyxl test 08 建立新檔 完整資料 + 格式")
print("開啟空白的活頁簿")
workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

print("開啟工作表")
# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料, 直接修改/設定工作表內的資料
sheet["A1"] = "動物列表"
sheet["A3"] = "中文名"
sheet["B3"] = "英文名"
sheet["C3"] = "體重"
sheet["D3"] = "全名"

""" same
listtitle=['中文名', '英文名', '體重', '全名']
sheet.append(listtitle)  
"""

# 以串列寫入資料
animal01 = ["鼠", "mouse", 3]#直接使用數值, 這樣才可以計算
animal02 = ["牛", "ox", 48]
animal03 = ["虎", "tiger", 33]
animal04 = ["兔", "rabbit", 8]
sheet.append(animal01)  # 逐筆添加到最後一列
sheet.append(animal02)  # 逐筆添加到最後一列
sheet.append(animal03)  # 逐筆添加到最後一列
sheet.append(animal04)  # 逐筆添加到最後一列

# 使用.cell()填入資料 設定公式
sheet["C8"].value = "總體重"
sheet["C9"].value = "=SUM(C4:C7)" # 寫入公式

print("------------------------------")  # 30個

print("一次寫入一個二維陣列")
print("建立 二維串列, 裡面都是list")
animals_2d_list = [
    ["鼠", "mouse", 3, "米老鼠"],
    ["牛", "ox", 48, "班尼牛"],
    ["虎", "tiger", 33, "跳跳虎"],
    ["兔", "rabbit", 8, "彼得兔"],
    ["龍", "dragon", 38, "逗逗龍"],
    ["蛇", "snake", 16, "貪吃蛇"],
    ["馬", "horse", 31, "草泥馬"],
    ["羊", "goat", 29, "喜羊羊"],
    ["猴", "monkey", 22, "山道猴"],
    ["雞", "chicken", 5, "肯德雞"],
    ["狗", "dog", 17, "貴賓狗"],
    ["豬", "pig", 42, "佩佩豬"],
]
for y in range(len(animals_2d_list)):
    for x in range(len(animals_2d_list[y])):
        row = 13 + y  # 寫入資料的範圍從 row=13 開始
        col = 1 + x  # 寫入資料的範圍從 column=1 開始
        sheet.cell(row, col).value = animals_2d_list[y][x]

print("------------------------------")  # 30個

#移動資料位置
#sheet.move_range("A13:D24", rows=6, cols=4)  # 區塊向下移6格 向右移4格

print("------------------------------")  # 30個

print('使用數值與使用字串比較')
sheet.cell(10, 1).value = "使用數值"
sheet.cell(10, 2).value = 123  # 儲存格 B1 內容 ( row=12, column=1 ) 為 123
sheet.cell(10, 3).value = 456  # 儲存格 B2 內容 ( row=13, column=1 ) 為 456
sheet.cell(row=10, column=4).value = 789
sheet["E10"] = "總和"
sheet["F10"] = "=SUM(B10:D10)"  # 寫入公式

sheet.cell(11, 1).value = "使用字串"
sheet.cell(11, 2).value = str(123)  # 儲存格 B1 內容 ( row=12, column=2 ) 為 123
sheet.cell(11, 3).value = str(456)  # 儲存格 B2 內容 ( row=13, column=2 ) 為 456
sheet.cell(row=11, column=4).value = str(789)
sheet["E11"] = "總和"
#sheet["F11"] = "=SUM(B11:D11)"  # 寫入公式
sheet["F11"] = "不可對字串計算總和"

print("------------------------------")  # 30個

# 設定欄寬
sheet.column_dimensions["A"].width = 18  # 設定A欄欄寬
sheet.column_dimensions["B"].width = 18  # 設定B欄欄寬
sheet.column_dimensions["C"].width = 18  # 設定C欄欄寬
sheet.column_dimensions["D"].width = 18  # 設定D欄欄寬
sheet.column_dimensions["E"].width = 18  # 設定E欄欄寬

# 設定列高
sheet.row_dimensions[1].height = 30  # 設定第1列列高
sheet.row_dimensions[2].height = 30  # 設定第2列列高
sheet.row_dimensions[3].height = 30  # 設定第3列列高
sheet.row_dimensions[6].height = 30  # 設定第6列列高

print("------------------------------")  # 30個

#將 A1 B1 C1 D1 A2 B2 C2 D2 合併為一格

print("合併儲存格: 方法一")
print("合併儲存格 A1+B1+C1+D1+A2+B2+C2+D2 合成一格")
sheet.merge_cells('A1:D2')
#print("原本合併儲存格再解開")
#sheet.unmerge_cells('A1:D2')#解除合併儲存格

print("------------------------------")  # 30個
"""
print("合併儲存格: 方法二")
print("合併儲存格 A1+B1+C1+D1+A2+B2+C2+D2 合成一格")
sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=4)
print("原本合併儲存格再解開")#解除合併儲存格
sheet.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=4)
"""
print("------------------------------")  # 30個

print("設定儲存格格式 字型 顏色 大小 背景色")
sheet.cell(1, 1).value = "動物列表"
sheet.cell(1, 1).font = openpyxl.styles.Font(size=24, color="0000CC")  # 設定儲存格的文字大小與文字顏色
sheet.cell(1, 1).fill = openpyxl.styles.PatternFill("solid", fgColor="66CCFF")  # 設定儲存格的背景色

sheet["A3"].font = openpyxl.styles.Font(name="Arial", color="ff0000", size=20, bold=True)  # 設定儲存格的文字樣式
sheet["B3"].font = openpyxl.styles.Font(name="Arial", color="00ff00", size=20, bold=True)  # 設定儲存格的文字樣式
sheet["C3"].font = openpyxl.styles.Font(name="Arial", color="0000ff", size=20, bold=True)  # 設定儲存格的文字樣式
sheet["D3"].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="FFFF00")  # 設定儲存格的背景樣式

print("------------------------------")  # 30個

#預設靠左對齊
#改成中間對齊
sheet.cell(1, 1).alignment = openpyxl.styles.Alignment("center")

print("------------------------------")  # 30個

#貼上一張圖
from openpyxl.drawing.image import Image
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
filename = "C:/_git/vcs/_4.python/_data/picture1.jpg"
img = Image(filename)
sheet.add_image(img, "E13")  # 把圖貼在E13

print("------------------------------")  # 30個

sheet["E3"] = "設定日期格式"
import datetime
#sheet = workbook["animals1"]
sheet["E4"] = datetime.datetime(1928, 11, 18, 12, 34, 56)
#sheet['E4'].number_format = 'yyyy-mm-dd h:mm:ss'
sheet['E4'].number_format = 'yyyy-mm-dd'
#sheet["E4"].number_format = "dd-mm-yyyy"

print("------------------------------")  # 30個

filename_w = "tmp_excel_openpyxl_b2_new_all.xlsx"
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

print("openpyxl test 09 讀 寫 從檔案後面附加資料")

filename_w = "tmp_excel_openpyxl_c.xlsx"

if not os.path.exists(filename_w):
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    heading = ["中文名", "英文名", "體重", "全名"]
    sheet.append(heading)
    workbook.save(filename_w)

workbook = openpyxl.load_workbook(filename_w)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

animal01 = ["鼠", "mouse", "3", "米老鼠"]
sheet.append(animal01)

workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

print("openpyxl test 10 創建Excel文件")

# 這個有問題

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

sheet.append(["中文名", "英文名", "體重", "全名"])
data =[
    ["鼠", "mouse", "3", "米老鼠"],
    ["牛", "ox", 48, "班尼牛"],
    ["虎", "tiger", "33", "跳跳虎"]
]

for row in data:
    sheet.append(row)

tab = openpyxl.worksheet.table.Table(displayName="Table1", ref="A1:E5")
tab.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True,
)
#sheet.add_table(tab) 問題在這裡

filename_w = "tmp_excel_openpyxl_d_add_table.xlsx"
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

import pathlib  # 標準函式庫
import csv  # 標準函式庫

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

list_row = 1
path = pathlib.Path("data\sales")  # 指定相對路徑
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        print(pass_obj)
        workbook = openpyxl.load_workbook(pass_obj)
        for sh in workbook:
            for dt_row in range(9, 19):
                if sh.cell(dt_row, 2).value != None:
                    # 更便於說明的代碼
                    # sheet.cell(row=list_row, column=1).value = \
                    #    sh.cell(row=2, column=7).value   #傳票NO
                    sheet.cell(list_row, 1).value = sh.cell(2, 7).value  # 傳票NO
                    sheet.cell(list_row, 2).value = sh.cell(3, 7).value  # 日期
                    sheet.cell(list_row, 3).value = sh.cell(4, 3).value  # 客戶代碼
                    sheet.cell(list_row, 4).value = sh.cell(3, 2).value.strip(
                        "敬啟"
                    )  # 客戶名稱
                    sheet.cell(list_row, 5).value = sh.cell(7, 8).value  # 負責人代碼
                    sheet.cell(list_row, 6).value = sh.cell(7, 7).value  # 負責人姓名
                    sheet.cell(list_row, 7).value = sh.cell(dt_row, 1).value  # No
                    sheet.cell(list_row, 8).value = sh.cell(dt_row, 2).value  # 商品代碼
                    sheet.cell(list_row, 9).value = sh.cell(dt_row, 3).value  # 商品名稱
                    sheet.cell(list_row, 10).value = sh.cell(dt_row, 4).value  # 數量
                    sheet.cell(list_row, 11).value = sh.cell(dt_row, 5).value  # 單價
                    sheet.cell(list_row, 12).value = (
                        sh.cell(dt_row, 4).value * sh.cell(dt_row, 5).value
                    )  # 金額
                    sheet.cell(list_row, 13).value = sh.cell(dt_row, 7).value  # 備註
                    list_row += 1

filename_w = "tmp_excel_openpyxl_f.xlsx"
workbook.save(filename_w)
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

import calendar

year = 2024
month = 5
dayname = ["日", "一", "二", "三", "四", "五", "六"]


# 【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = "tmp_excel_openpyx_" + str(year) + "_" + str(month) + "a.xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    c = sheet.cell(1, 4)
    c.value = str(year) + "年" + str(month) + "月"
    for col in range(7):  # 一週的每一天
        c = sheet.cell(2, col + 1)
        c.value = dayname[col]
    for col, week in enumerate(cal.monthdayscalendar(year, month)):
        for row, day in enumerate(week):
            if day > 0:
                c = sheet.cell((col + 3), row + 1)
                c.value = day
    workbook.save(savefile)  # Excel轉存檔案
    return "轉存" + savefile + "了。"


msg = makecalendar(year, month)
print(msg)

print("------------------------------------------------------------")  # 60個

import calendar

value1 = "2024"
value2 = "5"
dayname = ["日", "一", "二", "三", "四", "五", "六"]

fontN = openpyxl.styles.Font(size=24)
fontB = openpyxl.styles.Font(size=24, color="0000FF")
fontR = openpyxl.styles.Font(size=24, color="FF0000")
fillB = openpyxl.styles.PatternFill(patternType="solid", fgColor="AAAAFF")
fillR = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFAAAA")


# 【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = "tmp_excel_openpyx_" + str(year) + "_" + str(month) + "b.xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    for c in ["A", "B", "C", "D", "E", "F", "G"]:
        sheet.column_dimensions[c].width = 20
    c = sheet.cell(1, 4)
    c.value = str(year) + "年" + str(month) + "月"
    c.font = fontN
    for row in range(7):
        c = sheet.cell(2, row + 1)
        c.value = dayname[row]
        c.font = fontN
        c.alignment = openpyxl.styles.Alignment("center")
        if row == 6:
            c.font = fontB
            c.fill = fillB
        if row == 0:
            c.font = fontR
            c.fill = fillR
    for col, week in enumerate(cal.monthdayscalendar(year, month)):
        sheet.row_dimensions[col + 3].height = 50
        for row, day in enumerate(week):
            if day > 0:
                c = sheet.cell((col + 3), row + 1)
                c.value = day
                c.font = fontN
                if row == 6:
                    c.font = fontB
                if row == 0:
                    c.font = fontR
    workbook.save(savefile)  # Excel轉存檔案
    return "轉存" + savefile + "了。"


msg = makecalendar(value1, value2)
print(msg)

print("------------------------------------------------------------")  # 60個

print("openpyxl test 11 加入圖表1 雷達圖")

filename_r = "data/radar_chart.xlsx"
workbook = openpyxl.load_workbook(filename_r)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

data = openpyxl.chart.Reference(sheet, min_col=2, max_col=4, min_row=1, max_row=sheet.max_row)
labels = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = openpyxl.chart.RadarChart()
# 預設為standard
# filled為填色
# chart.type = "filled"
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "F2")

filename_w = "tmp_excel_openpyxl_e_add_radar_chart.xlsx"
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

print("openpyxl test 11 加入圖表2 統計圖")

# 官方範例

workbook = openpyxl.Workbook(write_only=True)
sheet = workbook.create_sheet()

rows = [
    ("Number", "Batch 1", "Batch 2"),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]


for row in rows:
    sheet.append(row)

chart1 = openpyxl.chart.BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = "Test number"
chart1.x_axis.title = "Sample length (mm)"

data = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_row=7, max_col=3)
cats = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
sheet.add_chart(chart1, "E1")

import copy

chart2 = copy.deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"

sheet.add_chart(chart2, "N1")


chart3 = copy.deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = "Stacked Chart"

sheet.add_chart(chart3, "E15")

chart4 = copy.deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = "Percent Stacked Chart"

sheet.add_chart(chart4, "N15")

filename_w = "tmp_excel_openpyxl_g_bar.xlsx"
workbook.save(filename_w)
print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

""" 暫存文字資料

print("顯示資料 方法七");

filename_r = 'data/python_ReadWrite_EXCEL.xlsx'

workbook = openpyxl.load_workbook(filename_r, data_only=True)  # 要excel開啟才可以看到值，否則會顯示None
workbook = openpyxl.load_workbook(filename_r, data_only=False) # 要excel開啟才可以看到值，否則會顯示None

sheet1 = workbook["animals1"]  # 開啟工作表 1

filename_w = 'tmp_excel_openpyxl_h.xlsx'
workbook.save(filename_w)  # 儲存檔案
print("建立 xlsx OK, 檔案 : " + filename_w)


21.3 Excel files
>>> workbook = openpyxl.load_workbook('temp_data_01.xlsx')
>>> results = []
>>> sheet = workbook.worksheets[0]
>>> for row in sheet.iter_rows():
...     results.append([cell.value for cell in row])
...  
>>> print(results)

[['Notes', 'State', 'State Code', 'Month Day, Year', 'Month Day, Year Code', 'Avg Daily Max Air Temperature (F)', 'Record Count for Daily Max Air Temp (F)', 'Min Temp for Daily Max Air Temp (F)', 'Max Temp for Daily Max Air Temp (F)', 'Avg Daily Max Heat Index (F)', 'Record Count for Daily Max Heat Index (F)', 'Min for Daily Max Heat Index (F)', 'Max for Daily Max Heat Index (F)', 'Daily Max Heat Index (F) % Coverage'], [None, 'Illinois', 17, 'Jan 01, 1979', '1979/01/01', 17.48, 994, 6, 30.5, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'Jan 02, 1979', '1979/01/02', 4.64, 994, -6.4, 15.8, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'Jan 03, 1979', '1979/01/03', 11.05, 994, -0.7, 24.7, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'Jan 04, 1979', '1979/01/04', 9.51, 994, 0.2, 27.6, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'May 15, 1979', '1979/05/15', 68.42, 994, 61, 75.1, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'May 16, 1979', '1979/05/16', 70.29, 994, 63.4, 73.5, 'Missing', 0, 'Missing', 'Missing', '0.00%'], [None, 'Illinois', 17, 'May 17, 1979', '1979/05/17', 75.34, 994, 64, 80.5, 82.6, 2, 82.4, 82.8, '0.20%'], [None, 'Illinois', 17, 'May 18, 1979', '1979/05/18', 79.13, 994, 75.5, 82.1, 81.42, 349, 80.2, 83.4, '35.11%'], [None, 'Illinois', 17, 'May 19, 1979', '1979/05/19', 74.94, 994, 66.9, 83.1, 82.87, 78, 81.6, 85.2, '7.85%']]




21.5.2 Writing Excel files

>>> import csv
>>> data_rows = [fields for fields in csv.reader(open("temp_data_01.csv"))]
>>> 
>>> sheet = workbook.active
>>> for row in data_rows:
...     sheet.append(row)
... 
>>> workbook.save("temp_data_02.xlsx")


workbook = openpyxl.load_workbook('tmp.xlsx')
sheets = workbook.sheetnames
print(sheets)

for i in range(len(sheets)):
    sheet = workbook[sheets[i]]
    print('title', sheet.title)
    for col in sheet.iter_cols(min_row=0, min_col=0, max_row=3, max_col=3):
        for cell in col:
            print(cell.value)

print("------------------------------------------------------------")  # 60個

#workbook = openpyxl.load_workbook(filename_r)

workbook = openpyxl.load_workbook(
    filename_r, data_only=True
)  # 設定 data_only=True 只讀取計算後的數值

print("修改目前工作表")

workbook = openpyxl.load_workbook(filename_r)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print("excel活動工作表： ", workbook.active)

workbook.active = 0
print("excel活動工作表： ", workbook.active)




print("隱藏/取消隱藏工作表")
    sheet = workbook["animals2"]
    sheet.sheet_state = "hidden"
    # sheet = workbook['BBB']
    # sheet.sheet_state = 'visible'


print("刪除工作表")
    sheet = workbook["animals1"]
    workbook.remove(sheet)


print("複製工作表")
    sheet = workbook["animals2"]
    target = workbook.copy_worksheet(sheet)
    target.title = "new_animals2"

    workbook.save(filename_w)
    print("建立 xlsx OK, 檔案 : " + filename_w)

print("------------------------------------------------------------")  # 60個

所有的修改都只是針對記憶體中的excel檔
只有save才會把修改的內容寫到儲存媒體上

print("------------------------------------------------------------")  # 60個

保留

for row in range(2, 7):  # 第 2~6 ROW
    for col in range(65, 70): # 65~69 A~E
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()

print("------------------------------------------------------------")  # 60個

print("建立大綱並折疊")
sheet.column_dimensions.group("A", "D", hidden=True)
sheet.row_dimensions.group(1, 10, hidden=True)

print("------------------------------------------------------------")  # 60個

print("改變字體 ＆ 加上註解")
print("改變字體 ＆ 加上註解")
sheet["A1"].font = openpyxl.styles.Font(name="Courier", size=24, color="FF0000")
sheet["A2"].comment = openpyxl.comments.Comment(text="這是註解", author="Amos")

print(sheet["A2"].comment.text)
print(sheet["A2"].comment.author)

sheet["A2"].comment = None  # 取消註解 不能無此行

print("------------------------------------------------------------")  # 60個

print('新增列在第4列')
sheet.insert_rows(4)

print('新增欄在第3欄, 插入兩欄')
sheet.insert_cols(3, 2)

print('刪除列 刪除第4列 刪除2列')
sheet.delete_rows(4, 2)

print('刪除欄 刪除第3欄 刪除2欄')
sheet.delete_cols(3, 2)

print("------------------------------------------------------------")  # 60個

print("鎖定、解除鎖定")

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
# sheet.freeze_panes = 'B2'
sheet.freeze_panes = None


"""

