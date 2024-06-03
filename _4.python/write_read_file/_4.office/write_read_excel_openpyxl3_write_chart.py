"""
讀寫 Excel 檔案, 使用 openpyxl

插入圖表

"""

import os
import sys
import time
import openpyxl

print("------------------------------------------------------------")  # 60個

# line_chart.py

from openpyxl.chart import LineChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart1_line.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=2, max_col=9, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = LineChart()
chart.title = "每月業績"
chart.y_axis.title = "銷售數量"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "A9")

workbook.save("tmp_01_line_chart.xlsx")

print("------------------------------------------------------------")  # 60個

print("openpyxl test 01 插入圖表")

# easy_bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook("data/python_add_chart2_bubble.xlsx")
sheet = workbook.active

chart = BubbleChart()
chart.style = 18
xvalues = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
yvalues = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
print(xvalues)
print(yvalues)
size = Reference(sheet, min_col=4, min_row=2, max_row=sheet.max_row)
#print(size)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size)
#print(series)
chart.series.append(series)

sheet.add_chart(chart, "F2")

workbook.save("tmp_02_bubble_chart_a.xlsx")

print("------------------------------------------------------------")  # 60個

# bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook("data/python_add_chart2_bubble.xlsx")
sheet = workbook.active

chart = BubbleChart()
chart.style = 18
for row in range(2, sheet.max_row+1):
    print(row)
    xvalues = Reference(sheet, min_col=3, min_row=row)
    yvalues = Reference(sheet, min_col=2, min_row=row)
    size = Reference(sheet, min_col=4, min_row=row)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size,  title=sheet.cell(row,1).value)
    chart.series.append(series)

sheet.add_chart(chart, "F2")

workbook.save("tmp_02_bubble_chart_b.xlsx")

print("------------------------------------------------------------")  # 60個

# easy_pie_chart.py

from openpyxl.chart import PieChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart3_pie.xlsx")
sheet = workbook.active
#print(sheet.max_row)
data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "D3")

workbook.save("tmp_03_pie_charta.xlsx")

print("------------------------------------------------------------")  # 60個

# pie_chart.py

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

workbook = openpyxl.load_workbook("data/python_add_chart3_pie.xlsx")
sheet = workbook.active
#print(sheet.max_row)
data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

#切出第一個扇形
slice = DataPoint(idx=0, explosion=10)
chart.series[0].data_points = [slice]

sheet.add_chart(chart, "D3")

workbook.save("tmp_03_pie_chartb.xlsx")

print("------------------------------------------------------------")  # 60個

# column_chart.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart4_column.xlsx")
sheet = workbook.active
#print(sheet.max_row)
data = Reference(sheet, min_col=3, max_col=3, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #橫條圖
# 轉換成橫條圖之後，有部分的客戶名稱會消失，所以要指定chart整體的高度
#chart.height = 10

chart.style = 28    #1灰色、11藍色、28橙色、30黃色、37圖表的背景設定為淺灰色、45整體的背景設定為黑色
chart.title = "各客戶業績"
chart.y_axis.title = "營業額"
chart.x_axis.title = "客戶名稱"

chart.add_data(data,titles_from_data=True)  #以當月業績作為圖例
chart.set_categories(labels)
sheet.add_chart(chart, "E3")

workbook.save("tmp_04_column_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# column_chart_stacked.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart4_column_stacked.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=3, max_col=7, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.overlap = 100
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "I2")

workbook.save("tmp_04_column_chart_stacked.xlsx")

print("------------------------------------------------------------")  # 60個

print("各類別業績（尺寸堆疊長條圖）")

from openpyxl.chart import AreaChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart5_area.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=3, max_col=7, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = AreaChart()
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "I2")

workbook.save("tmp_05_area_chart.xlsx")

print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個
