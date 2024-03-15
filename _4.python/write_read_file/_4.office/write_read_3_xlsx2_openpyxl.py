#各種檔案寫讀範例 xlsx 1

import sys
import openpyxl

print('------------------------------------------------------------')	#60個

print('excel讀寫測試 1')

print("寫讀 xlsx")

print("建立 xlsx")


workbook=openpyxl.Workbook()    # 建立一個工作簿

# 取得第 1 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'

# 以串列寫入資料
listtitle=["姓名","電話"]
sheet.append(listtitle)  
listdata=["David","0999-1234567"]
sheet.append(listdata)

# 儲存檔案
filename = 'tmp_test1.xlsx'
workbook.save(filename)

print("建立 xlsx OK, 檔案 : " + filename)

print('------------------------------------------------------------')	#60個

import csv
import openpyxl

print('excel讀寫測試 2')

csvfile = open("../_3.csv/data/animals.csv")  # 開啟 CSV 檔案
raw_data = csv.reader(csvfile)  # 讀取 CSV 檔案
data = list(raw_data)  # 轉換成二維串列
print(data)

workbook = openpyxl.Workbook()  # 建立空白的 Excel 活頁簿物件
sheet = workbook.create_sheet("animal")  # 建立新工作表 名為animal

"""
# 以儲存格位置寫入資料
sheet['A1'] = '中文名'
sheet['B1'] = '英文名'
sheet['C1'] = '體重'
sheet['D1'] = '全名'
"""

for i in data:
    print(i)
    sheet.append(i)  # 逐筆添加到最後一列

# 儲存檔案
filename = "tmp_test2_fail.xlsx"
workbook.save(filename)

print("建立 xlsx OK, 檔案 : " + filename)


print('------------------------------------------------------------')	#60個

print('excel讀寫測試 3')

filename = 'data/python_ReadWrite_EXCEL3.xlsx'

print("讀取 xlsx, 檔案 : " + filename)

#  讀取檔案
workbook = openpyxl.load_workbook(filename)
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
print("顯示資料");
# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)
# 取得總行、列數
print(sheet.max_row, sheet.max_column)
# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()
sheet['A3'] = 'David' 

print("另存新檔");
filename2 = 'tmp_test2b.xlsx'
workbook.save(filename2)
print("另存新檔 OK, 檔案 : " + filename2)

print('------------------------------------------------------------')	#60個
print('excel讀寫測試 4')

#各種檔案寫讀範例 xlsx 2

import openpyxl
filename = 'data/python_ReadWrite_EXCEL5_population2.xlsx'

print("讀取 xlsx, 檔案 : " + filename)

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

list_values = list(sheet.values)
print('所有資料')
print(list_values)
print()

print('標題欄')
cols = list_values[0]
print(cols)

cnt = 0
for value_tuple in list_values[1:]:
    print('第', cnt, '筆資料 :', value_tuple)
    #print(type(value_tuple))
    cnt += 1

print('------------------------------------------------------------')	#60個

import openpyxl

filename = 'data/sample.xlsx'

workbook = openpyxl.load_workbook(filename)
for sheet in workbook:
    print('sheet')
    for row in sheet:
        #print('row')
        for cell in row:
            print(cell.value, end = ' ')
        print()

print('------------------------------------------------------------')	#60個

import os
import time
import openpyxl

# User info
firstname = 'David'
lastname = 'Wang'
        
title = 'kkkkkkkk'
age = 22
nationality = 'Taiwan'
            
# Course info
registration_status = 'aaa'
numcourses = 'cccc'
numsemesters = 'ddddd'
            
print("First name: ", firstname, "Last name: ", lastname)
print("Title: ", title, "Age: ", age, "Nationality: ", nationality)
print("# Courses: ", numcourses, "# Semesters: ", numsemesters)
print("Registration status", registration_status)
print("------------------------------------------")

filepath = 'tmp_excel_' + time.strftime("%Y%m%d_%H%M%S", time.localtime()) + '.xlsx'
            
if not os.path.exists(filepath):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    heading = ["First Name", "Last Name", "Title", "Age", "Nationality",
               "# Courses", "# Semesters", "Registration status"]
    sheet.append(heading)
    workbook.save(filepath)
workbook = openpyxl.load_workbook(filepath)
sheet = workbook.active
sheet.append([firstname, lastname, title, age, nationality, numcourses,
              numsemesters, registration_status])
workbook.save(filepath)



print('------------------------------------------------------------')	#60個

import openpyxl
from openpyxl.chart import RadarChart, Reference

print('在excel檔案內加入雷達圖')

workbook = openpyxl.load_workbook(r"data/radar_chart.xlsx")
sh = workbook.active

data = Reference(sh, min_col=2, max_col=4, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = RadarChart()
#預設為standard
#filled為填色
#chart.type = "filled"
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "F2")
workbook.save(r"tmp_add_radar_chart.xlsx")


print('------------------------------------------------------------')	#60個

#創建Excel文件


from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

workbook = Workbook()
sheet = workbook.active
data = [
    ['1001', '白元芳', '男', '13123456789'],
    ['1002', '白潔', '女', '13233445566']
]
sheet.append(['學號', '姓名', '性別', '電話'])
for row in data:
    sheet.append(row)
tab = Table(displayName="Table1", ref="A1:E5")
tab.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
sheet.add_table(tab)
workbook.save('./tmp_全班學生數據.xlsx')


print('------------------------------------------------------------')	#60個

#讀取Excel文件

from openpyxl import load_workbook
from openpyxl import Workbook

workbook = load_workbook('./data/學生明細表.xlsx')
print(workbook.sheetnames)
sheet = workbook[workbook.sheetnames[0]]
print(sheet.title)
for row in range(2, 7):
    for col in range(65, 70):
        cell_index = chr(col) + str(row)
        print(sheet[cell_index].value, end='\t')
    print()

print('------------------------------------------------------------')	#60個

print('------------------------------------------------------------')	#60個

print('------------------------------------------------------------')	#60個



print('------------------------------------------------------------')	#60個
print('作業完成')
print('------------------------------------------------------------')	#60個


