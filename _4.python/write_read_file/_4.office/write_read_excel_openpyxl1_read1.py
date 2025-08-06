"""
使用 openpyxl 讀取 Excel 檔案

pip3 install openpyxl

專有名詞
workbook（活頁簿）：也就是一個excel檔案
worksheet（工作表）：一個excel可以有很多個工作表，目前正在觀看或處理的工作表又稱作「活動中的工作表（active sheet）」
欄（column）：工作表中的直欄，以A、B、C…代表
行（row）：工作表中的橫列，以1、2、3…代表
儲存格（cell）：工作表中的每一個都是一個儲存格

1. 讀取檔案info
2. 讀取檔案所有資料

活頁簿 workbook
工作表 sheet
"""

import os
import sys
import time
import openpyxl

print("------------------------------------------------------------")  # 60個

print("通用訊息")

filename = "data/python_ReadWrite_EXCEL.xlsx"
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

sheetnames = workbook.sheetnames  # 讀取 excel檔案 裏所有工作表名稱
print("所有工作表名稱 :", sheetnames)

length = len(workbook.sheetnames)
print("共有", length, "個工作表")

for i in range(length):
    print("工作表", i + 1, "名稱 :", workbook.sheetnames[i])
    print("取得第", i + 1, "個工作表", end="\t")
    sheet = workbook.worksheets[i]
    # 取得工作表參數
    sheet_name = sheet.title
    ROW_ST = sheet.min_row
    COL_ST = sheet.min_column
    ROW_SP = sheet.max_row
    COL_SP = sheet.max_column
    print("名稱 :", sheet_name, end="\t")
    print("欄位 :", COL_ST, "~", COL_SP, end="\t")
    print("列位 :", ROW_ST, "~", ROW_SP)
    print(sheet.sheet_properties)  # 印出工作表屬性


print("取得最後編輯的那個工作表")
workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

print(workbook.active)
print("目前工作表： ", workbook.active.title)
print("excel活動工作表： ", sheet)

print("取得第 1 個工作表")
sheet1 = workbook.worksheets[0]

print("取得第 1 個工作表 same")
sheet1 = workbook[workbook.sheetnames[0]]

print("取得第 1 個工作表")
sheet1 = workbook["animals1"]  # 取得工作表名稱為「animals1」的內容

print("取得第 2 個工作表")
sheet2 = workbook["animals2"]  # 取得工作表名稱為「animals2」的內容

print("第 1 個工作表名稱 :", sheet1.title)
print("第 2 個工作表名稱 :", sheet2.title)

print("------------------------------")  # 30個

print("顯示資料 方法一 使用cell方法取得資料")

print("取得第 1 個工作表")
sheet1 = workbook["animals1"]

print("取得第 2 個工作表")
sheet2 = workbook["animals2"]

print("讀取 sheet1 的所有內容 ST")
sheet = sheet1

ROW_ST = sheet.min_row
COL_ST = sheet.min_column
ROW_SP = sheet.max_row
COL_SP = sheet.max_column

# 顯示 cell資料
for i in range(ROW_ST, ROW_SP + 1):
    for j in range(COL_ST, COL_SP + 1):
        print(sheet.cell(row=i, column=j).value, end="   ")
    print()

print("讀取 sheet1 的所有內容 SP")

print("------------------------------")  # 30個

print("顯示資料 方法二")

list_values = list(sheet.values)
print("資料長度 :", len(list_values))
print("所有資料 :", list_values)

print("列出 第1列 標題")
print(list_values[0])

cnt = 0
for value_tuple in list_values[1:]:
    print("第", cnt, "筆資料 :", value_tuple)
    # print(type(value_tuple))
    cnt += 1

print("------------------------------")  # 30個

print("顯示資料 方法三 所有工作表")

for sheet in workbook:
    print("工作表 sheet :", sheet, ", 名稱 :", sheet.title)
    for row in sheet:
        # print('row')
        for cell in row:
            print(cell.value, end=" ")
        print()
    print()

print("------------------------------")  # 30個

print("顯示資料 方法四")
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

for sheetname in workbook.sheetnames:  # 所有工作表
    print(sheetname)
    sheet = workbook[sheetname]
    ROW_ST = sheet.min_row
    COL_ST = sheet.min_column
    ROW_SP = sheet.max_row
    COL_SP = sheet.max_column

    print("此工作表名稱、總列數、總行數 :", sheet_name, ROW_SP - ROW_ST + 1, COL_SP - COL_ST + 1)
    for c in range(ROW_ST, COL_SP + 1):
        for r in range(ROW_ST, ROW_SP + 1):
            cell = sheet.cell(row=r, column=c)  # 儲存格
            if cell.value != None:
                print(cell.value, end=" ")
        print()

print("------------------------------")  # 30個

print("顯示資料 方法五")
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

print("取得第 1 個工作表")
sheet1 = workbook.worksheets[0]

results = []  # 第一層串列
for row in sheet1:
    results2 = []  # 第二層串列
    for column in row:
        results2.append(column.value)  # 寫入內容
    results.append(results2)

for _ in results:
    print(_)

print("------------------------------")  # 30個

print("顯示資料 方法六")

print("取得第 1 個工作表")
sheet1 = workbook.worksheets[0]

v = sheet1.iter_rows(min_row=1, min_col=1, max_col=5, max_row=4)  # 取出四格內容
# print(v)
for i in v:
    for j in i:
        print(j.value, end=" ")
    print()

print("------------------------------")  # 30個

print("顯示資料 方法七")

filename = "data/python_ReadWrite_EXCEL.xlsx"
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

print("取得第 1 個工作表")
sheet1 = workbook.worksheets[0]

for row in sheet1:
    for cell in row:
        print(cell.value, end=" ")
    print()
print()

print("------------------------------")  # 30個

print("讀取工作表某欄某列的資料")

print("讀取 C 欄的所有資料")
for cell in sheet1["C"]:  # 讀取C欄所有資料，並列印出來
    print(cell.value, end=" ")
print()

print("讀取 第 4 列的所有資料")
for cell in sheet1["4"]:  # 讀取第2列所有資料，並列印出來
    print(cell.value, end=" ")
print()

select_data = sheet1["A2":"D7"]
for a, b, c, d in select_data:
    print("{0} {1} {2} {3}".format(a.value, b.value, c.value, d.value))
print()

# 取得指定儲存格資料
print(sheet1["A1"])
print(sheet1["A1"].value)  # 取出 A1 的內容
print(sheet1.cell(1, 1).value)  # 等同取出 A1 的內容

print(sheet2["B2"].value)  # 取出 B2 的內容
print(sheet2.cell(2, 2).value)  # 等同取出 B2 的內容

print("讀取指定儲存格的內容")
print(sheet1["A1"].value)
print(sheet1["A2"].value)
print(sheet1["B2"].value)
print(sheet1["C3"].value)

print("讀取指定區域內容")

for row in sheet1["A2":"D5"]:
    for cell in row:
        print(cell.value, end=" ")
    print()

print("------------------------------")  # 30個

print("資料所在位置 sheet1.dimensions :", sheet1.dimensions)

for a, b, c, d in sheet1[sheet1.dimensions]:
    print(a.value, b.value, c.value, d.value)

print("------------------------------")  # 30個

print("顯示資料 方法九")
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

sheetnames = workbook.sheetnames  # 讀取 excel檔案 裏所有工作表名稱
print("所有工作表名稱 :", sheetnames)

length = len(workbook.sheetnames)
print("共有", length, "個工作表")

for i in range(length):
    sheet = workbook[sheetnames[i]]
    print("title", sheet.title)
    for col in sheet.iter_cols(min_row=0, min_col=0, max_row=3, max_col=3):
        for cell in col:
            print(cell.value, end=" ")
    print()

print("------------------------------")  # 30個

ROW_ST = sheet.min_row
COL_ST = sheet.min_column
ROW_SP = sheet.max_row
COL_SP = sheet.max_column

for i in range(ROW_ST, ROW_SP + 1):
    sheet.row_dimensions[i].height = 18
    for j in range(COL_ST, COL_SP + 1):
        print(sheet.cell(row=i, column=j).value, end=" ")
    print()

print("------------------------------")  # 30個

print("循欄列印")

print("第1欄")
for cell in list(sheet.columns)[0]:
    print(cell.value, end=" ")
print()

print("第2欄")
for cell in list(sheet.columns)[1]:
    print(cell.value, end=" ")
print()

print("第3欄")
for cell in list(sheet.columns)[2]:
    print(cell.value, end=" ")
print()

print("第4欄")
for cell in list(sheet.columns)[3]:
    print(cell.value, end=" ")
print()

print("------------------------------")  # 30個

print("循列列印")

print("第1列")
for cell in list(sheet.rows)[0]:
    print(cell.value, end=" ")
print()

print("第2列")
for cell in list(sheet.rows)[1]:
    print(cell.value, end=" ")
print()

print("第3列")
for cell in list(sheet.rows)[2]:
    print(cell.value, end=" ")
print()

print("第4列")
for cell in list(sheet.rows)[3]:
    print(cell.value, end=" ")
print()

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

filename = "data/python_ReadWrite_EXCEL.xlsx"
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

print("取得第 1 個工作表")
sheet = workbook.worksheets[0]

print("印出整個工作表1")
i = 0
for row in sheet:
    print("第", i, "row", end=" : ")
    i += 1
    for cell in row:
        print(cell.value, end=" ")
    print()
print()

print("印出整個工作表2")
ROW_ST = sheet.min_row
COL_ST = sheet.min_column
ROW_SP = sheet.max_row
COL_SP = sheet.max_column
i = 0
for row in range(ROW_ST, ROW_SP + 1):
    print("第", i, "row", end=" : ")
    i += 1
    for col in range(COL_ST, COL_SP + 1):
        print(sheet.cell(row, col).value, end=" ")
    print()
print()

print("印出工作表部分資料1")
i = 0
for rows in sheet["B2":"C3"]:
    for cell in rows:
        print(cell.value, end=" ")
    print()
print()

print("印出工作表部分資料2")

getted_list = []
i = 0
for rows in sheet.iter_rows(min_row=2, min_col=2, max_row=3, max_col=3):
    print("第", i, "row", end=" : ")
    for cell in rows:
        print(cell.value, end=" ")
    print()
print()

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

filename = "data/list-countries-world.xlsx"
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

print("取得第 1 個工作表")
sheet = workbook.worksheets[0]

ROW_ST = sheet.min_row
COL_ST = sheet.min_column
ROW_SP = sheet.max_row
COL_SP = sheet.max_column

list_values = list(sheet.values)
print("資料長度 :", len(list_values))
print("所有資料 :", list_values)

print("列出 第1列 標題")
print(list_values[0])

print("列出 第2列 開始的資料 10 筆")
row_index = 0
for value_tuple in list_values[1:]:
    col_index = 0
    for value in value_tuple:
        # print(row_index , col_index, str(value))
        print(str(value), end=" ")
        col_index += 1
    print()
    row_index += 1
    if row_index > 10:
        break
print()

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

filename = "data/108Q2.xlsx"
print("讀取 xlsx, 檔案 : " + filename)
workbook = openpyxl.load_workbook(filename)

# 取得第 1 個工作表
sheet = workbook.worksheets[0]

# 取得總行、列數
print(sheet.max_row, sheet.max_column)

""" many
# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()   
"""


print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個
sys.exit()

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個

filename = "data/python_ReadWrite_EXCEL.xlsx"


"""
#開啟舊檔 + 參數
workbook = openpyxl.load_workbook(filename, data_only=True)  # 要excel開啟才可以看到值，否則會顯示None
workbook = openpyxl.load_workbook(filename, data_only=False) # 要excel開啟才可以看到值，否則會顯示None
"""

print("------------------------------")  # 30個

""" CSV 轉 EXCEL
import csv
data_rows = [fields for fields in csv.reader(open("animals.csv"))]

print("取得第 1 個工作表")
sheet = workbook.worksheets[0]

for row in data_rows:
    sheet.append(row)

workbook.save("temp_data_02.xlsx")
"""

print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

# 所有的修改都只是針對記憶體中的excel檔
# 只有save才會把修改的內容寫到儲存媒體上

print("------------------------------------------------------------")  # 60個

print("建立大綱並折疊")
sheet.column_dimensions.group("A", "D", hidden=True)
sheet.row_dimensions.group(1, 10, hidden=True)

print("------------------------------------------------------------")  # 60個

print("改變字體 ＆ 加上註解")
sheet["A1"].font = openpyxl.styles.Font(name="Courier", size=24, color="FF0000")
sheet["A2"].comment = openpyxl.comments.Comment(text="這是註解", author="Amos")

print(sheet["A2"].comment.text)
print(sheet["A2"].comment.author)

sheet["A2"].comment = None  # 取消註解 不能無此行

print("------------------------------------------------------------")  # 60個

print("新增列在第4列")
sheet.insert_rows(4)

print("新增欄在第3欄, 插入兩欄")
sheet.insert_cols(3, 2)

print("刪除列 刪除第4列 刪除2列")
sheet.delete_rows(4, 2)

print("刪除欄 刪除第3欄 刪除2欄")
sheet.delete_cols(3, 2)

print("------------------------------------------------------------")  # 60個

print("鎖定、解除鎖定")

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
# sheet.freeze_panes = 'B2'
sheet.freeze_panes = None

print("------------------------------------------------------------")  # 60個

設定粗體
sheet.cell(row=i, column=j).font = Font(bold=True)

# 千分位樣式
sheet.cell(row=i, column=j).number_format = "#,##0"

# 固定為第1列、第2欄
sheet.freeze_panes = "C2"

# 隱藏A欄
# sheet.column_dimensions["A"].hidden=False

print("------------------------------------------------------------")  # 60個

# color_scale.py

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.formatting.rule import ColorScaleRule

workbook = openpyxl.Workbook()

print("取得第 1 個工作表")
sheet = workbook.worksheets[0]

values = [11, 22, 33, 44, 55, 66, 77, 88, 99, 55]

# A欄 Sugar可看出漸層效果  Kilo看不出效果
for i, value in enumerate(values):
    sheet.cell(i + 1, 1).value = value

two_color_scale = ColorScaleRule(
    start_type="min", start_color="FF0000", end_type="max", end_color="FFFFFF"
)

sheet.conditional_formatting.add("A1:A10", two_color_scale)

print("------------------------------")  # 60個

# C欄

for i, value in enumerate(values):
    sheet.cell(i + 1, 3).value = value

less_than_rule = CellIsRule(
    operator="lessThan",
    formula=[100],
    stopIfTrue=True,
    fill=PatternFill("solid", start_color="FF0000", end_color="FF0000"),
)
sheet.conditional_formatting.add("C1:C10", less_than_rule)

workbook.save("tmp_01_fill_red.xlsx")

from openpyxl.utils import get_column_letter

print("取得欄位名稱 前30個")
for i in range(30 + 1):
    print(get_column_letter(i + 1), end=" ")
print()

from openpyxl.utils import column_index_from_string

print("由欄位名稱 取得索引")
print(column_index_from_string("A"))  # 1
print(column_index_from_string("Z"))  # 26
print(column_index_from_string("AA"))  # 27
print(column_index_from_string("ZZ"))  # 702
print(column_index_from_string("A"))
print(column_index_from_string("B"))


print("儲存格 欄名", sheet["C5"].column)
print("儲存格 列名", sheet["C5"].row)
print("儲存格名", sheet["C5"].coordinate)
