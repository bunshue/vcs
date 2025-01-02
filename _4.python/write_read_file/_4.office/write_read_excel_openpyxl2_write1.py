"""
使用 openpyxl 寫入 Excel 檔案

pip3 install openpyxl

python讀取/修改excel

寫入檔案
1 建立新檔 簡易資料
2 建立新檔 完整資料 + 格式
2 建立新檔 多工作頁
3 建立新檔 寫入資料
4 建立新檔 寫入資料 加 格式
5 讀取檔案 修改、寫入資料
6 插入圖片 表格

"""

import os
import sys
import time
import openpyxl

pic_filename = "C:/_git/vcs/_4.python/_data/picture1.jpg"

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

# 選取正在工作中的工作表，空白活頁簿有一個預存工作表，名稱是Sheet
ws = workbook.active

# 設定工作表名稱為Sheet1
ws.title = "Sheet1"

# 指定字串給A1儲存格
ws["A1"] = "1"

# 指定數值給A2儲存格
ws["A2"] = 1

# 指定數值給B2儲存格
ws["B2"] = 2

# 指定數值公式給C2儲存格
ws["C2"] = "=SUM(A2:B2)"

# 指定數值計算結果給A3儲存格
ws["A3"] = 1 + 2

# 指定字串計算結果給B3儲存格
ws["B3"] = "1" + "2"

# 指定數值計算公式給A4儲存格
ws["A4"] = "=1+2"

# 指定字串公式給C2儲存格
ws["B4"] = '=CONCATENATE("1","2")'

# 新增一個工作表名稱為Sheet2
workbook.create_sheet("Sheet2")

# 選擇Sheet2工作表為正在工作中的工作表
ws = workbook.get_sheet_by_name("Sheet2")

# 一次填入一列數值
ws.append([1, 2, 3])

# 一次填入一列文字
ws.append(["1", "2", "3"])

# 建立圖片物件
img = openpyxl.drawing.image.Image(pic_filename)
img.height = img.height * 0.1
img.width = img.width * 0.1

# 加入圖片於C1儲存格
ws.add_image(img, "C1")

filename_w = "tmp_excel_openpyxl_a.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

print("openpyxl test 01 建立多工作表之Excel活頁簿")

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

# 預設為名為Sheet之工作表

# 新增工作表
workbook.create_sheet("工作表3")  # 插入工作表 3 在最後方
workbook.create_sheet("工作表1.5", 1)  # 插入工作表 1.5 在第二個位置 ( 工作表 1 和 2 的中間 )
workbook.create_sheet("工作表0", 0)  # 插入工作表 0 在第一個位置
workbook.create_sheet("工作表aa")  # 插入工作表aa 在最後方
workbook.create_sheet("工作表bb")  # 插入工作表bb 在最後方
workbook.create_sheet("Mysheet1", 1)  # 新增工作表並指定放置位置
workbook.create_sheet("Mysheet0", 0)

# 建立新工作表
sheet = workbook.create_sheet("新animal")  # 建立新工作表 名為 新animal
data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]  # 二維陣列資料
for i in data:
    sheet.append(i)  # 逐筆添加到最後一列

# 將工作表的頁箋著色
sheet1 = workbook["工作表aa"]  # 開啟工作表aa
sheet1.sheet_properties.tabColor = "ff0000"  # 修改工作表 1 頁籤顏色為紅色

sheet2 = workbook["工作表bb"]  # 開啟工作表bb
sheet2.sheet_properties.tabColor = "ffff00"  # 修改工作表 2 頁籤顏色為黃色

workbook.copy_worksheet(sheet1)  # 複製工作表aa 放到最後方

# 修改工作表的名稱
sheet1.title = "新工作表A"  # 修改工作表aa 的名稱為 新工作表A
sheet2.title = "新工作表B"  # 修改工作表bb 的名稱為 新工作表B

"""
# 新增工作表，若名稱已經存在則原本名稱之後加數字
workbook.create_sheet(title="amos")

# 修改工作表
workbook["amos"].title = "carol"

# 刪除工作表
workbook.remove(workbook["carol"])
"""

"""
print("openpyxl test 06 新增工作表 ＆ 修改工作表名稱")
workbook.active = 0
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
"""

filename_w = "tmp_excel_openpyxl_a_sheet.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

print("openpyxl test 02a 建立新檔, 簡易資料")

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

"""
print("openpyxl test 02b 開啟舊檔, 修改資料, 存檔, 或另存新檔")
filename_r = "data/python_ReadWrite_EXCEL.xlsx"
print("讀取 xlsx, 檔案 : " + filename_r)
workbook = openpyxl.load_workbook(filename_r)
"""

print("開啟工作表")
# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料, 直接修改/設定工作表內的資料
sheet["A1"] = "中文名"
sheet["B1"] = "英文名"
sheet["C1"] = "體重"
sheet["D1"] = "全名"
sheet["A2"] = "鼠"
sheet["B2"] = "mouse"
sheet["C2"] = "3"
sheet["D2"] = "米老鼠"
sheet["A3"] = "牛"
sheet["B3"] = "ox"
sheet["C3"] = "48"
sheet["D3"] = "班尼牛"

filename_w = "tmp_excel_openpyxl_b1_new_simple.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 02c 建立新檔 完整資料 + 格式")
print("開啟空白的活頁簿")
workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

print("開啟工作表")
# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料, 直接修改/設定工作表內的資料
sheet["A1"] = "動物列表"
sheet["A3"] = "中文名"
sheet["B3"] = "英文名"
sheet["C3"] = "體重"
sheet["D3"] = "全名"

""" same
listtitle=['中文名', '英文名', '體重', '全名']
sheet.append(listtitle)  
"""

# 以串列寫入資料
animal01 = ["鼠", "mouse", 3]  # 直接使用數值, 這樣才可以計算
animal02 = ["牛", "ox", 48]
animal03 = ["虎", "tiger", 33]
animal04 = ["兔", "rabbit", 8]
sheet.append(animal01)  # 逐筆添加到最後一列
sheet.append(animal02)  # 逐筆添加到最後一列
sheet.append(animal03)  # 逐筆添加到最後一列
sheet.append(animal04)  # 逐筆添加到最後一列

# 使用.cell()填入資料 設定公式
sheet["C8"].value = "總體重"
sheet["C9"].value = "=SUM(C4:C7)"  # 寫入公式

print("------------------------------")  # 30個

print("一次寫入一個二維陣列")
print("建立 二維串列, 裡面都是list")
animals_2d_list = [
    ["鼠", "mouse", 3, "米老鼠"],
    ["牛", "ox", 48, "班尼牛"],
    ["虎", "tiger", 33, "跳跳虎"],
    ["兔", "rabbit", 8, "彼得兔"],
    ["龍", "dragon", 38, "逗逗龍"],
    ["蛇", "snake", 16, "貪吃蛇"],
    ["馬", "horse", 31, "草泥馬"],
    ["羊", "goat", 29, "喜羊羊"],
    ["猴", "monkey", 22, "山道猴"],
    ["雞", "chicken", 5, "肯德雞"],
    ["狗", "dog", 17, "貴賓狗"],
    ["豬", "pig", 42, "佩佩豬"],
]
for y in range(len(animals_2d_list)):
    for x in range(len(animals_2d_list[y])):
        row = 13 + y  # 寫入資料的範圍從 row=13 開始
        col = 1 + x  # 寫入資料的範圍從 column=1 開始
        sheet.cell(row, col).value = animals_2d_list[y][x]

print("------------------------------")  # 30個

# 移動資料位置
# sheet.move_range("A13:D24", rows=6, cols=4)  # 區塊向下移6格 向右移4格

print("------------------------------")  # 30個

print("使用數值與使用字串比較")
sheet.cell(10, 1).value = "使用數值"
sheet.cell(10, 2).value = 123  # 儲存格 B1 內容 ( row=12, column=1 ) 為 123
sheet.cell(10, 3).value = 456  # 儲存格 B2 內容 ( row=13, column=1 ) 為 456
sheet.cell(row=10, column=4).value = 789
sheet["E10"] = "總和"
sheet["F10"] = "=SUM(B10:D10)"  # 寫入公式

sheet.cell(11, 1).value = "使用字串"
sheet.cell(11, 2).value = str(123)  # 儲存格 B1 內容 ( row=12, column=2 ) 為 123
sheet.cell(11, 3).value = str(456)  # 儲存格 B2 內容 ( row=13, column=2 ) 為 456
sheet.cell(row=11, column=4).value = str(789)
sheet["E11"] = "總和"
# sheet["F11"] = "=SUM(B11:D11)"  # 寫入公式
sheet["F11"] = "不可對字串計算總和"

print("------------------------------")  # 30個

# 設定欄寬
sheet.column_dimensions["A"].width = 18  # 設定A欄欄寬
sheet.column_dimensions["B"].width = 18  # 設定B欄欄寬
sheet.column_dimensions["C"].width = 18  # 設定C欄欄寬
sheet.column_dimensions["D"].width = 18  # 設定D欄欄寬
sheet.column_dimensions["E"].width = 18  # 設定E欄欄寬

# 設定列高
sheet.row_dimensions[1].height = 30  # 設定第1列列高
sheet.row_dimensions[2].height = 30  # 設定第2列列高
sheet.row_dimensions[3].height = 30  # 設定第3列列高
sheet.row_dimensions[6].height = 30  # 設定第6列列高

print("------------------------------")  # 30個

# 將 A1 B1 C1 D1 A2 B2 C2 D2 合併為一格

print("合併儲存格: 方法一")
print("合併儲存格 A1+B1+C1+D1+A2+B2+C2+D2 合成一格")
sheet.merge_cells("A1:D2")
# print("原本合併儲存格再解開")
# sheet.unmerge_cells('A1:D2')#解除合併儲存格

print("------------------------------")  # 30個
"""
print("合併儲存格: 方法二")
print("合併儲存格 A1+B1+C1+D1+A2+B2+C2+D2 合成一格")
sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=4)
print("原本合併儲存格再解開")#解除合併儲存格
sheet.unmerge_cells(start_row=1, start_column=1, end_row=2, end_column=4)
"""
print("------------------------------")  # 30個

print("合併儲存格: 方法三")

print("合併儲存格的測試")
sheet["A28"] = "合併儲存格的測試"
sheet.merge_cells("A28:D31")
sheet["A28"].alignment = openpyxl.styles.Alignment(horizontal="center")
# sheet.unmerge_cells("A28:D31")

print("------------------------------")  # 30個

print("設定儲存格格式 字型 顏色 大小 背景色")
sheet.cell(1, 1).value = "動物列表"
sheet.cell(1, 1).font = openpyxl.styles.Font(size=24, color="0000CC")  # 設定儲存格的文字大小與文字顏色
sheet.cell(1, 1).fill = openpyxl.styles.PatternFill(
    "solid", fgColor="66CCFF"
)  # 設定儲存格的背景色

sheet["A3"].font = openpyxl.styles.Font(
    name="Arial", color="ff0000", size=20, bold=True
)  # 設定儲存格的文字樣式
sheet["B3"].font = openpyxl.styles.Font(
    name="Arial", color="00ff00", size=20, bold=True
)  # 設定儲存格的文字樣式
sheet["C3"].font = openpyxl.styles.Font(
    name="Arial", color="0000ff", size=20, bold=True
)  # 設定儲存格的文字樣式
sheet["D3"].fill = openpyxl.styles.PatternFill(
    fill_type="solid", fgColor="FFFF00"
)  # 設定儲存格的背景樣式

print("------------------------------")  # 30個

# 預設靠左對齊
# 改成中間對齊
sheet.cell(1, 1).alignment = openpyxl.styles.Alignment("center")

print("------------------------------")  # 30個

# 貼上一張圖
from openpyxl.drawing.image import Image

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
img = Image(pic_filename)
sheet.add_image(img, "E13")  # 把圖貼在E13

print("------------------------------")  # 30個

sheet["E3"] = "設定日期格式"
import datetime

# sheet = workbook["animals1"]
sheet["E4"] = datetime.datetime(1928, 11, 18, 12, 34, 56)
# sheet['E4'].number_format = 'yyyy-mm-dd h:mm:ss'
sheet["E4"].number_format = "yyyy-mm-dd"
# sheet["E4"].number_format = "dd-mm-yyyy"

print("------------------------------")  # 30個

print("儲存格格式")
from openpyxl.styles import Border, Side

side1 = Side(style="hair", color="FF0000")  # R
side2 = Side(style="dashDotDot", color="00FF00")  # G
side3 = Side(style="double", color="0000FF")  # B

for rows in sheet["A13":"D15"]:
    for cell in rows:
        cell.border = Border(left=side1, right=side1, top=side1, bottom=side1)

for rows in sheet["A17":"D19"]:
    for cell in rows:
        cell.border = Border(left=side2, right=side2, top=side2, bottom=side2)

for rows in sheet["A21":"D23"]:
    for cell in rows:
        cell.border = Border(left=side3, right=side3, top=side3, bottom=side3)

print("------------------------------")  # 30個

filename_w = "tmp_excel_openpyxl_b2_new_all1.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 02 建立新檔 完整資料 + 格式")
print("開啟空白的活頁簿")
workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件

print("開啟工作表")
# 取得第 0 個工作表
sheet = workbook.worksheets[0]

# 以儲存格位置寫入資料, 直接修改/設定工作表內的資料
sheet["A1"] = "動物列表"
sheet["A2"] = "中文名"
sheet["B2"] = "英文名"
sheet["C2"] = "體重"
sheet["D2"] = "全名"

""" same
listtitle=['中文名', '英文名', '體重', '全名']
sheet.append(listtitle)  
"""
# 以串列寫入資料
animal01 = ["鼠", "mouse", 3, "米老鼠"]  # 直接使用數值, 這樣才可以計算
animal02 = ["牛", "ox", 48, "班尼牛"]
animal03 = ["虎", "tiger", 33, "跳跳虎"]
animal04 = ["兔", "rabbit", 8, "彼得兔"]
sheet.append(animal01)  # 逐筆添加到最後一列
sheet.append(animal02)  # 逐筆添加到最後一列
sheet.append(animal03)  # 逐筆添加到最後一列
sheet.append(animal04)  # 逐筆添加到最後一列

print("------------------------------")  # 30個

# 設定欄寬
col_widths = {"A": 15, "B": 15, "C": 10, "D": 10, "E": 10, "F": 10, "G": 10, "H": 10}
for col_name in col_widths:
    sheet.column_dimensions[col_name].width = col_widths[col_name]

# 建立字體
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

font_header = Font(name="MS PGothic", size=12, bold=True, color="FFFFFF")

# 設定背景色
TITLE_CELL_COLOR = "AA8866"
for rows in sheet["A1":"H1"]:
    for cell in rows:
        # 儲存格背景色
        cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
        # 水平置中
        cell.alignment = Alignment(horizontal="distributed")
        # 設定字型
        cell.font = font_header

side = Side(style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
for row in sheet:
    for cell in row:
        cell.border = border
        # sheet[cell.coordinate].border = border

print("------------------------------")  # 30個

filename_w = "tmp_excel_openpyxl_b2_new_all2.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 03 讀 寫 從檔案後面附加資料")

if not os.path.exists(filename_w):
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    heading = ["中文名", "英文名", "體重", "全名"]
    sheet.append(heading)
    workbook.save(filename_w)  # 儲存檔案

workbook = openpyxl.load_workbook(filename_w)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

animal01 = ["鼠", "mouse", "3", "米老鼠"]
sheet.append(animal01)

filename_w = "tmp_excel_openpyxl_c.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 03 讀 寫 讀後再寫")

filename_r = "data/python_ReadWrite_EXCEL.xlsx"

workbook = openpyxl.load_workbook(
    filename_r, data_only=True
)  # 設定 data_only=True 只讀取計算後的數值

print("修改目前工作表")

workbook = openpyxl.load_workbook(filename_r)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
print("excel活動工作表： ", workbook.active)

workbook.active = 0
print("excel活動工作表： ", workbook.active)

print("隱藏/取消隱藏工作表")
sheet = workbook["animals2"]
sheet.sheet_state = "hidden"
# sheet = workbook['BBB']
# sheet.sheet_state = 'visible'

print("刪除工作表")
sheet = workbook["animals1"]
workbook.remove(sheet)

print("複製工作表")
sheet = workbook["animals2"]
target = workbook.copy_worksheet(sheet)
target.title = "new_animals2"

filename_w = "tmp_excel_openpyxl_c_sheet.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 04 創建Excel文件")

# 這個有問題

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

sheet.append(["中文名", "英文名", "體重", "全名"])
data = [["鼠", "mouse", "3", "米老鼠"], ["牛", "ox", 48, "班尼牛"], ["虎", "tiger", "33", "跳跳虎"]]

for row in data:
    sheet.append(row)

tab = openpyxl.worksheet.table.Table(displayName="Table1", ref="A1:E5")
tab.tableStyleInfo = openpyxl.worksheet.table.TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=True,
)
# sheet.add_table(tab) 問題在這裡

filename_w = "tmp_excel_openpyxl_d_add_table.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 05")

import pathlib  # 標準函式庫
import csv  # 標準函式庫

workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

list_row = 1
path = pathlib.Path("data\sales")  # 指定相對路徑
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        print(pass_obj)
        workbook = openpyxl.load_workbook(pass_obj)
        for sh in workbook:
            for dt_row in range(9, 19):
                if sh.cell(dt_row, 2).value != None:
                    # 更便於說明的代碼
                    # sheet.cell(row=list_row, column=1).value = \
                    #    sh.cell(row=2, column=7).value   #傳票NO
                    sheet.cell(list_row, 1).value = sh.cell(2, 7).value  # 傳票NO
                    sheet.cell(list_row, 2).value = sh.cell(3, 7).value  # 日期
                    sheet.cell(list_row, 3).value = sh.cell(4, 3).value  # 客戶代碼
                    sheet.cell(list_row, 4).value = sh.cell(3, 2).value.strip(
                        "敬啟"
                    )  # 客戶名稱
                    sheet.cell(list_row, 5).value = sh.cell(7, 8).value  # 負責人代碼
                    sheet.cell(list_row, 6).value = sh.cell(7, 7).value  # 負責人姓名
                    sheet.cell(list_row, 7).value = sh.cell(dt_row, 1).value  # No
                    sheet.cell(list_row, 8).value = sh.cell(dt_row, 2).value  # 商品代碼
                    sheet.cell(list_row, 9).value = sh.cell(dt_row, 3).value  # 商品名稱
                    sheet.cell(list_row, 10).value = sh.cell(dt_row, 4).value  # 數量
                    sheet.cell(list_row, 11).value = sh.cell(dt_row, 5).value  # 單價
                    sheet.cell(list_row, 12).value = (
                        sh.cell(dt_row, 4).value * sh.cell(dt_row, 5).value
                    )  # 金額
                    sheet.cell(list_row, 13).value = sh.cell(dt_row, 7).value  # 備註
                    list_row += 1

filename_w = "tmp_excel_openpyxl_e.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
"""
import calendar

year = 2024
month = 5
dayname = ["日", "一", "二", "三", "四", "五", "六"]


# 【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = "tmp_excel_openpyx_" + str(year) + "_" + str(month) + "a.xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    c = sheet.cell(1, 4)
    c.value = str(year) + "年" + str(month) + "月"
    for col in range(7):  # 一週的每一天
        c = sheet.cell(2, col + 1)
        c.value = dayname[col]
    for col, week in enumerate(cal.monthdayscalendar(year, month)):
        for row, day in enumerate(week):
            if day > 0:
                c = sheet.cell((col + 3), row + 1)
                c.value = day
    workbook.save(savefile)  # 儲存檔案
    return "轉存" + savefile + "了。"


msg = makecalendar(year, month)
print(msg)

print("------------------------------------------------------------")  # 60個

import calendar

value1 = "2024"
value2 = "5"
dayname = ["日", "一", "二", "三", "四", "五", "六"]

fontN = openpyxl.styles.Font(size=24)
fontB = openpyxl.styles.Font(size=24, color="0000FF")
fontR = openpyxl.styles.Font(size=24, color="FF0000")
fillB = openpyxl.styles.PatternFill(patternType="solid", fgColor="AAAAFF")
fillR = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFAAAA")


# 【在Excel檔新增月曆的函數】
def makecalendar(value1, value2):
    year = int(value1)
    month = int(value2)
    savefile = "tmp_excel_openpyx_" + str(year) + "_" + str(month) + "b.xlsx"

    cal = calendar.Calendar(calendar.SUNDAY)
    workbook = openpyxl.Workbook()  # 建立空白的Excel活頁簿物件
    sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)
    for c in ["A", "B", "C", "D", "E", "F", "G"]:
        sheet.column_dimensions[c].width = 20
    c = sheet.cell(1, 4)
    c.value = str(year) + "年" + str(month) + "月"
    c.font = fontN
    for row in range(7):
        c = sheet.cell(2, row + 1)
        c.value = dayname[row]
        c.font = fontN
        c.alignment = openpyxl.styles.Alignment("center")
        if row == 6:
            c.font = fontB
            c.fill = fillB
        if row == 0:
            c.font = fontR
            c.fill = fillR
    for col, week in enumerate(cal.monthdayscalendar(year, month)):
        sheet.row_dimensions[col + 3].height = 50
        for row, day in enumerate(week):
            if day > 0:
                c = sheet.cell((col + 3), row + 1)
                c.value = day
                c.font = fontN
                if row == 6:
                    c.font = fontB
                if row == 0:
                    c.font = fontR
    workbook.save(savefile)  # 儲存檔案
    return "轉存" + savefile + "了。"


msg = makecalendar(value1, value2)
print(msg)
"""

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個
"""
print("匯出 pdf 檔案")

import pathlib
from win32com import client

path = pathlib.Path("data/sales")    #指定相對路徑

xlApp = client.Dispatch("Excel.Application")
for pass_obj in path.iterdir():
    print("\n-------------------------------------------------\n原檔案 :", pass_obj)
    if pass_obj.match("*001.xlsx") and not pass_obj.match("~$"):
        print("match :", pass_obj)
        book = xlApp.workbooks.open(str(pass_obj.resolve()))
        print("aaaa :", str(pass_obj.resolve()))
        for sheet in book.Worksheets:
            print('------------------------------')
            slip_no = str(int(sheet.Range("G2").value))
            file_name = "tmp_sales_data_" + slip_no + ".pdf"
            pdf_path = path / "pdf" / file_name
            sheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            print(pdf_path)
            print(str(pdf_path.resolve()))
            print()
        book.Close()
    else:
        print('不是excel檔案')
xlApp.Quit()

print("------------------------------------------------------------")  # 60個

print("匯出 pdf 檔案")

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm

import shutil
import pathlib  
import datetime

from PIL import Image

def load_informatiom():
    workbook = openpyxl.load_workbook("data/特銷說明會導覽.xlsx")
    sheet = workbook.active
    sale_dict = {} 
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row,1).value == "導覽內容":
            info_list = [sheet.cell(row,2).value]
            for info_row in range(row + 1 , sheet.max_row + 1):
                info_list.append(sheet.cell(info_row,2).value)
            sale_dict.setdefault("導覽內容", info_list)
        elif sheet.cell(row,1).value is not None:     
            sale_dict.setdefault(sheet.cell(row,1).value, sheet.cell(row,2).value)
    return sale_dict


target_dir = 'tmp_pdf'
#準備輸出資料夾 若已存在, 則先刪除再建立 若不存在, 則建立
if os.path.exists(target_dir):
        #os.remove(target_dir)  #存取被拒 不可用
        shutil.rmtree(target_dir)
if not os.path.exists(target_dir):
        os.mkdir(target_dir)

sale_dict = load_informatiom()
path = pathlib.Path(target_dir)
workbook = openpyxl.load_workbook("data/客戶聯絡資料.xlsx")
sheet = workbook["收件人資料"]
for row in range(1, sheet.max_row + 1):
    file_name = (sheet.cell(row,2).value) + "先生／小姐特銷會說明.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("特銷說明會導覽")
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawCentredString(6*cm, 27*cm, sheet.cell(row,2).value + " " \
        + sheet.cell(row,3).value + " 先生／小姐")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm) #在客戶名稱套用底線
    cv.setFont("HeiseiKakuGo-W5", 14)
    #cv.drawCentredString(10*cm, 24*cm, sale_dict["主題"])
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawString(2*cm, 22*cm, "舉辦時間：" + sale_dict["舉辦時間"])
    cv.drawString(2*cm, 21*cm, "舉辦地點：" + sale_dict["舉辦地點"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HeiseiKakuGo-W5", 12)
    for line in sale_dict["導覽內容"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14) # POSITIVE Y moves down!!!
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    #logo_filename = 'C:/_git/vcs/_1.data/______test_files1/__pic/_logo/matlab.png'
    #image =Image.open(logo_filename)
    image =Image.open("data/logo.png")
    cv.drawInlineImage(image,13*cm,13*cm)
    cv.showPage()
    cv.save()

print("------------------------------------------------------------")  # 60個
"""

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

# 插入圖表

print("openpyxl test 11 加入圖表1 雷達圖")

filename_r = "data/radar_chart.xlsx"
workbook = openpyxl.load_workbook(filename_r)

sheet = workbook.active  # 取得開啟試算表後立刻顯示的工作表(即最後編輯的工作表)

data = openpyxl.chart.Reference(
    sheet, min_col=2, max_col=4, min_row=1, max_row=sheet.max_row
)
labels = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = openpyxl.chart.RadarChart()
# 預設為standard
# filled為填色
# chart.type = "filled"
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "F2")

filename_w = "tmp_excel_openpyxl_e_add_radar_chart.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

print("openpyxl test 11 加入圖表2 統計圖")

# 官方範例

workbook = openpyxl.Workbook(write_only=True)
sheet = workbook.create_sheet()

rows = [
    ("Number", "Batch 1", "Batch 2"),
    (2, 10, 30),
    (3, 40, 60),
    (4, 50, 70),
    (5, 20, 10),
    (6, 10, 40),
    (7, 50, 30),
]


for row in rows:
    sheet.append(row)

chart1 = openpyxl.chart.BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Bar Chart"
chart1.y_axis.title = "Test number"
chart1.x_axis.title = "Sample length (mm)"

data = openpyxl.chart.Reference(sheet, min_col=2, min_row=1, max_row=7, max_col=3)
cats = openpyxl.chart.Reference(sheet, min_col=1, min_row=2, max_row=7)
chart1.add_data(data, titles_from_data=True)
chart1.set_categories(cats)
chart1.shape = 4
sheet.add_chart(chart1, "E1")

import copy

chart2 = copy.deepcopy(chart1)
chart2.style = 11
chart2.type = "bar"
chart2.title = "Horizontal Bar Chart"

sheet.add_chart(chart2, "N1")


chart3 = copy.deepcopy(chart1)
chart3.type = "col"
chart3.style = 12
chart3.grouping = "stacked"
chart3.overlap = 100
chart3.title = "Stacked Chart"

sheet.add_chart(chart3, "E15")

chart4 = copy.deepcopy(chart1)
chart4.type = "bar"
chart4.style = 13
chart4.grouping = "percentStacked"
chart4.overlap = 100
chart4.title = "Percent Stacked Chart"

sheet.add_chart(chart4, "N15")

filename_w = "tmp_excel_openpyxl_g_bar.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個

# line_chart.py

from openpyxl.chart import LineChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart1_line.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=2, max_col=9, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = LineChart()
chart.title = "每月業績"
chart.y_axis.title = "銷售數量"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "A9")

filename_w = "tmp_01_line_chart.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("openpyxl test 01 插入圖表")

# easy_bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook("data/python_add_chart2_bubble.xlsx")
sheet = workbook.active

chart = BubbleChart()
chart.style = 18
xvalues = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row)
yvalues = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)
print(xvalues)
print(yvalues)
size = Reference(sheet, min_col=4, min_row=2, max_row=sheet.max_row)
# print(size)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size)
# print(series)
chart.series.append(series)

sheet.add_chart(chart, "F2")

filename_w = "tmp_02_bubble_chart_a.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

# bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook("data/python_add_chart2_bubble.xlsx")
sheet = workbook.active

chart = BubbleChart()
chart.style = 18
for row in range(2, sheet.max_row + 1):
    print(row)
    xvalues = Reference(sheet, min_col=3, min_row=row)
    yvalues = Reference(sheet, min_col=2, min_row=row)
    size = Reference(sheet, min_col=4, min_row=row)
    series = Series(
        values=yvalues, xvalues=xvalues, zvalues=size, title=sheet.cell(row, 1).value
    )
    chart.series.append(series)

sheet.add_chart(chart, "F2")

filename_w = "tmp_02_bubble_chart_b.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

# easy_pie_chart.py

from openpyxl.chart import PieChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart3_pie.xlsx")
sheet = workbook.active
# print(sheet.max_row)
data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "D3")

filename_w = "tmp_03_pie_charta.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

# pie_chart.py

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

workbook = openpyxl.load_workbook("data/python_add_chart3_pie.xlsx")
sheet = workbook.active
# print(sheet.max_row)
data = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

# 切出第一個扇形
slice = DataPoint(idx=0, explosion=10)
chart.series[0].data_points = [slice]

sheet.add_chart(chart, "D3")

filename_w = "tmp_03_pie_chartb.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

# column_chart.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart4_column.xlsx")
sheet = workbook.active
# print(sheet.max_row)
data = Reference(sheet, min_col=3, max_col=3, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = BarChart()
chart.type = "col"
# chart.type = "bar" #橫條圖
# 轉換成橫條圖之後，有部分的客戶名稱會消失，所以要指定chart整體的高度
# chart.height = 10

chart.style = 28  # 1灰色、11藍色、28橙色、30黃色、37圖表的背景設定為淺灰色、45整體的背景設定為黑色
chart.title = "各客戶業績"
chart.y_axis.title = "營業額"
chart.x_axis.title = "客戶名稱"

chart.add_data(data, titles_from_data=True)  # 以當月業績作為圖例
chart.set_categories(labels)
sheet.add_chart(chart, "E3")

filename_w = "tmp_04_column_chart.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

# column_chart_stacked.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart4_column_stacked.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=3, max_col=7, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
# chart.grouping = "percentStacked"
chart.overlap = 100
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "I2")

filename_w = "tmp_04_column_chart_stacked.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個

print("各類別業績（尺寸堆疊長條圖）")

from openpyxl.chart import AreaChart, Reference

workbook = openpyxl.load_workbook("data/python_add_chart5_area.xlsx")
sheet = workbook.active

data = Reference(sheet, min_col=3, max_col=7, min_row=1, max_row=sheet.max_row)
labels = Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=sheet.max_row)
chart = AreaChart()
chart.grouping = "stacked"
# chart.grouping = "percentStacked"
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sheet.add_chart(chart, "I2")

filename_w = "tmp_05_area_chart.xlsx"
workbook.save(filename_w)  # 儲存檔案

print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個
sys.exit()

print("------------------------------------------------------------")  # 60個

# 對齊方式 ST
from openpyxl.styles import Alignment

workbook = openpyxl.Workbook()
sh = workbook.active


sh.column_dimensions["A"].width = 20
sh["a1"] = "left,bottom"
sh["a1"].alignment = Alignment(horizontal="left", vertical="bottom")
sh["a2"] = "center,center"
sh["a2"].alignment = Alignment(horizontal="center", vertical="center")
sh["a3"] = "right,top"
sh["a3"].alignment = Alignment(horizontal="right", vertical="top")
sh["a4"] = "distributed,bottom"
sh["a4"].alignment = Alignment(horizontal="distributed", vertical="bottom")

workbook.save(r"tmp_format_test.xlsx")  # 儲存檔案
# 對齊方式 SP
