import openpyxl

print("------------------------------------------------------------")  # 60個

workbook = openpyxl.load_workbook(r"data\range.xlsx")
sheet = workbook.active

getted_list = []
for row in sheet:
    for cell in row:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for row in range(2, sheet.max_row+1):
    for col in range(2,sheet.max_column+1):
        getted_list.append(sheet.cell(row,col).value)

print(getted_list)

getted_list = []
for rows in sheet["B2":"C3"]:
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for rows in sheet.iter_rows(min_row=2, min_col=2, max_row=3, max_col=3):
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)

print("------------------------------------------------------------")  # 60個

# color_scale.py

import random
from openpyxl.formatting.rule import ColorScaleRule


workbook = openpyxl.Workbook()
sh = workbook.active
values = random.sample(range(50,150), 10)
for i, value in enumerate(values):
    sh.cell(i + 1, 1 ).value = value

two_color_scale = ColorScaleRule(        
    start_type="min", start_color="FF0000",
    end_type="max", end_color="FFFFFF"
)

sh.conditional_formatting.add("A1:A10", two_color_scale)
 

workbook.save(r"tmp_01_color_scale.xlsx")

print("------------------------------------------------------------")  # 60個

# fill_red.py

import random
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

workbook = openpyxl.Workbook()
sh = workbook.active

values = random.sample(range(50,150), 10)

for i, value in enumerate(values):
    sh.cell(i + 1, 1 ).value = value

less_than_rule = CellIsRule( 
    operator="lessThan",
    formula=[100],
    stopIfTrue=True,
    fill=PatternFill("solid", start_color="FF0000", end_color="FF0000")
)
sh.conditional_formatting.add("A1:A10", less_than_rule)

workbook.save(r"tmp_02_fill_red.xlsx")

print("------------------------------------------------------------")  # 60個

# format_sheet.py

from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

#常數
TITLE_CELL_COLOR = "AA8866"

workbook = openpyxl.load_workbook("data\orders_aggregate.xlsx")
sh = workbook.active

#固定為第1列、第2欄
sh.freeze_panes = "C2"
#設定欄寬
col_widths = {"A":8, "B":15, "C":10, "D":10, "E":10, \
    "F":10, "G":10, "H":10}
for col_name in col_widths:
    sh.column_dimensions[col_name].width = col_widths[col_name]

#隱藏A欄
#sh.column_dimensions["A"].hidden=False

for i in range(2, sh.max_row+1):
    sh.row_dimensions[i].height = 18
    for j in range(3, sh.max_column+1):
        #千分位樣式
        sh.cell(row=i,column=j).number_format = "#,##0" 
        if j == 8:
            sh.cell(row=i,column=j).font = Font(bold=True)


#建立字體
font_header = Font(name="MS PGothic",size=12,bold=True,color="FFFFFF")

for rows in sh["A1":"H1"]:
    for cell in rows:
        #儲存格背景色
        cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
        #水平置中
        cell.alignment = Alignment(horizontal="distributed")
        #設定字型
        cell.font = font_header

side = Side(style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
for row in sh:
    for cell in row:
        cell.border = border
        #sh[cell.coordinate].border = border

workbook.save("tmp_03_orders_aggregate_ed.xlsx")

print("------------------------------------------------------------")  # 60個

# format_sheet2.py

workbook = openpyxl.Workbook()
sh = workbook.active

sh["b2"] = "合併儲存格的測試"
sh.merge_cells("b2:c2")
sh["b2"].alignment = openpyxl.styles.Alignment(horizontal="center")
#sh.unmerge_cells("b2:c2")

workbook.save(r"tmp_04_format_test.xlsx")

print("------------------------------------------------------------")  # 60個

# format_sheet3.py

from openpyxl.styles import Border, Side

workbook = openpyxl.load_workbook(r"data\border.xlsx")
sh = workbook.active

side1 = Side(style="hair", color="00FF00")
side2 = Side(style="dashDotDot", color="0000FF")
side3 = Side(style="double", color="FF0000")
for rows in sh["B2":"C4"]:
    for cell in rows:
        cell.border = Border(left=side1, right=side1, top=side1, bottom=side1 )
for rows in sh["E2":"F4"]:
    for cell in rows:
        cell.border = Border(left=side2, right=side2, top=side2, bottom=side2)
for rows in sh["H2":"I4"]:
    for cell in rows:
        cell.border = Border(left=side3, right=side3, top=side3, bottom=side3 )

workbook.save(r"tmp_05_border_ed.xlsx")

print("------------------------------------------------------------")  # 60個

from openpyxl.chart import AreaChart, Reference

workbook = openpyxl.load_workbook(r"data\area_chart.xlsx")
sh = workbook.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = AreaChart()
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
workbook.save(r"tmp_06_area_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook(r"data\bubble_chart.xlsx")
sh = workbook.active

chart = BubbleChart()
chart.style = 18
for row in range(2,sh.max_row+1):
    xvalues = Reference(sh, min_col=3, min_row=row)
    yvalues = Reference(sh, min_col=2, min_row=row)
    size = Reference(sh, min_col=4, min_row=row)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=size,  title=sh.cell(row,1).value)
    chart.series.append(series)

sh.add_chart(chart, "F2")
workbook.save(r"tmp_07_bubble_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# column_chart.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data\column_chart.xlsx")
sh = workbook.active
#print(sh.max_row)
data = Reference(sh, min_col=3, max_col=3, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
#chart.type = "bar" #橫條圖
# 轉換成橫條圖之後，有部分的客戶名稱會消失，所以要指定chart整體的高度
#chart.height = 10

chart.style = 28    #1灰色、11藍色、28橙色、30黃色、37圖表的背景設定為淺灰色、45整體的背景設定為黑色
chart.title = "各客戶業績"
chart.y_axis.title = "營業額"
chart.x_axis.title = "客戶名稱"

chart.add_data(data,titles_from_data=True)  #以當月業績作為圖例
chart.set_categories(labels)
sh.add_chart(chart, "E3")

workbook.save("tmp_08_column_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# column_chart_stacked.py

from openpyxl.chart import BarChart, Reference

workbook = openpyxl.load_workbook("data\column_chart_stacked.xlsx")
sh = workbook.active

data = Reference(sh, min_col=3, max_col=7, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=2, max_col=2, min_row=2, max_row=sh.max_row)
chart = BarChart()
chart.type = "col"
chart.grouping = "stacked"
#chart.grouping = "percentStacked"
chart.overlap = 100
chart.title = "各類別業績（尺寸堆疊長條圖）"
chart.x_axis.title = "分類"
chart.y_axis.title = "尺寸"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "I2")
workbook.save("tmp_09_column_chart_stacked.xlsx")

print("------------------------------------------------------------")  # 60個

# easy_bubble_chart.py

from openpyxl.chart import Series, Reference, BubbleChart

workbook = openpyxl.load_workbook(r"data\bubble_chart.xlsx")
sh = workbook.active

chart = BubbleChart()
chart.style = 18
xvalues = Reference(sh, min_col=3, min_row=2, max_row=sh.max_row)
yvalues = Reference(sh, min_col=2, min_row=2, max_row=sh.max_row)
size = Reference(sh, min_col=4, min_row=2, max_row=sh.max_row)
series = Series(values=yvalues, xvalues=xvalues, zvalues=size)
chart.series.append(series)

sh.add_chart(chart, "F2")
workbook.save(r"tmp_10_bubble_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# easy_pie_chart.py

from openpyxl.chart import PieChart, Reference

workbook = openpyxl.load_workbook("data\pie_chart.xlsx")
sh = workbook.active
#print(sh.max_row)
data = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "D3")
workbook.save("tmp_11_pie_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# line_chart.py

from openpyxl.chart import LineChart, Reference

workbook = openpyxl.load_workbook("data\line_chart.xlsx")
sh = workbook.active

data = Reference(sh, min_col=2, max_col=9, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = LineChart()
chart.title = "每月業績"
chart.y_axis.title = "銷售數量"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

sh.add_chart(chart, "A9")
workbook.save("tmp_12_line_chart.xlsx")

print("------------------------------------------------------------")  # 60個

# pie_chart.py

from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

workbook = openpyxl.load_workbook("data\pie_chart.xlsx")
sh = workbook.active
#print(sh.max_row)
data = Reference(sh, min_col=2, min_row=1, max_row=sh.max_row)
labels = Reference(sh, min_col=1, min_row=2, max_row=sh.max_row)

chart = PieChart()
chart.title = "各部門業績"
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)

#切出第一個扇形
slice = DataPoint(idx=0, explosion=10)
chart.series[0].data_points = [slice]

sh.add_chart(chart, "D3")
workbook.save("tmp_13_pie_chart.xlsx")

print("------------------------------------------------------------")  # 60個

import pathlib  
from win32com import client

path = pathlib.Path("data")    #指定相對路徑

xlApp = client.Dispatch("Excel.Application")
for pass_obj in path.iterdir():
    print("\n-------------------------------------------------\n原檔案 :", pass_obj)
    if pass_obj.match("*.xlsx"):
        print("match :", pass_obj)
        book = xlApp.workbooks.open(str(pass_obj.resolve()))
        print("aaaa :", str(pass_obj.resolve()))
        for sheet in book.Worksheets:
            print('------------------------------')
            slip_no = str(int(sheet.Range("G2").value))
            file_name = slip_no + ".pdf"
            pdf_path = path / "pdf" / file_name
            
            sheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            print(pdf_path)
            print(str(pdf_path.resolve()))
            print()

        book.Close()
xlApp.Quit()

print("------------------------------------------------------------")  # 60個

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm

import os
import shutil
import pathlib  
import datetime

from PIL import Image

def load_informatiom():
    workbook = openpyxl.load_workbook("data/特銷說明會導覽.xlsx")
    sh = workbook.active
    sale_dict = {} 
    for row in range(1, sh.max_row + 1):
        if sh.cell(row,1).value == "導覽內容":
            info_list = [sh.cell(row,2).value]
            for info_row in range(row + 1 , sh.max_row + 1):
                info_list.append(sh.cell(info_row,2).value)
            sale_dict.setdefault("導覽內容", info_list)
        elif sh.cell(row,1).value is not None:     
            sale_dict.setdefault(sh.cell(row,1).value, sh.cell(row,2).value)
    return sale_dict


target_dir = 'tmp_pdf'
#準備輸出資料夾 若已存在, 則先刪除再建立 若不存在, 則建立
if os.path.exists(target_dir):
        #os.remove(target_dir)  #存取被拒 不可用
        shutil.rmtree(target_dir)
if not os.path.exists(target_dir):
        os.mkdir(target_dir)

sale_dict = load_informatiom()
path = pathlib.Path(target_dir)
workbook = openpyxl.load_workbook("data/客戶聯絡資料.xlsx")
sh = wb["收件人資料"]
for row in range(1, sh.max_row + 1):
    file_name = (sh.cell(row,2).value) + "先生／小姐特銷會說明.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("特銷說明會導覽")
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawCentredString(6*cm, 27*cm, sh.cell(row,2).value + " " \
        + sh.cell(row,3).value + " 先生／小姐")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm) #在客戶名稱套用底線
    cv.setFont("HeiseiKakuGo-W5", 14)
    #cv.drawCentredString(10*cm, 24*cm, sale_dict["主題"])
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawString(2*cm, 22*cm, "舉辦時間：" + sale_dict["舉辦時間"])
    cv.drawString(2*cm, 21*cm, "舉辦地點：" + sale_dict["舉辦地點"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HeiseiKakuGo-W5", 12)
    for line in sale_dict["導覽內容"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14) # POSITIVE Y moves down!!!
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    #logo_filename = 'C:/_git/vcs/_1.data/______test_files1/__pic/_logo/matlab.png'
    #image =Image.open(logo_filename)
    image =Image.open("data/logo.png")
    cv.drawInlineImage(image,13*cm,13*cm)
    cv.showPage()
    cv.save()

print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print('------------------------------------------------------------')	#60個


