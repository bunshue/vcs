import openpyxl

wb = openpyxl.load_workbook(r"data\range.xlsx")
sheet = wb.active

getted_list = []
for row in sheet:
    for cell in row:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for row in range(2, sheet.max_row+1):
    for col in range(2,sheet.max_column+1):
        getted_list.append(sheet.cell(row,col).value)

print(getted_list)

getted_list = []
for rows in sheet["B2":"C3"]:
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)

getted_list = []
for rows in sheet.iter_rows(min_row=2, min_col=2, max_row=3, max_col=3):
    for cell in rows:
        getted_list.append(cell.value)

print(getted_list)



print("------------------------------------------------------------")  # 60個




#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\color_scale.py

import openpyxl
import random
from openpyxl.formatting.rule import ColorScaleRule


wb = openpyxl.Workbook()
sh = wb.active
values = random.sample(range(50,150), 10)
for i, value in enumerate(values):
    sh.cell(i + 1, 1 ).value = value

two_color_scale = ColorScaleRule(        
    start_type="min", start_color="FF0000",
    end_type="max", end_color="FFFFFF"
)

sh.conditional_formatting.add("A1:A10", two_color_scale)
 

wb.save(r"tmp_color_scale.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\fill_red.py

import openpyxl
import random
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


wb = openpyxl.Workbook()
sh = wb.active

values = random.sample(range(50,150), 10)

for i, value in enumerate(values):
    sh.cell(i + 1, 1 ).value = value

less_than_rule = CellIsRule( 
    operator="lessThan",
    formula=[100],
    stopIfTrue=True,
    fill=PatternFill("solid", start_color="FF0000", end_color="FF0000")
)
sh.conditional_formatting.add("A1:A10", less_than_rule)

wb.save(r"tmp_fill_red.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\format_sheet.py

import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

#常數
TITLE_CELL_COLOR = "AA8866"

wb = openpyxl.load_workbook("data\orders_aggregate.xlsx")
sh = wb.active

#固定為第1列、第2欄
sh.freeze_panes = "C2"
#設定欄寬
col_widths = {"A":8, "B":15, "C":10, "D":10, "E":10, \
    "F":10, "G":10, "H":10}
for col_name in col_widths:
    sh.column_dimensions[col_name].width = col_widths[col_name]

#隱藏A欄
#sh.column_dimensions["A"].hidden=False

for i in range(2, sh.max_row+1):
    sh.row_dimensions[i].height = 18
    for j in range(3, sh.max_column+1):
        #千分位樣式
        sh.cell(row=i,column=j).number_format = "#,##0" 
        if j == 8:
            sh.cell(row=i,column=j).font = Font(bold=True)


#建立字體
font_header = Font(name="MS PGothic",size=12,bold=True,color="FFFFFF")

for rows in sh["A1":"H1"]:
    for cell in rows:
        #儲存格背景色
        cell.fill = PatternFill(patternType="solid", fgColor=TITLE_CELL_COLOR)
        #水平置中
        cell.alignment = Alignment(horizontal="distributed")
        #設定字型
        cell.font = font_header

side = Side(style="thin", color="000000")
border = Border(left=side, right=side, top=side, bottom=side)
for row in sh:
    for cell in row:
        cell.border = border
        #sh[cell.coordinate].border = border


wb.save("tmp_orders_aggregate_ed.xlsx")


print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\format_sheet1.py


print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\format_sheet2.py

import openpyxl

wb = openpyxl.Workbook()
sh = wb.active

sh["b2"] = "合併儲存格的測試"
sh.merge_cells("b2:c2")
sh["b2"].alignment = openpyxl.styles.Alignment(horizontal="center")
#sh.unmerge_cells("b2:c2")

wb.save(r"tmp_format_test.xlsx")

print("------------------------------------------------------------")  # 60個

#檔案 : C:\_git\vcs\_4.python\__code\圖解零基礎入門Excel╳Python高效工作術\05\python_prg\format_sheet3.py

import openpyxl
from openpyxl.styles import Border, Side

wb = openpyxl.load_workbook(r"data\border.xlsx")
sh = wb.active

side1 = Side(style="hair", color="00FF00")
side2 = Side(style="dashDotDot", color="0000FF")
side3 = Side(style="double", color="FF0000")
for rows in sh["B2":"C4"]:
    for cell in rows:
        cell.border = Border(left=side1, right=side1, top=side1, bottom=side1 )
for rows in sh["E2":"F4"]:
    for cell in rows:
        cell.border = Border(left=side2, right=side2, top=side2, bottom=side2)
for rows in sh["H2":"I4"]:
    for cell in rows:
        cell.border = Border(left=side3, right=side3, top=side3, bottom=side3 )


wb.save(r"tmp_border_ed.xlsx")

print("------------------------------------------------------------")  # 60個


