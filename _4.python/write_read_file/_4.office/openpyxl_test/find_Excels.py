#【函數: 搜尋Excel檔案之内的字串】

import pathlib
import openpyxl

infolder = "testfolder"
value1 = "這個是"
value2 = "*.xlsx"


def findfile(readfile, findword):
    try:
        msg = ""
        wb = openpyxl.load_workbook(readfile)
        cnt = 0
        for sheetname in wb.sheetnames:             #依序開啟所有工作表
            sheet = wb[sheetname]
            for c in range(1, sheet.max_column+1):  #各欄的
                for r in range(1, sheet.max_row+1): #各列的
                    cell = sheet.cell(row=r, column=c)
                    cellstr = str(cell.value)       #將儲存格的值轉換成字串
                    cnt += cellstr.count(findword)  #要搜尋的字串的個數
        if cnt > 0:                         #找到的話
            msg = readfile+"："+"找到" + str(cnt)+"個了。\n"
        return msg
    except:
        return readfile + "：程式執行失敗。"

#【函數: 搜尋資料夾與子資料夾的文字檔】
def findfiles(infolder, findword, ext):
    msg = ""
    filelist = []
    for f in pathlib.Path(infolder).rglob(ext): #將這個資料夾以及子資料夾的所有檔案
        filelist.append(str(f))         #新增至列表
    for filename in sorted(filelist):   #再替每個檔案排序
        msg += findfile(filename, findword)
    return msg


#【執行函數】
msg = findfiles(infolder, value1, value2)
print(msg)
