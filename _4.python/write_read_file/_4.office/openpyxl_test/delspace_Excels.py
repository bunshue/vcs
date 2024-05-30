"""

看起來是 把excel檔的 空白 tab 刪除之

"""

import pathlib
import openpyxl

infolder = "testfolder"
value1 = "outputfolder"
value2 = "*.xlsx"

#【函數：執行unicode正規化】
def normalizefile(readfile, savefolder):
    print('normalizefile :', readfile)
    try:
        msg = ""
        wb = openpyxl.load_workbook(readfile)
        for sheetname in wb.sheetnames:             #依序開啟所有工作表
            sheet = wb[sheetname]
            for c in range(1, sheet.max_column+1):  #各欄的
                for r in range(1, sheet.max_row+1): #各列的
                    cell = sheet.cell(row=r, column=c)
                    if type(cell.value) is str: #儲存格的值若是字串就刪除空白字元
                        cell.value = cell.value.replace(" ","")
                        cell.value = cell.value.replace("　","")
                        cell.value = cell.value.replace("\t","")
        savedir = pathlib.Path(savefolder)
        savedir.mkdir(exist_ok=True)            #建立轉存資料夾
        filename = pathlib.Path(readfile).name
        newname = savedir.joinpath(filename)
        wb.save(newname)                        #Excel轉存檔案
        msg = "在" + savefolder+"轉存"+ filename + "了喲。\n"
        return msg
    except:
        return readfile + "：程式執行失敗。"

#【函數: 對資料夾之內的文字檔執行unicode正規化】
def normalizefiles(infolder, savefolder, ext):
    msg = ""
    filelist = []
    for p in pathlib.Path(infolder).glob(ext):  #將這個資料夾的檔案
        filelist.append(str(p))         #新增至列表
    for filename in sorted(filelist):   #再替每個檔案排序
        msg += normalizefile(filename, savefolder)
    return msg


#【執行函數】
msg = normalizefiles(infolder, value1, value2)
print(msg)
