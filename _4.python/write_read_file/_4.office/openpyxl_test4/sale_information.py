import pathlib  
import openpyxl 
from win32com import client

path = pathlib.Path("data")    #指定相對路徑

xlApp = client.Dispatch("Excel.Application")
for pass_obj in path.iterdir():
    print(pass_obj)
    if pass_obj.match("*.xlsx"):
        print("match :", pass_obj)
        book = xlApp.workbooks.open(str(pass_obj.resolve()))
        for sheet in book.Worksheets:
            print('------------------------------')
            slip_no = str(int(sheet.Range("G2").value))
            file_name = slip_no + ".pdf"
            pdf_path = path / "pdf" / file_name
            
            sheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            print(pdf_path)
            print(str(pdf_path.resolve()))
            print()

        book.Close()
xlApp.Quit()

print("------------------------------------------------------------")  # 60個


print('------------------------------------------------------------')	#60個




from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm

import os
import shutil
import pathlib  
import datetime
import openpyxl 

from PIL import Image

def load_informatiom():
    wb = openpyxl.load_workbook("特銷說明會導覽.xlsx")
    sh = wb.active
    sale_dict = {} 
    for row in range(1, sh.max_row + 1):
        if sh.cell(row,1).value == "導覽內容":
            info_list = [sh.cell(row,2).value]
            for info_row in range(row + 1 , sh.max_row + 1):
                info_list.append(sh.cell(info_row,2).value)
            sale_dict.setdefault("導覽內容", info_list)
        elif sh.cell(row,1).value is not None:     
            sale_dict.setdefault(sh.cell(row,1).value, sh.cell(row,2).value)
    return sale_dict


target_dir = 'tmp_pdf'
#準備輸出資料夾 若已存在, 則先刪除再建立 若不存在, 則建立
if os.path.exists(target_dir):
        #os.remove(target_dir)  #存取被拒 不可用
        shutil.rmtree(target_dir)
if not os.path.exists(target_dir):
        os.mkdir(target_dir)

sale_dict = load_informatiom()
path = pathlib.Path(target_dir)
wb = openpyxl.load_workbook("客戶聯絡資料.xlsx")
sh = wb["收件人資料"]
for row in range(1, sh.max_row + 1):
    file_name = (sh.cell(row,2).value) + "先生／小姐特銷會說明.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("特銷說明會導覽")
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawCentredString(6*cm, 27*cm, sh.cell(row,2).value + " " \
        + sh.cell(row,3).value + " 先生／小姐")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm) #在客戶名稱套用底線
    cv.setFont("HeiseiKakuGo-W5", 14)
    #cv.drawCentredString(10*cm, 24*cm, sale_dict["主題"])
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawString(2*cm, 22*cm, "舉辦時間：" + sale_dict["舉辦時間"])
    cv.drawString(2*cm, 21*cm, "舉辦地點：" + sale_dict["舉辦地點"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HeiseiKakuGo-W5", 12)
    for line in sale_dict["導覽內容"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14) # POSITIVE Y moves down!!!
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    #logo_filename = 'C:/_git/vcs/_1.data/______test_files1/__pic/_logo/matlab.png'
    #image =Image.open(logo_filename)
    image =Image.open("logo.png")
    cv.drawInlineImage(image,13*cm,13*cm)
    cv.showPage()
    cv.save()



print("------------------------------------------------------------")  # 60個


print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print('------------------------------------------------------------')	#60個

