"""
讀寫 Excel 檔案, 使用 openpyxl


"""

import os
import sys
import time
import openpyxl

print("------------------------------------------------------------")  # 60個
'''
print("匯出 pdf 檔案")

import pathlib
from win32com import client

path = pathlib.Path("data/sales")    #指定相對路徑

xlApp = client.Dispatch("Excel.Application")
for pass_obj in path.iterdir():
    print("\n-------------------------------------------------\n原檔案 :", pass_obj)
    if pass_obj.match("*001.xlsx") and not pass_obj.match("~$"):
        print("match :", pass_obj)
        book = xlApp.workbooks.open(str(pass_obj.resolve()))
        print("aaaa :", str(pass_obj.resolve()))
        for sheet in book.Worksheets:
            print('------------------------------')
            slip_no = str(int(sheet.Range("G2").value))
            file_name = "tmp_sales_data_" + slip_no + ".pdf"
            pdf_path = path / "pdf" / file_name
            sheet.ExportAsFixedFormat(0, str(pdf_path.resolve()))
            print(pdf_path)
            print(str(pdf_path.resolve()))
            print()
        book.Close()
    else:
        print('不是excel檔案')
xlApp.Quit()

print("------------------------------------------------------------")  # 60個

print("匯出 pdf 檔案")

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, portrait
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib.units import cm

import os
import shutil
import pathlib  
import datetime

from PIL import Image

def load_informatiom():
    workbook = openpyxl.load_workbook("data/特銷說明會導覽.xlsx")
    sheet = workbook.active
    sale_dict = {} 
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row,1).value == "導覽內容":
            info_list = [sheet.cell(row,2).value]
            for info_row in range(row + 1 , sheet.max_row + 1):
                info_list.append(sheet.cell(info_row,2).value)
            sale_dict.setdefault("導覽內容", info_list)
        elif sheet.cell(row,1).value is not None:     
            sale_dict.setdefault(sheet.cell(row,1).value, sheet.cell(row,2).value)
    return sale_dict


target_dir = 'tmp_pdf'
#準備輸出資料夾 若已存在, 則先刪除再建立 若不存在, 則建立
if os.path.exists(target_dir):
        #os.remove(target_dir)  #存取被拒 不可用
        shutil.rmtree(target_dir)
if not os.path.exists(target_dir):
        os.mkdir(target_dir)

sale_dict = load_informatiom()
path = pathlib.Path(target_dir)
workbook = openpyxl.load_workbook("data/客戶聯絡資料.xlsx")
sheet = workbook["收件人資料"]
for row in range(1, sheet.max_row + 1):
    file_name = (sheet.cell(row,2).value) + "先生／小姐特銷會說明.pdf"
    out_path =  path / file_name
    cv = canvas.Canvas(str(out_path), pagesize=portrait(A4))
    cv.setTitle("特銷說明會導覽")
    pdfmetrics.registerFont(UnicodeCIDFont("HeiseiKakuGo-W5"))
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawCentredString(6*cm, 27*cm, sheet.cell(row,2).value + " " \
        + sheet.cell(row,3).value + " 先生／小姐")
    cv.line(1.8*cm, 26.8*cm,10.8*cm,26.8*cm) #在客戶名稱套用底線
    cv.setFont("HeiseiKakuGo-W5", 14)
    #cv.drawCentredString(10*cm, 24*cm, sale_dict["主題"])
    cv.setFont("HeiseiKakuGo-W5", 12)
    cv.drawString(2*cm, 22*cm, "舉辦時間：" + sale_dict["舉辦時間"])
    cv.drawString(2*cm, 21*cm, "舉辦地點：" + sale_dict["舉辦地點"])

    textobject = cv.beginText()
    textobject.setTextOrigin(2*cm, 19*cm,)
    textobject.setFont("HeiseiKakuGo-W5", 12)
    for line in sale_dict["導覽內容"]:
        textobject.textOut(line)
        textobject.moveCursor(0,14) # POSITIVE Y moves down!!!
    
    cv.drawText(textobject)
    now = datetime.datetime.now()
    cv.drawString(14.4*cm, 14.8*cm, now.strftime("%Y/%m/%d"))
    #logo_filename = 'C:/_git/vcs/_1.data/______test_files1/__pic/_logo/matlab.png'
    #image =Image.open(logo_filename)
    image =Image.open("data/logo.png")
    cv.drawInlineImage(image,13*cm,13*cm)
    cv.showPage()
    cv.save()

print("------------------------------------------------------------")  # 60個
'''


import openpyxl
from openpyxl import Workbook
#建立一個空白活頁簿
wb=Workbook()
#選取正在工作中的工作表，空白活頁簿有一個預存工作表，名稱是Sheet
ws=wb.active
#設定工作表名稱為Sheet1
ws.title="Sheet1"
#指定字串給A1儲存格
ws['A1']='1'
#指定數值給A2儲存格
ws['A2']=1
#指定數值給B2儲存格
ws['B2']=2
#指定數值公式給C2儲存格
ws['C2']='=SUM(A2:B2)'
#指定數值計算結果給A3儲存格
ws['A3']=1+2
#指定字串計算結果給B3儲存格
ws['B3']='1'+'2'
#指定數值計算公式給A4儲存格
ws['A4']="=1+2"
#指定字串公式給C2儲存格
ws['B4']='=CONCATENATE("1","2")'
#新增一個工作表名稱為Sheet2
wb.create_sheet("Sheet2")
#選擇Sheet2工作表為正在工作中的工作表
ws=wb.get_sheet_by_name("Sheet2")
#一次填入一列數值
ws.append([1,2,3])
#一次填入一列文字
ws.append(['1','2','3'])
#建立圖片物件
img=openpyxl.drawing.image.Image('data/Google-Colab-Guide-1024x683.jpg')
img.height=img.height*0.1
img.width=img.width*0.1
#加入圖片於C1儲存格
ws.add_image(img,'C1')
#儲存
wb.save('tmp_excel.xlsx')

print("------------------------------------------------------------")  # 60個
print("作業完成")
print("------------------------------------------------------------")  # 60個

print('------------------------------------------------------------')	#60個


print("------------------------------------------------------------")  # 60個

print("------------------------------------------------------------")  # 60個

