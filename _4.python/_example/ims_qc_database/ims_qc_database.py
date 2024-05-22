"""
ims_qc_database

把原始資料.csv檔放在某個資料夾內

這個轉換程式去那個資料夾內撈.csv檔案

每撈一個檔案，分析裡面的資料，就匯出至excel檔，

匯出至excel檔時，由csv資料找到年月日與各站別，決定輸出檔案之檔名

要先檢查檔案是否存在，若不存在，則要寫檔頭，否則直接輸出資料至excel檔即可。

匯出資料完畢後，將此.csv檔移到別的資料夾

T 060201 01 環境溫濕度管制紀錄表_B (1月).xlsx


每個csv檔匯出的csv_data檢查一次日期

若是未曾出現的年月excel檔，
則建立一空白之excel檔，裡面有8個頁箋(其實就是複製template檔)


T 060201 01 環境溫濕度管制紀錄表_B (1月).xlsx


402除菌區數據紀錄_211201_000000



最後要匯出的資料:

每個月一個excel檔案
檔案裡面有8個頁箋 放8個區域資料

"""

print('------------------------------------------------------------')	#60個
print('準備工作')

import os
import sys
import csv
import time
import glob
import shutil
import openpyxl
import tkinter as tk
from tkinter.filedialog import askopenfile #tk之openFileDialog
from tkinter.filedialog import asksaveasfile #tk之saveFileDialog
from tkinter.filedialog import askopenfilename
from tkinter.filedialog import asksaveasfilename

flag_debug_mode = True

# 建立csv二維串列資料
csv_data = list()

stage_no = [
'402除菌區', '403包裝區', '404組裝區', '405燒機測試區',
'406電信測試區', '407原料除菌區', '408清洗區', '409無塵室'
]

stage = -1
tablename = ""
excel_filename = ""
sheetname = '1月402' # 1月402 1月402 1月402 1月402 1月402 1月402

dummy_data = 'abcd'

print('------------------------------------------------------------')	#60個

#公用副程式

#清除記憶體內的資料
def clear_csv_data():
    csv_data = [
    ]
    
print('------------------------------------------------------------')	#60個

def process_csv_file1(filename, source):
    global stage
    global tablename
    #global filename
    global sheetname
    
    print("aaa :", filename)
    filename = filename.replace(source+"\\", "")
    print("bbb :", filename)

    print("ccc :", filename)

    #year_month_day_data = filename
   
    if filename.startswith('402除菌區') == True:
        print('402除菌區')
        stage = 402
        tablename = 'Room402'
    elif filename.startswith('403包裝區') == True:
        print('403包裝區')
        stage = 403
        tablename = 'Room403'
    elif filename.startswith('404組裝區') == True:
        print('404組裝區')
        stage = 404
        tablename = 'Room404'
    elif filename.startswith('405燒機測試區') == True:
        print('405燒機測試區')
        stage = 405
        tablename = 'Room405'
    elif filename.startswith('406電信測試區') == True:
        print('406電信測試區')
        stage = 406
        tablename = 'Room406'
    elif filename.startswith('407原料除菌區') == True:
        print('407原料除菌區')
        stage = 407
        tablename = 'Room407'
    elif filename.startswith('408清洗區') == True:
        print('408清洗區')
        stage = 408
        tablename = 'Room408'
    elif filename.startswith('409無塵室') == True:
        print('409無塵室')
        stage = 409
        tablename = 'Room409'
    else:
        print('第XXXXXXXXX站')
        stage = 0
    if stage == 0:
        print('不合法的csv檔 : ', filename, ',\t跳過')

def save_data_in_list(datas):
    #print(type(datas))
    data_length = len(datas)
    if stage == 1:
        print('第1站')
    else:
        print('第XXXXXXXXX站')

def save_data_in_list2(datas):
    #print(type(datas))
    data_length = len(datas)
    length = len(csv_data)
    if length == 0:
        csv_data.append(datas[0])   #print('加入檔頭')
    for i in range(1, data_length):
        csv_data.append(datas[i])

def process_csv_file2(filename):
    global stage
    global tablename
    #global filename
    global sheetname

    #with open(filename, newline = '') as csvfile:
    with open(filename, encoding = 'big5') as csvfile:
        rows = csv.reader(csvfile)  # 讀取 csv 檔案內容
        datas = list(rows)    #將資料轉成list

    length = len(datas)
    #print('len = ', length)

    save_data_in_list(datas)

    #print(datas)
    data_column = len(datas[0])
    #print('data_column = ', data_column)

    for row in datas:
        for i in range(0, len(row)):
            row[i] = row[i].strip()

    '''
    cnt = 0
    for row in datas:
        print(row)
        print(type(row))
        #print(len(row))
        cnt += 1
        if cnt == 3:
            break
    '''

def process_csv_file3(filename):
    global stage
    global tablename
    #global filename
    global sheetname

    print("process_csv_file3")
    print(filename)
    print(tablename)

    #with open(filename, newline = '') as csvfile:
    with open(filename, encoding = 'utf_16_le') as csvfile:
        rows = csv.reader(csvfile, delimiter="\t")  # 讀取 csv 檔案內容
        datas = list(rows)    #將資料轉成list

    length = len(datas)
    print('len = ', length)

    save_data_in_list2(datas)
    #print(datas)
    
    data_column = len(datas[0])
    print('data_column = ', data_column)

    for row in datas:
        for i in range(0, len(row)):
            row[i] = row[i].strip()

    '''
    cnt = 0
    for row in datas:
        print(row)
        print(type(row))
        #print(len(row))
        cnt += 1
        if cnt == 3:
            break
    '''

    
def import_csv_data():
    global stage
    global tablename
    global filename
    global sheetname
    global csv_data

    csv_data = list() #重新計數csv資料

    source_dir = 'QC/csv'
    target_dir = 'QC/csv_old'

    if flag_debug_mode == True:
        source_dir = 'QC_debug/csv'
        target_dir = 'QC_debug/csv_old'

    #準備輸出資料夾 若不存在, 則建立
    if not os.path.exists(target_dir):
            os.mkdir(target_dir)
            #os.makedirs(target_dir, exist_ok = True)

    #尋找檔案
    import glob
    print('222尋找目前目錄下之 *.csv')
    print((source_dir + "/*.csv"))
    files = glob.glob(source_dir + "/*.csv") 
    for filename in files:
        print(filename)
        stage = 0
        tablename = 'table00'
        process_csv_file1(filename, source_dir)
        if stage != 0:
            print('繼續')
            process_csv_file3(filename)
            #若能正確處理完畢, 再搬到old資料夾
            #shutil.move(filename, target_dir)

    message = "匯入生產資料 完成"
    print(message)
    text1.insert('end', message)

def show_info():
    print('show_info')
    length = len(csv_data)
    print('資料長度 : ', length)

def show_csv_data():
    length = len(csv_data)
    print('csv_data 資料長度 : ', length)
    for i in range(0, length):
        print(csv_data[i])

def export_data_to_excel0(csv_data):
    global sheetname
    print('真的匯出資料到excel')
    print('取得excel檔名 :', excel_filename)

    length = len(csv_data)
    print('共有資料 :', length, '筆')
    """
    if length > 0:
        for _ in csv_data:
            print(_)
    """
    print(sheetname)
    print(sheetname)
    print(sheetname)
    print(sheetname)

    workbook = openpyxl.load_workbook(excel_filename)
    print('所有工作表名稱 :', workbook.sheetnames)

    names = workbook.sheetnames  # 讀取 excel檔案 裏所有工作表名稱
    print('所有工作表名稱 :', names)
    
    # 取得 指定的 工作表
    print(sheetname)
    print(sheetname)
    print(sheetname)
    sheet = workbook[sheetname]

    if length > 0:
        for cc in csv_data:
            #print(cc)
            time = cc[1][:2]
            if time == "09" or time == "13" or time == "17":
                print(cc)
                day = int(cc[0][-2:])
                print(day)
                data1 = str(cc[4])#溫度
                data2 = str(cc[2])#濕度
                data3 = str(cc[3])#靜壓
                print(data1, data2, data3)
                if time == "09":
                    sheet.cell(8, 2+day).value = str(data1)
                    sheet.cell(9, 2+day).value = str(data2)
                    sheet.cell(10, 2+day).value = str(data3)
                elif time == "13":
                    sheet.cell(11, 2+day).value = str(data1)
                    sheet.cell(12, 2+day).value = str(data2)
                    sheet.cell(13, 2+day).value = str(data3)
                elif time == "17":
                    sheet.cell(14, 2+day).value = str(data1)
                    sheet.cell(15, 2+day).value = str(data2)
                    sheet.cell(16, 2+day).value = str(data3)
    workbook.save(excel_filename)  # 儲存檔案

def check_excel_filename(data1, data2):
    global sheetname
    print(type(data1))
    print(data1)
    year, month, day = data1.split('/')
    print(year)
    print(month)
    print(day)
    filename = "T 060201 01 環境溫濕度管制紀錄表_B (%d月).xlsx" % int(month)
    room = data2[:3]
    sheetname = "%d月"%int(month)+room

    if os.path.exists(filename):
        print(filename, "檔案存在")
    else:
        print(filename, "檔案不存在, 建立一個")
        shutil.copy("template/template.xlsx", filename)  # 檔案複製
        if month != "01":
            print('需要改sheet名')
            year_month = "年月：     20%s      年        %02d        月" % (str(year), int(month))
            print(year_month)
           
            workbook = openpyxl.load_workbook(filename)

            sheet0 = workbook.worksheets[0]
            sheet0.title = "%d月" % int(month)+"402"
            sheet0.cell(4,1).value = year_month
            sheet1 = workbook.worksheets[1]
            sheet1.title = "%d月" % int(month)+"403"
            sheet1.cell(4,1).value = year_month
            sheet2 = workbook.worksheets[2]
            sheet2.title = "%d月" % int(month)+"404"
            sheet2.cell(4,1).value = year_month
            sheet3 = workbook.worksheets[3]
            sheet3.title = "%d月" % int(month)+"405"
            sheet3.cell(4,1).value = year_month
            sheet4 = workbook.worksheets[4]
            sheet4.title = "%d月" % int(month)+"406"
            sheet4.cell(4,1).value = year_month
            sheet5 = workbook.worksheets[5]
            sheet5.title = "%d月" % int(month)+"407"
            sheet5.cell(4,1).value = year_month
            sheet6 = workbook.worksheets[6]
            sheet6.title = "%d月" % int(month)+"408"
            sheet6.cell(4,1).value = year_month
            sheet7 = workbook.worksheets[7]
            sheet7.title = "%d月" % int(month)+"409"
            sheet7.cell(4,1).value = year_month

            workbook.save(filename)  # 儲存檔案
    return filename

def export_data_to_excel():
    global excel_filename
    global sheetname
    length = len(csv_data)
    """
    print('共有資料 :', length, '筆')
    if length > 0:
        for _ in csv_data:
            print(_)
    """

    excel_filename = check_excel_filename(csv_data[1][0], csv_data[0][2])
    print('取得excel檔名 :', excel_filename)

    length = len(csv_data)
    print('資料長度 : ', length)
    if length == 0:
        message = '無資料, 離開'
        print(message)
        text1.insert('end', message)
        main_message2.set(message)
    else:
        message = '匯出資料 完成'
        print(message)
        text1.insert('end', message)

        export_data_to_excel0(csv_data)


def button00Click():
    print('你按了 匯入生產資料')
    import_csv_data()

def button01Click():
    print('你按了 匯出至Excel檔案')
    export_data_to_excel()

def button02Click():
    print('你按了 info')
    show_info()
    show_csv_data()
   
def button03Click():
    print('你按了button03')
    window.destroy()

def button04Click():
    print('你按了xxx')

def button05Click():
    print('你按了button05')
    clear_text1()

def clear_text1():
    text1.delete(1.0, 'end')
    # 執行 clear 函式時，清空內容


filename = "402除菌區數據紀錄_211214_000000.csv"
            
year_month_day_data = filename[-17:-11]

print(year_month_day_data)

year = filename[-17:-15]
month = filename[-15:-13]
day = filename[-13:-11]

print(int(year)+2000)
print(month)
print(day)



window = tk.Tk()

main_message1 = tk.StringVar()
main_message2 = tk.StringVar()

# 設定主視窗大小
W = 800
H = 800
x_st = 100
y_st = 100
#size = str(W)+'x'+str(H)
#size = str(W)+'x'+str(H)+'+'+str(x_st)+'+'+str(y_st)
#window.geometry(size)
window.geometry("{0:d}x{1:d}+{2:d}+{3:d}".format(W, H, x_st, y_st))
#print('{0:d}x{1:d}+{2:d}+{3:d}'.format(W, H, x_st, y_st))

# 設定主視窗標題
window.title('環境溫 / 濕度管制紀錄表 資料轉換')

# 設定主視窗之背景色
#window.configure(bg = "#7AFEC6")

icon_filename = 'C:/_git/vcs/_1.data/______test_files1/_material/ims.ico'
window.iconbitmap(icon_filename)   # 更改圖示

x_st = 50
y_st = 50
dx = 120
dy = 80
w = 12
h = 3

button00 = tk.Button(window, width = w, height = h, command = button00Click, text = '匯入生產資料')
button00.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button01 = tk.Button(window, width = w, height = h, command = button01Click, text = '匯出至Excel檔案')
button01.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button02 = tk.Button(window, width = w, height = h, command = button02Click, text = 'info')
button02.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button03 = tk.Button(window, width = w, height = h, command = button03Click, text = '關閉視窗')
button03.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button04 = tk.Button(window, width = w, height = h, command = button04Click, text = '')
button04.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button05 = tk.Button(window, width = w, height = h, command = button05Click, text = 'Clear')
button05.pack(side = tk.LEFT, ipadx = 25, ipady = 25, expand = tk.YES)
button00.place(x = x_st + dx * 0, y = y_st + dy * 0)
button01.place(x = x_st + dx * 1, y = y_st + dy * 0)
button02.place(x = x_st + dx * 2, y = y_st + dy * 0)
button03.place(x = x_st + dx * 3, y = y_st + dy * 0)
button04.place(x = x_st + dx * 4, y = y_st + dy * 0)
button05.place(x = x_st + dx * 5, y = y_st + dy * 0)

font_size = 20

# 加入 Label
label_message1 = tk.Label(window, font=("標楷體", font_size), fg = 'red', textvariable = main_message1)
label_message1.pack()
label_message1.place(x = 5, y = 0 + 10)
main_message1.set('')

label_message2 = tk.Label(window, font=("標楷體", font_size), fg = 'red', textvariable = main_message2)
label_message2.pack()
label_message2.place(x = 5 + W * 3 / 4, y = 0 + 10)
main_message2.set('')

msg = tk.StringVar()
label1 = tk.Label(window, text = '')
label1.pack()
label1.place(x = x_st + dx * 0, y = y_st + dy * 2 - 20)
label2 = tk.Label(window, fg = 'red', textvariable = msg)
label2.pack()
label2.place(x = x_st + dx * 0, y = y_st + dy * 2 + 80)

# 加入 Text
text1 = tk.Text(window, width = 100, height = 30)  # 放入多行輸入框
text1.pack()
text1.place(x = x_st + dx * 0, y = y_st + dy * 3 + 50)

message = "匯入生產資料"
main_message1.set(message)

window.mainloop()


"""

    print('新建資料庫, 清除記憶體資料')
    clear_csv_data()

    button01_text.set("開啟資料庫")
    print('開啟資料庫, 清除記憶體資料')
    clear_csv_data()



old 或許沒有用了

def precheck_csv_data():
    print('預先檢查這些csv檔案是否皆可用')
    
    #1. 無舊資料
    #2. 有舊資料 要考慮序號是否
    #3. 第一站資料必定要先存在
    
    source_dir = 'QC/csv'
    target_dir = 'QC/csv_old'

    if flag_debug_mode == True:
        source_dir = 'QC_debug/csv'
        target_dir = 'QC_debug/csv_old'

    #準備輸出資料夾 若不存在, 則建立
    if not os.path.exists(target_dir):
            os.mkdir(target_dir)
            #os.makedirs(target_dir, exist_ok = True)

    #尋找檔案
    import glob
    print('111尋找目前目錄下之 *.csv')
    files = glob.glob(source_dir + "/*.csv") 
    for filename in files:
        print(filename)
        stage = 0
        tablename = 'table00'
        #process_csv_file1(filename)
        if stage != 0:
            print('繼續')
            #process_csv_file2(filename)
            #若能正確處理完畢, 再搬到old資料夾
            #shutil.move(filename, target_dir)

    #message = "匯入生產資料 完成"
    #print(message)
    #text1.insert('end', message)

def export_data():
    global csv_data
    length = len(csv_data)
    print('資料長度 : ', length)
    if length == 0:
        message = '無資料, 離開'
        print(message)
        text1.insert('end', message)
        main_message2.set(message)
        return
    else:
        # 開啟輸出的 csv 檔案
        filename_w = '匯出資料範例.csv'
        with open(filename_w, 'w', newline = '') as csvfile:
            # 建立 csv 檔寫入物件
            writer = csv.writer(csvfile)

            # 寫入二維串列資料
            writer.writerows(csv_data)

        message = '匯出資料 完成'
        #print(message)
        #text1.insert('end', message)



"""




