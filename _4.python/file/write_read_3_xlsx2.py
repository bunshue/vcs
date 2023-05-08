#各種檔案寫讀範例 xlsx 2

import openpyxl

print("寫讀 xlsx")

print("建立 xlsx")

filename = 'C:/_git/vcs/_1.data/______test_files2/test2.xlsx'

workbook=openpyxl.Workbook()   #建立一個工作簿
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
# 以儲存格位置寫入資料
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'
# 以串列寫入資料
listtitle=["姓名","電話"]
sheet.append(listtitle)  
listdata=["David","0999-1234567"]
sheet.append(listdata)  

# 儲存檔案    
workbook.save(filename)    

print("建立 xlsx OK, 檔案 : " + filename)

 
#  讀取檔案
workbook = openpyxl.load_workbook(filename)
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
print("顯示資料");
# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)

# 取得總行、列數
print(sheet.max_row, sheet.max_column)

# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()

sheet['A3'] = 'David'

filename2 = 'C:/_git/vcs/_1.data/______test_files2/test2b.xlsx'

workbook.save(filename2)









