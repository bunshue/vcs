#各種檔案寫讀範例 xlsx 1

import openpyxl

print('------------------------------------------------------------')	#60個
print('excel讀寫測試 1')

print("寫讀 xlsx")

print("建立 xlsx")

filename = 'C:/_git/vcs/_1.data/______test_files2/test1.xlsx'

workbook=openpyxl.Workbook()    # 建立一個工作簿
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
# 以儲存格位置寫入資料
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'
# 以串列寫入資料
listtitle=["姓名","電話"]
sheet.append(listtitle)  
listdata=["David","0999-1234567"]
sheet.append(listdata)

# 儲存檔案
workbook.save(filename)

print("建立 xlsx OK, 檔案 : " + filename)

print('------------------------------------------------------------')	#60個
print('excel讀寫測試 2')

filename = 'C:/_git/vcs/_1.data/______test_files1/__RW/_excel/python_ReadWrite_EXCEL3.xlsx'

print("讀取 xlsx, 檔案 : " + filename)

#  讀取檔案
workbook = openpyxl.load_workbook(filename)
# 取得第 1 個工作表
sheet = workbook.worksheets[0]
print("顯示資料");
# 取得指定儲存格
print(sheet['A1'], sheet['A1'].value)
# 取得總行、列數
print(sheet.max_row, sheet.max_column)
# 顯示 cell資料
for i in range(1, sheet.max_row+1):
    for j in range(1, sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value,end="   ")
    print()
sheet['A3'] = 'David' 

print("另存新檔");
filename2 = 'C:/_git/vcs/_1.data/______test_files2/test2b.xlsx'
workbook.save(filename2)
print("另存新檔 OK, 檔案 : " + filename2)

print('------------------------------------------------------------')	#60個
print('excel讀寫測試 3')

#各種檔案寫讀範例 xlsx 2

import openpyxl
filename = 'C:/_git/vcs/_1.data/______test_files1/__RW/_excel/python_ReadWrite_EXCEL5_population2.xlsx'

print("讀取 xlsx, 檔案 : " + filename)

workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

list_values = list(sheet.values)
print('所有資料')
print(list_values)
print()

print('標題欄')
cols = list_values[0]
print(cols)

cnt = 0
for value_tuple in list_values[1:]:
    print('第', cnt, '筆資料 :', value_tuple)
    #print(type(value_tuple))
    cnt += 1


print('------------------------------------------------------------')	#60個
print('完成')

